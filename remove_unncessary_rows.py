import os
import openpyxl
from openpyxl import load_workbook

# ─── PATHS ────────────────────────────────────────────────────────────────────
MASTER_FILE   = r"c:\Users\Tanush.Bidkar\Downloads\Final_Rate_Linked_ICICI 2 (1).xlsx"
TRAINING_DIR  = r"C:\Users\Tanush.Bidkar\Downloads\Zone Wise Grouped ICICI\East"

# ─── COLUMN INDEX → KEYWORD TO MATCH IN Summary Sheet (column B, "Particulars") ─
# Master Summary data columns start at index 3 (0-based), i.e. Excel col D
COL_INDEX_TO_KEYWORD = {
    3:  "CIVIL",
    4:  "POP",
    5:  "CARPENTRY",
    6:  "PAINTING",
    7:  "ROLLING",
    8:  "ELECTRIF",
    9:  "ADDITIONAL",
    10: "TOTAL",
}

YELLOW_FILL = "FFFF00"  # Standard yellow

def is_yellow(cell):
    """Return True if cell background is yellow."""
    fill = cell.fill
    if fill and fill.fgColor:
        color = fill.fgColor
        # Handle theme colors vs rgb
        if color.type == "rgb":
            rgb = color.rgb  # e.g. 'FFFFFF00' or 'FFFF00'
            # Strip alpha channel if present (8 chars) → last 6 chars
            hex_color = rgb[-6:].upper()
            return hex_color == YELLOW_FILL
    return False

def get_yellow_column_indices(master_ws, row_num, data_start_col=4, data_end_col=11):
    """
    For a given row in Master Summary, return list of 0-based column indices
    (matching COL_INDEX_TO_KEYWORD keys) where the cell is yellow.
    data_start_col=4 = Excel col D (1-based), data_end_col=11 = Excel col K
    """
    yellow_indices = []
    for excel_col in range(data_start_col, data_end_col + 1):
        cell = master_ws.cell(row=row_num, column=excel_col)
        if is_yellow(cell):
            zero_based_idx = excel_col - 1  # Convert to 0-based
            if zero_based_idx in COL_INDEX_TO_KEYWORD:
                yellow_indices.append(zero_based_idx)
    return yellow_indices

def find_training_file(file_branch_name, training_dir):
    """
    Match the File/Branch Name from Master Summary to an actual .xlsx file
    in the training directory (top-level only, no subfolders).
    Uses fuzzy prefix matching on filename stem.
    """
    target = file_branch_name.strip().lower().replace(" ", "_")
    
    for fname in os.listdir(training_dir):
        if not fname.endswith(('.xlsx', '.xls')):
            continue
        # Skip files inside East/West/North/South subfolders (we only scan top level)
        stem = os.path.splitext(fname)[0].lower().replace(" ", "_")
        # Check if target is substring of stem or stem starts with target
        if target in stem or stem.startswith(target[:15]):
            return os.path.join(training_dir, fname)
    
    return None

def delete_rows_in_summary_sheet(training_file_path, keywords_to_remove):
    """
    Open the training file, go to 'Summary Sheet', find rows where
    column B (Particulars) contains any of the keywords, delete those rows.
    Saves the file in place.
    """
    try:
        wb = load_workbook(training_file_path)
    except Exception as e:
        print(f"    ERROR opening {training_file_path}: {e}")
        return

    if 'Summary Sheet' not in wb.sheetnames:
        print(f"    WARNING: 'Summary Sheet' not found in {os.path.basename(training_file_path)}")
        wb.close()
        return

    ws = wb['Summary Sheet']
    
    # Collect rows to delete (iterate in reverse to not mess up row indices)
    rows_to_delete = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.column == 2:  # Column B = Particulars
                cell_val = str(cell.value).upper() if cell.value else ""
                for keyword in keywords_to_remove:
                    if keyword.upper() in cell_val:
                        rows_to_delete.append(cell.row)
                        print(f"    → Marking row {cell.row} for deletion: '{cell.value}' (matched keyword: {keyword})")
                        break
    
    if not rows_to_delete:
        print(f"    No matching rows found (already removed or not present) — skipping save.")
        wb.close()
        return
    
    # Delete in reverse order
    for row_num in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row_num)
        print(f"    ✓ Deleted row {row_num}")
    
    wb.save(training_file_path)
    print(f"    ✓ Saved: {os.path.basename(training_file_path)}")
    wb.close()

def main():
    print("=" * 70)
    print("Loading Master Summary from:")
    print(f"  {MASTER_FILE}")
    print("=" * 70)

    try:
        master_wb = load_workbook(MASTER_FILE, data_only=True)
    except Exception as e:
        print(f"ERROR: Cannot open master file: {e}")
        return

    if 'Master Summary' not in master_wb.sheetnames:
        print("ERROR: 'Master Summary' sheet not found in master file!")
        return

    master_ws = master_wb['Master Summary']

    # Get all training files at top level (exclude subfolders)
    training_files = [
        f for f in os.listdir(TRAINING_DIR)
        if f.endswith(('.xlsx', '.xls'))
        and os.path.isfile(os.path.join(TRAINING_DIR, f))
    ]
    print(f"\nFound {len(training_files)} training files at top level of training folder.\n")

    # Iterate Master Summary rows (skip header row 1)
    processed = 0
    for row_num in range(2, master_ws.max_row + 1):
        # Column B = File / Branch Name (Excel col 2)
        file_branch_cell = master_ws.cell(row=row_num, column=2)
        file_branch_name = file_branch_cell.value
        
        if not file_branch_name or str(file_branch_name).strip() == "":
            continue

        file_branch_name = str(file_branch_name).strip()
        
        # Find yellow columns in this row (Excel cols D to K = indices 4 to 11)
        yellow_col_indices = get_yellow_column_indices(master_ws, row_num, data_start_col=4, data_end_col=11)
        
        if not yellow_col_indices:
            continue  # No yellow highlights → skip this file
        
        # Map yellow column indices to keywords
        keywords_to_remove = [COL_INDEX_TO_KEYWORD[idx] for idx in yellow_col_indices if idx in COL_INDEX_TO_KEYWORD]
        
        print(f"\nRow {row_num}: {file_branch_name}")
        print(f"  Yellow columns detected → keywords to remove: {keywords_to_remove}")
        
        # Find the matching training file
        training_file_path = find_training_file(file_branch_name, TRAINING_DIR)
        
        if not training_file_path:
            print(f"  WARNING: No matching training file found for '{file_branch_name}' — skipping.")
            continue
        
        print(f"  Matched training file: {os.path.basename(training_file_path)}")
        
        # Delete the rows in Summary Sheet
        delete_rows_in_summary_sheet(training_file_path, keywords_to_remove)
        processed += 1

    master_wb.close()
    print("\n" + "=" * 70)
    print(f"DONE. Processed {processed} files with yellow highlights.")
    print("=" * 70)

if __name__ == "__main__":
    main()