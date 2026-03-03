import os
import pandas as pd
from flask import Flask, request, render_template, send_from_directory, flash, redirect, url_for, jsonify, session
from werkzeug.utils import secure_filename
import pickle
import numpy as np
import re
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from collections import defaultdict
import io
import base64
import html # For escaping HTML characters in data attributes
import re

# --- Matplotlib Setup (for backend graph generation) ---
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
# Clear memory after each plot
import gc
from scipy.stats import norm

def _default_dict_factory():
    return list()
import __main__
__main__._default_dict_factory = _default_dict_factory

# --- CWIValidatorTester Class ---
class CWIValidatorTester:
    def __init__(self):
        self.sqfeet_ranges = {}
        self.vectorizers = {}
        self.particular_vectors = {}
        self.particular_names = {}
        self.training_stats = {}

    def extract_sqfeet_from_filename(self, filename, mode='quantity'):
        """Extract square feet range from filename based on mode"""
        filename_lower = filename.lower()

        # Check for explicit range patterns first (work for both modes)
        # 250 sqft ranges (amount mode)
        amount_patterns = [
            (r'500\s*-?\s*1000', '500-1000'),
            (r'1001\s*-?\s*1500', '1001-1500'),
            (r'1501\s*-?\s*2000', '1501-2000'),
            (r'2001\s*-?\s*2500', '2001-2500'),
            (r'2501\s*-?\s*3000', '2501-3000'),
            (r'3001\s*-?\s*3500', '3001-3500'),
            (r'3501\s*-?\s*4000', '3501-4000'),
            (r'4001\s*-?\s*4500', '4001-4500'),
            (r'4501\s*-?\s*5000', '4501-5000')
        ]
        # 500 sqft ranges (quantity mode)
        quantity_patterns = [
            (r'500\s*-?\s*1000', '500-1000'),
            (r'1001\s*-?\s*1500', '1001-1500'),
            (r'1501\s*-?\s*2000', '1501-2000'),
            (r'2001\s*-?\s*2500', '2001-2500'),
            (r'2501\s*-?\s*3000', '2501-3000'),
            (r'3001\s*-?\s*3500', '3001-3500'),
            (r'3501\s*-?\s*4000', '3501-4000'),
            (r'4001\s*-?\s*4500', '4001-4500'),
            (r'4501\s*-?\s*5000', '4501-5000')
        ]

        # Try explicit patterns first
        for pattern, range_key in amount_patterns:
            if re.search(pattern, filename_lower):
                return range_key

        for pattern, range_key in quantity_patterns:
            if re.search(pattern, filename_lower):
                return range_key

        # Check for 5000+ pattern
        if re.search(r'5000\+|above\s*5000|more\s*than\s*5000', filename_lower):
            return '5000+'

        # Fallback: Extract number and determine range based on mode
        numbers = re.findall(r'\d+', filename_lower)
        if numbers:
            for num_str in numbers:
                try:
                    num = int(num_str)

                    if mode == 'amount':
                        # Use 500 sqft ranges for AMOUNT mode
                        if 500 <= num <= 1000:
                            return '500-1000'
                        elif 1001 <= num <= 1500:
                            return '1001-1500'
                        elif 1501 <= num <= 2000:
                            return '1501-2000'
                        elif 2001 <= num <= 2500:
                            return '2001-2500'
                        elif 2501 <= num <= 3000:
                            return '2501-3000'
                        elif 3001 <= num <= 3500:
                            return '3001-3500'
                        elif 3501 <= num <= 4000:
                            return '3501-4000'
                        elif 4001 <= num <= 4500:
                            return '4001-4500'
                        elif 4501 <= num <= 5000:
                            return '4501-5000'
                        elif num > 5000:
                            return '5000+'
                    elif mode == 'rate':
                        # Use 250 sqft ranges for RATE mode
                        if 500 <= num <= 750:
                            return '500-750'
                        elif 751 <= num <= 1000:
                            return '751-1000'
                        elif 1001 <= num <= 1250:
                            return '1001-1250'
                        elif 1251 <= num <= 1500:
                            return '1251-1500'
                        elif 1501 <= num <= 1750:
                            return '1501-1750'
                        elif 1751 <= num <= 2000:
                            return '1751-2000'
                        elif 2001 <= num <= 2250:
                            return '2001-2250'
                        elif 2251 <= num <= 2500:
                            return '2251-2500'
                        elif 2501 <= num <= 2750:
                            return '2501-2750'
                        elif 2751 <= num <= 3000:
                            return '2751-3000'
                        elif 3001 <= num <= 3250:
                            return '3001-3250'
                        elif 3251 <= num <= 3500:
                            return '3251-3500'
                        elif 3501 <= num <= 3750:
                            return '3501-3750'
                        elif 3751 <= num <= 4000:
                            return '3751-4000'
                        elif 4001 <= num <= 4250:
                            return '4001-4250'
                        elif 4251 <= num <= 4500:
                            return '4251-4500'
                        elif 4501 <= num <= 4750:
                            return '4501-4750'
                        elif 4751 <= num <= 5000:
                            return '4751-5000'
                        elif num > 5000:
                            return '5000+'
                    else:
                        # Use 500 sqft ranges (default for quantity mode)
                        if 500 <= num <= 1000:
                            return '500-1000'
                        elif 1001 <= num <= 1500:
                            return '1001-1500'
                        elif 1501 <= num <= 2000:
                            return '1501-2000'
                        elif 2001 <= num <= 2500:
                            return '2001-2500'
                        elif 2501 <= num <= 3000:
                            return '2501-3000'
                        elif 3001 <= num <= 3500:
                            return '3001-3500'
                        elif 3501 <= num <= 4000:
                            return '3501-4000'
                        elif 4001 <= num <= 4500:
                            return '4001-4500'
                        elif 4501 <= num <= 5000:
                            return '4501-5000'
                        elif num > 5000:
                            return '5000+'
                except ValueError:
                    continue

        return None
    def extract_sqfeet_from_filename_regional(self, filename):
        """Extract square feet range for REGIONAL models (500-2500, 2500-5000, 5000+)"""
        filename_lower = filename.lower()

        # Regional patterns - LARGER intervals
        patterns = [
            (r'500\s*-?\s*2500', '500-2500'),
            (r'2500\s*-?\s*5000', '2500-5000'),
            (r'5000\+|above\s*5000|more\s*than\s*5000', '5000+')
        ]

        for pattern, range_key in patterns:
            if re.search(pattern, filename_lower):
                return range_key

        # Fallback: Extract number and determine regional range
        numbers = re.findall(r'\d+', filename_lower)
        if numbers:
            for num_str in numbers:
                try:
                    num = int(num_str)
                    if 500 <= num <= 2500:
                        return '500-2500'
                    elif 2501 <= num <= 5000:
                        return '2500-5000'
                    elif num > 5000:
                        return '5000+'
                except ValueError:
                    continue

        return None
    def extract_sqfeet_from_filename_reliance(self, filename):
        """Extract square feet range for RELIANCE models based on specific numeric ranges"""
        filename_lower = filename.lower()

        # Extract all numeric values from the filename
        # This ensures we find '10191' even if it's not at the very start
        numbers = re.findall(r'\d+', filename_lower)

        if numbers:
            for n in numbers:
                try:
                    num = int(n)

                    # Filter out small numbers that aren't square footage (like dates or indices)
                    if num < 1000:
                        continue

                    # Apply your specific 11 ranges
                    if 1000 <= num <= 3999:
                        return '1000-3999'
                    elif 4000 <= num <= 4999:
                        return '4000-4999'
                    elif 5000 <= num <= 5999:
                        return '5000-5999'
                    elif 6000 <= num <= 6999:
                        return '6000-6999'
                    elif 7000 <= num <= 7999:
                        return '7000-7999'
                    elif 8000 <= num <= 8999:
                        return '8000-8999'
                    elif 9000 <= num <= 10500:
                        return '9000-10500'
                    elif 10501 <= num <= 12999:
                        return '10501-12999'
                    elif 13000 <= num <= 14999:
                        return '13000-14999'
                    elif 15000 <= num <= 19999:
                        return '15000-19999'
                    elif num >= 20000:
                        return '20000+'
                except ValueError:
                    continue

        # Fallback: If filenames contain explicit range strings (Optional)
        patterns = [
            (r'1000\s*-?\s*3999', '1000-3999'),
            (r'4000\s*-?\s*4999', '4000-4999'),
            (r'5000\s*-?\s*5999', '5000-5999'),
            (r'6000\s*-?\s*6999', '6000-6999'),
            (r'7000\s*-?\s*7999', '7000-7999'),
            (r'8000\s*-?\s*8999', '8000-8999'),
            (r'9000\s*-?\s*10500', '9000-10500'),
            (r'10501\s*-?\s*12999', '10501-12999'),
            (r'13000\s*-?\s*14999', '13000-14999'),
            (r'15000\s*-?\s*19999', '15000-19999'),
            (r'20000\+|above\s*20000', '20000+')
        ]

        for pattern, range_key in patterns:
            if re.search(pattern, filename_lower):
                return range_key

        return None
    def extract_sqfeet_from_filename_smart(self, filename):
        """Extract square feet range for SMART models (specific ranges from training)"""
        filename_lower = filename.lower()

        # Extract all numeric values from the filename
        numbers = re.findall(r'\d+', filename_lower)

        if numbers:
            for n in numbers:
                try:
                    num = int(n)

                    # Filter out small numbers that aren't square footage
                    if num < 1000:
                        continue

                    # Apply Smart-specific ranges
                    if 1000 <= num <= 3999:
                        return '1000-3999'
                    elif 4000 <= num <= 4999:
                        return '4000-4999'
                    elif 5000 <= num <= 5999:
                        return '5000-5999'
                    elif 6000 <= num <= 6999:
                        return '6000-6999'
                    elif 7000 <= num <= 7999:
                        return '7000-7999'
                    elif 8000 <= num <= 8999:
                        return '8000-8999'
                    elif 9000 <= num <= 10500:
                        return '9000-10500'
                    elif 10501 <= num <= 12999:
                        return '10501-12999'
                    elif 13000 <= num <= 13999:
                        return '13000-13999'
                    elif 14000 <= num <= 15999:
                        return '14000-15999'
                    elif 16000 <= num <= 17999:
                        return '16000-17999'
                    elif 18000 <= num <= 19999:
                        return '18000-19999'
                    elif 20000 <= num <= 21999:
                        return '20000-21999'
                    elif 22000 <= num <= 23999:
                        return '22000-23999'
                    elif 24000 <= num <= 25999:
                        return '24000-25999'
                    elif 26000 <= num <= 27999:
                        return '26000-27999'
                    elif 28000 <= num <= 29999:
                        return '28000-29999'
                    elif num >= 30000:
                        return '30000+'
                except ValueError:
                    continue

        return None
    def clean_text(self, text):
        """Clean and normalize text for better matching"""
        if pd.isna(text):
            return ""
        text = str(text).lower()
        text = re.sub(r'[^\w\s:/.-]', ' ', text)
        text = ' '.join(text.split())
        return text

    def extract_quantity(self, qty_value):
        """Extract numeric quantity from various formats"""
        if pd.isna(qty_value):
            print(f"      ⚠️  Value is NaN/None")
            return 0

        qty_str = str(qty_value).strip()
        print(f"      📝 Raw string: '{qty_str}'")

        # Remove common currency symbols and formatting
        qty_str = qty_str.replace('₹', '').replace('Rs', '').replace(',', '').replace('INR', '').strip()

        # Handle negative values
        qty_str = qty_str.replace('(', '-').replace(')', '')

        print(f"      🧹 Cleaned string: '{qty_str}'")

        numbers = re.findall(r'-?\d+\.?\d*', qty_str)
        print(f"      🔢 Found numbers: {numbers}")

        if numbers:
            result = float(numbers[0])
            print(f"      ✅ Extracted: {result}")
            return result

        print(f"      ❌ No numbers found, returning 0")
        return 0

    def extract_amount(self, amount_value):
        """Extract numeric amount from various formats"""
        if pd.isna(amount_value):
            print(f"      ⚠️  Amount value is NaN/None")
            return 0

        amount_str = str(amount_value).strip()
        print(f"      📝 Raw amount string: '{amount_str}'")

        # Remove currency symbols and commas
        amount_str = amount_str.replace('₹', '').replace('Rs', '').replace(',', '').replace('INR', '').strip()

        # Handle negative values
        amount_str = amount_str.replace('(', '-').replace(')', '')

        print(f"      🧹 Cleaned amount string: '{amount_str}'")

        numbers = re.findall(r'-?\d+\.?\d*', amount_str)
        print(f"      🔢 Found amount numbers: {numbers}")

        if numbers:
            result = float(numbers[0])
            print(f"      ✅ Extracted amount: {result}")
            return result

        print(f"      ❌ No numbers found in amount, returning 0")
        return 0

    def normalize_unit(self, unit_value):
        """Normalize unit values to check if it's number-based"""
        if pd.isna(unit_value):
            return None
        unit_str = str(unit_value).strip().upper()
        # Number-based units
        number_units = ['NOS', 'NO', 'NOS.', 'NO.', 'NUMBER', 'NUMBERS', 'QTY', 'QUANTITY', 'LSUM', 'L.SUM', 'LUMPSUM', 'LUMP SUM', 'LS', 'L/S','L SUM']
        for nu in number_units:
            if nu in unit_str:
                return 'NUMBER_BASED'
        return unit_str

    def is_number_based_unit(self, unit_value):
        """Check if unit is number-based"""
        normalized = self.normalize_unit(unit_value)
        return normalized == 'NUMBER_BASED'

    def round_quantity_if_number_based(self, quantity, unit_value):
        """Round quantity if unit is number-based"""
        if self.is_number_based_unit(unit_value):
            return round(quantity)
        return quantity

    def infer_unit_from_particular(self, particular_text):
        """Use AI logic to infer unit from particular description"""
        particular_lower = str(particular_text).lower()

        # Keywords indicating number-based quantities
        number_keywords = ['panel', 'door', 'window', 'unit', 'piece', 'item', 'set', 
                           'fixture', 'board', 'sheet', 'nos', 'number', 'quantity']

        # Keywords indicating area-based quantities
        area_keywords = ['sqft', 'sq ft', 'square feet', 'sft', 'area', 'flooring', 
                         'painting', 'plastering', 'tiling']

        # Keywords indicating length-based quantities
        length_keywords = ['rft', 'running feet', 'linear', 'length', 'pipe', 'wire', 
                           'cable', 'railing']

        # Check for number-based
        for keyword in number_keywords:
            if keyword in particular_lower:
                return 'NOS'

        # Check for area-based
        for keyword in area_keywords:
            if keyword in particular_lower:
                return 'SFT'

        # Check for length-based
        for keyword in length_keywords:
            if keyword in particular_lower:
                return 'RFT'

        # Default to NOS if uncertain
        return 'NOS'

    def load_model(self, model_path):
        """Load trained model from file with proper key normalization"""
        try:
            with open(model_path, 'rb') as f:
                model_data = pickle.load(f)

            # --- 1. Load Primary sqfeet_ranges ---
            if 'sqfeet_ranges' in model_data:
                self.sqfeet_ranges = model_data['sqfeet_ranges']
            elif 'rate_sqfeet_ranges' in model_data:
                self.sqfeet_ranges = model_data['rate_sqfeet_ranges']
            elif 'amount_cwi_ranges' in model_data:
                self.sqfeet_ranges = model_data['amount_cwi_ranges']
            else:
                self.sqfeet_ranges = {}

            # 🔥 CRITICAL FIX: Normalize all range data keys
            for range_name, range_data in self.sqfeet_ranges.items():
                # Count which keys exist
                has_quantity = 'quantity_ranges' in range_data and range_data['quantity_ranges']
                has_rate = 'rate_ranges' in range_data and range_data['rate_ranges']
                has_amount = 'amount_ranges' in range_data and range_data['amount_ranges']

                # Determine which key has the actual data
                if has_quantity:
                    source_key = 'quantity_ranges'
                    source_data = range_data['quantity_ranges']
                elif has_rate:
                    source_key = 'rate_ranges'
                    source_data = range_data['rate_ranges']
                elif has_amount:
                    source_key = 'amount_ranges'
                    source_data = range_data['amount_ranges']
                else:
                    # No data in this range
                    continue

                # 🔥 COPY SOURCE DATA TO ALL THREE KEYS
                range_data['quantity_ranges'] = source_data
                range_data['rate_ranges'] = source_data
                range_data['amount_ranges'] = source_data

                print(f"  ✓ Normalized range {range_name}: copied '{source_key}' to all keys ({len(source_data)} items)")

            # --- 2. Load Text Similarity Assets ---
            self.vectorizers = model_data.get('vectorizers', {})
            self.particular_vectors = model_data.get('particular_vectors', {})
            self.particular_names = model_data.get('particular_names', {})

            # --- 3. Load Other Data Structures ---
            self.training_stats = model_data.get('training_stats', {})
            self.rate_sqfeet_ranges = model_data.get('rate_sqfeet_ranges', {})
            self.amount_cwi_ranges = model_data.get('amount_cwi_ranges', defaultdict(lambda: {
                'grouped_ranges': {}, 'particulars_data': defaultdict(list)
            }))
            self.rate_per_sqft_250_ranges = model_data.get('rate_per_sqft_250_ranges', {})
            self.rate_per_sqft_500_ranges = model_data.get('rate_per_sqft_500_ranges', defaultdict(lambda: {
                'grouped_ranges': {}, 'subcategory_data': defaultdict(list)
            }))
            self.amount_per_cwi_250_ranges = model_data.get('amount_per_cwi_250_ranges', defaultdict(lambda: {
                'grouped_ranges': {}, 'subcategory_data': defaultdict(list)
            }))
            self.known_groups = model_data.get('known_groups', {})
            self.fallback_group = model_data.get('fallback_group', 'Others')
            self.all_subcategories = model_data.get('subcategory_map', {})

            print(f"✓ Model loaded successfully: {model_path}")
            print(f"  - Ranges loaded: {list(self.sqfeet_ranges.keys())}")
            print(f"  - Similarity engines active: {len(self.vectorizers) > 0}")
            print(f"  - All keys normalized to: quantity_ranges, rate_ranges, amount_ranges")

            return True

        except Exception as e:
            print(f"❌ Error loading model: {str(e)}")
            import traceback
            traceback.print_exc()
            return False

    def normalize_range_key_for_reliance(self, range_key, model_key):
        """Convert between K format and numeric format for Reliance models"""
        if model_key and 'reliance' in model_key:
            # If using K format but vectors use numeric format
            if 'K' in range_key:
                # Convert '8K-12K' to '8000-12000'
                parts = range_key.replace('K', '000').replace('+', '').split('-')
                if len(parts) == 2:
                    return f"{parts[0]}-{parts[1]}"
                elif '+' in range_key:
                    return f"{parts[0]}+"
        return range_key
    def find_similar_particular(self, query_particular, sqfeet_range, model_key=None, threshold=0.5):
        # --- NEW FIX FOR RELIANCE KEY MISMATCH ---
        # Convert '8K-12K' to '8000-12000' for vector lookup if needed
        lookup_range = sqfeet_range
        if model_key and 'reliance' in model_key and 'K' in sqfeet_range:
            lookup_range = self.normalize_range_key_for_reliance(sqfeet_range, model_key)
            print(f"🔄 Reliance Lookup: Converted '{sqfeet_range}' to '{lookup_range}' for vectorizer")

        # Use lookup_range for dictionary access
        if lookup_range not in self.vectorizers:
            print(f"⚠️ Range '{lookup_range}' not found in vectorizers. Available: {list(self.vectorizers.keys())}")
            return None, 0
        
        vectorizer = self.vectorizers[lookup_range]
        particular_vectors = self.particular_vectors[lookup_range]
        particular_names = self.particular_names[lookup_range]
        # ... rest of your code ...
        
        query_vector = vectorizer.transform([self.clean_text(query_particular)])
        similarities = cosine_similarity(query_vector, particular_vectors)[0]
        best_match_idx = np.argmax(similarities)
        
        if similarities[best_match_idx] >= threshold:
            return particular_names[best_match_idx], similarities[best_match_idx]
        
        return None, similarities[best_match_idx]

    def validate_quantity(self, particular, quantity, sqfeet_range, is_regional=False, selected_region=None, mode='quantity', model_key=None):
        """Validate quantity for a particular item within specific sq feet range"""
        # ✅ ADD DEBUG OUTPUT AT THE VERY START
        print(f"\n🔍 VALIDATE_QUANTITY DEBUG:")
        print(f"   Model Key: {model_key}")
        print(f"   Requested Range: {sqfeet_range}")
        print(f"   Available Ranges in Model: {list(self.sqfeet_ranges.keys())}")
        print(f"   Particular (first 50 chars): {str(particular)[:50]}")

        # Check if range exists
        if sqfeet_range not in self.sqfeet_ranges:
            print(f"   ❌ RANGE MISMATCH! '{sqfeet_range}' not in model")
            print(f"   This suggests you're using the wrong model for this file")
            return True, f"⚠️ Range {sqfeet_range} not found in model", None, 0, None, None, None, None, [], []

        print(f"   ✅ Range found in model")

        # ✅ NEW: Special handling for Reliance models
        if model_key and 'reliance' in model_key:
            print(f"\n🏪 RELIANCE VALIDATION MODE")
            print(f"   Range: {sqfeet_range}")
            print(f"   Service Code (passed as 'particular'): {particular}")
            print(f"   Quantity: {quantity}")

            # For Reliance, 'particular' parameter actually contains the Service Code
            service_code = str(particular).strip()

            # Check if range exists
            if sqfeet_range not in self.sqfeet_ranges:
                print(f"   ❌ Range {sqfeet_range} not found")
                return True, f"⚠️ Range {sqfeet_range} not found in model", None, 0, None, None, None, None, [], []

            range_data = self.sqfeet_ranges[sqfeet_range]

            # Reliance uses quantity_ranges key
            if 'quantity_ranges' not in range_data or not range_data['quantity_ranges']:
                print(f"   ❌ No quantity_ranges data for this range")
                return True, "⚠️ No validation data available", None, 0, None, None, None, None, [], []

            # EXACT MATCH CHECK
            if service_code in range_data['quantity_ranges']:
                stats = range_data['quantity_ranges'][service_code]
                print(f"   ✅ EXACT MATCH found for service code: {service_code}")

                threshold_min = stats.get('threshold_min', 0)
                threshold_max = stats.get('threshold_max', 0)
                weighted_avg_min = stats.get('weighted_avg_min', stats.get('mean', 0))
                weighted_avg_max = stats.get('weighted_avg_max', stats.get('mean', 0))
                min_threshold_files = stats.get('min_files', [])
                max_threshold_files = stats.get('max_files', [])
                unit = stats.get('unit', '')

                print(f"   📊 Stats: Weighted {weighted_avg_min} - {weighted_avg_max}")
                print(f"   📊 Threshold: {threshold_min} - {threshold_max}")

                # Handle zero quantities
                if quantity == 0:
                    zero_msg = "Zero quantity - skipped validation (Reliance)"
                    return True, zero_msg, service_code, 1.0, threshold_min, threshold_max, weighted_avg_min, weighted_avg_max, min_threshold_files, max_threshold_files

                # Round if number-based unit
                if unit and self.is_number_based_unit(unit):
                        # ✅ NEW: Round to 2 decimal places for Reliance
                    threshold_min = round(threshold_min, 2)
                    threshold_max = round(threshold_max, 2)
                    weighted_avg_min = round(weighted_avg_min, 2)
                    weighted_avg_max = round(weighted_avg_max, 2)
                    quantity = round(quantity, 2)

                # Validate against weighted average range
                is_valid = weighted_avg_min <= quantity <= weighted_avg_max

                if is_valid:
                    msg = f"✅ Within range: {weighted_avg_min} - {weighted_avg_max}"
                else:
                    msg = f"❌ Outside range: {weighted_avg_min} - {weighted_avg_max}"

                print(f"   {msg}")
                return is_valid, msg, service_code, 1.0, threshold_min, threshold_max, weighted_avg_min, weighted_avg_max, min_threshold_files, max_threshold_files

            # SIMILARITY MATCH CHECK
            print(f"   🔍 No exact match, trying similarity search...")
            similar_service_code, similarity = self.find_similar_particular(service_code, sqfeet_range, model_key=model_key, threshold=0.7)

            if similar_service_code and similarity >= 0.7:
                stats = range_data['quantity_ranges'][similar_service_code]
                print(f"   ✅ SIMILAR MATCH found: {similar_service_code} (similarity: {similarity:.2f})")

                threshold_min = stats.get('threshold_min', 0)
                threshold_max = stats.get('threshold_max', 0)
                weighted_avg_min = stats.get('weighted_avg_min', stats.get('mean', 0))
                weighted_avg_max = stats.get('weighted_avg_max', stats.get('mean', 0))
                min_threshold_files = stats.get('min_files', [])
                max_threshold_files = stats.get('max_files', [])
                unit = stats.get('unit', '')

                if quantity == 0:
                    zero_msg = f"Zero quantity - skipped validation (Similar: {similar_service_code})"
                    return True, zero_msg, similar_service_code, similarity, threshold_min, threshold_max, weighted_avg_min, weighted_avg_max, min_threshold_files, max_threshold_files

                if unit and self.is_number_based_unit(unit):
                    # ✅ NEW: Round to 2 decimal places for Reliance
                    threshold_min = round(threshold_min, 2)
                    threshold_max = round(threshold_max, 2)
                    weighted_avg_min = round(weighted_avg_min, 2)
                    weighted_avg_max = round(weighted_avg_max, 2)
                    quantity = round(quantity, 2)

                is_valid = weighted_avg_min <= quantity <= weighted_avg_max

                if is_valid:
                    msg = f"✅ Similar match ({similarity:.2f}): {weighted_avg_min} - {weighted_avg_max}"
                else:
                    msg = f"❌ Similar match ({similarity:.2f}): Outside {weighted_avg_min} - {weighted_avg_max}"

                return is_valid, msg, similar_service_code, similarity, threshold_min, threshold_max, weighted_avg_min, weighted_avg_max, min_threshold_files, max_threshold_files

            # NO MATCH FOUND
            print(f"   ❌ No match found for service code: {service_code}")
            return True, "No matching service code found - automatically approved", None, 0, None, None, None, None, [], []

        
            

        # Check if range exists
        if sqfeet_range not in self.sqfeet_ranges:
            return True, f"⚠️ Range {sqfeet_range} not found in model", None, 0, None, None, None, None, [], []

        range_data = self.sqfeet_ranges[sqfeet_range]

        # 🔥 KEY FIX: Select the correct key based on MODE
        range_key = None

        if mode == 'quantity':
            # For quantity mode, prioritize quantity_ranges
            if 'quantity_ranges' in range_data and range_data['quantity_ranges']:
                range_key = 'quantity_ranges'
            elif 'rate_ranges' in range_data and range_data['rate_ranges']:
                range_key = 'rate_ranges'
            elif 'amount_ranges' in range_data and range_data['amount_ranges']:
                range_key = 'amount_ranges'

        elif mode == 'amount':
            # For amount mode, prioritize amount_ranges
            if 'amount_ranges' in range_data and range_data['amount_ranges']:
                range_key = 'amount_ranges'
            elif 'rate_ranges' in range_data and range_data['rate_ranges']:
                range_key = 'rate_ranges'
            elif 'quantity_ranges' in range_data and range_data['quantity_ranges']:
                range_key = 'quantity_ranges'

        else:  # mode == 'rate'
            # For rate mode, prioritize rate_ranges
            if 'rate_ranges' in range_data and range_data['rate_ranges']:
                range_key = 'rate_ranges'
            elif 'amount_ranges' in range_data and range_data['amount_ranges']:
                range_key = 'amount_ranges'
            elif 'quantity_ranges' in range_data and range_data['quantity_ranges']:
                range_key = 'quantity_ranges'

        if not range_key:
            available_keys = list(range_data.keys())
            print(f"⚠️ Could not find appropriate data key. Available: {available_keys}")
            return True, f"⚠️ No validation data found for {sqfeet_range}", None, 0, None, None, None, None, [], []

        print(f"✓ Using '{range_key}' for {mode} mode in range {sqfeet_range}")

        # Continue with rest of validation...
        cleaned_particular = self.clean_text(particular)
        
        # 1. Try fetching regional stats for EXACT match
        regional_stats = {}
        if is_regional and selected_region:
            regions = ['east', 'west', 'north', 'south']
            for region in regions:
                model_key = f'icici_rate_{region}'
                regional_validator = get_validator(model_key)
                if regional_validator is not None:
                    if sqfeet_range in regional_validator.sqfeet_ranges:
                        regional_range_data = regional_validator.sqfeet_ranges[sqfeet_range]
                        if range_key in regional_range_data and cleaned_particular in regional_range_data[range_key]:
                            regional_stats[region] = regional_range_data[range_key][cleaned_particular]
        
        # --- EXACT MATCH LOGIC ---
        if cleaned_particular in range_data[range_key]:
            stats = range_data[range_key][cleaned_particular]

            threshold_min = stats['overall_min']
            threshold_max = stats['overall_max']
            weighted_avg_min = stats.get('weighted_avg_min', stats['overall_mean'])
            weighted_avg_max = stats.get('weighted_avg_max', stats['overall_mean'])
            min_threshold_files = stats.get('min_threshold_files', [])
            max_threshold_files = stats.get('max_threshold_files', [])
            
            # ✅ NEW: ALWAYS treat zero as VALID (don't check zero_allowed flag)
            if quantity == 0:
                zero_count = stats.get('zero_count', 0)
                total_count = stats.get('total_count', 1)
                zero_msg = f"Zero quantity allowed - appears in {zero_count}/{total_count} training cases"
                
                if is_regional:
                    regional_messages = {}
                    for region in ['east', 'west', 'north', 'south']:
                        if region in regional_stats:
                            r_zero_count = regional_stats[region].get('zero_count', 0)
                            r_total_count = regional_stats[region].get('total_count', 1)
                            regional_messages[region] = f"Zero allowed ({r_zero_count}/{r_total_count} cases)"
                        else:
                            regional_messages[region] = "No data"
                    return True, regional_messages, cleaned_particular, 1.0, threshold_min, threshold_max, weighted_avg_min, weighted_avg_max, min_threshold_files, max_threshold_files
                
                return True, zero_msg, cleaned_particular, 1.0, threshold_min, threshold_max, weighted_avg_min, weighted_avg_max, min_threshold_files, max_threshold_files
            
            unit = stats.get('unit', None)
            
            if unit and self.is_number_based_unit(unit):
                threshold_min = round(threshold_min)
                threshold_max = round(threshold_max)
                weighted_avg_min = round(weighted_avg_min)
                weighted_avg_max = round(weighted_avg_max)
            
            # Build validation message with all regional ranges (EXACT MATCH)
            if is_regional:
                regional_messages = {}
                for region in ['east', 'west', 'north', 'south']:
                    if region in regional_stats:
                        r_stats = regional_stats[region]
                        r_min = r_stats.get('weighted_avg_min', r_stats.get('overall_mean', 0))
                        r_max = r_stats.get('weighted_avg_max', r_stats.get('overall_mean', 0))
                        
                        if unit and self.is_number_based_unit(unit):
                            r_min = round(r_min)
                            r_max = round(r_max)
                            regional_messages[region] = f"{int(r_min)} - {int(r_max)}"
                        else:
                            regional_messages[region] = f"{self.format_indian_currency(r_min)} - {self.format_indian_currency(r_max)}"
                    else:
                        regional_messages[region] = "No data available" # Better than NaN

                # Validate against SELECTED REGION for is_valid
                if selected_region and selected_region in regional_stats:
                    validation_stats = regional_stats[selected_region]
                    val_min = validation_stats.get('weighted_avg_min', validation_stats['overall_mean'])
                    val_max = validation_stats.get('weighted_avg_max', validation_stats['overall_mean'])

                    if unit and self.is_number_based_unit(unit):
                        is_valid = val_min <= round(quantity) <= val_max
                    else:
                        is_valid = val_min <= quantity <= val_max
                else:
                    # If regional mode but selected region has no data for this item, fallback to main model or mark invalid?
                    # Usually better to be lenient or fallback to main model check
                    is_valid = weighted_avg_min <= quantity <= weighted_avg_max

                return is_valid, regional_messages, cleaned_particular, 1.0, threshold_min, threshold_max, weighted_avg_min, weighted_avg_max, min_threshold_files, max_threshold_files
            
        # --- SIMILAR MATCH LOGIC ---
        similar_particular, similarity = self.find_similar_particular(particular, sqfeet_range, model_key=model_key)
        if similar_particular:
            stats = range_data[range_key][similar_particular]

            # >>>>> KEY FIX STARTS HERE <<<<<
            # If we found a similar match, we MUST re-fetch regional stats for that similar name
            if is_regional and selected_region:
                regional_stats = {} # Reset stats
                regions = ['east', 'west', 'north', 'south']
                for region in regions:
                    model_key = f'icici_rate_{region}'
                    regional_validator = get_validator(model_key)
                    if regional_validator is not None:
                        if sqfeet_range in regional_validator.sqfeet_ranges:
                            regional_range_data = regional_validator.sqfeet_ranges[sqfeet_range]
                            # Look for the SIMILAR match name in the regional data
                            if range_key in regional_range_data and similar_particular in regional_range_data[range_key]:
                                regional_stats[region] = regional_range_data[range_key][similar_particular]
            # >>>>> KEY FIX ENDS HERE <<<<<

            threshold_min = stats['overall_min']
            threshold_max = stats['overall_max']
            weighted_avg_min = stats.get('weighted_avg_min', stats['overall_mean'])
            weighted_avg_max = stats.get('weighted_avg_max', stats['overall_mean'])
            min_threshold_files = stats.get('min_threshold_files', [])
            max_threshold_files = stats.get('max_threshold_files', [])
            
            # ✅ NEW: ALWAYS treat zero as VALID for similar matches too
            if quantity == 0:
                zero_count = stats.get('zero_count', 0)
                total_count = stats.get('total_count', 1)
                zero_msg = f"Zero quantity allowed (Similar Match) - appears in {zero_count}/{total_count} training cases"
                
                if is_regional:
                    regional_messages = {}
                    for region in ['east', 'west', 'north', 'south']:
                        if region in regional_stats:
                            r_zero_count = regional_stats[region].get('zero_count', 0)
                            r_total_count = regional_stats[region].get('total_count', 1)
                            regional_messages[region] = f"Zero allowed ({r_zero_count}/{r_total_count} cases)"
                        else:
                            regional_messages[region] = "No data"
                    return True, regional_messages, similar_particular, similarity, threshold_min, threshold_max, weighted_avg_min, weighted_avg_max, min_threshold_files, max_threshold_files
                
                return True, zero_msg, similar_particular, similarity, threshold_min, threshold_max, weighted_avg_min, weighted_avg_max, min_threshold_files, max_threshold_files
            
            unit = stats.get('unit', None)
            
            if unit and self.is_number_based_unit(unit):
                threshold_min = round(threshold_min)
                threshold_max = round(threshold_max)
                weighted_avg_min = round(weighted_avg_min)
                weighted_avg_max = round(weighted_avg_max)
            
            # Build validation message for regional (SIMILAR MATCH)
            if is_regional:
                regional_messages = {}
                for region in ['east', 'west', 'north', 'south']:
                    if region in regional_stats:
                        r_stats = regional_stats[region]
                        r_min = r_stats.get('weighted_avg_min', r_stats.get('overall_mean', 0))
                        r_max = r_stats.get('weighted_avg_max', r_stats.get('overall_mean', 0))
                        
                        if unit and self.is_number_based_unit(unit):
                            r_min = round(r_min)
                            r_max = round(r_max)
                            regional_messages[region] = f"{int(r_min)} - {int(r_max)}"
                        else:
                            regional_messages[region] = f"{self.format_indian_currency(r_min)} - {self.format_indian_currency(r_max)}"
                    else:
                        regional_messages[region] = "No data available"

                # Validate against selected region
                if selected_region and selected_region in regional_stats:
                    validation_stats = regional_stats[selected_region]
                    val_min = validation_stats.get('weighted_avg_min', validation_stats['overall_mean'])
                    val_max = validation_stats.get('weighted_avg_max', validation_stats['overall_mean'])

                    if unit and self.is_number_based_unit(unit):
                        is_valid = val_min <= round(quantity) <= val_max
                    else:
                        is_valid = val_min <= quantity <= val_max
                else:
                    is_valid = weighted_avg_min <= quantity <= weighted_avg_max

                return is_valid, regional_messages, similar_particular, similarity, threshold_min, threshold_max, weighted_avg_min, weighted_avg_max, min_threshold_files, max_threshold_files
            
            # Standard (non-regional) similar match
            if unit and self.is_number_based_unit(unit):
                msg = f"Range (from similar match): {int(weighted_avg_min)} - {int(weighted_avg_max)}"
                is_valid = weighted_avg_min <= round(quantity) <= weighted_avg_max
            else:
                msg = f"Range (from similar match): {weighted_avg_min:.2f} - {weighted_avg_max:.2f}"
                is_valid = weighted_avg_min <= quantity <= weighted_avg_max
            
            return is_valid, msg, similar_particular, similarity, threshold_min, threshold_max, weighted_avg_min, weighted_avg_max, min_threshold_files, max_threshold_files
        
        # No match found
        if is_regional:
             # Return valid structure even if no match
             return True, {r: "No match" for r in ['east','west','north','south']}, None, 0, None, None, None, None, [], []

        return True, "No matching particular found - automatically approved", None, 0, None, None, None, None, [], []
    def validate_test_file(self, test_file_path, mode='quantity', is_regional=False, model_key=None):
        """Validate test file based on mode (quantity, amount, or rate)"""
        # ✅ FIX: Use appropriate extraction based on model type
        if model_key and 'reliance' in model_key:
            if 'smart' in model_key:
                sqfeet_range = self.extract_sqfeet_from_filename_smart(os.path.basename(test_file_path))
                print(f"🏪 Using SMART range: {sqfeet_range}")
            else:
                sqfeet_range = self.extract_sqfeet_from_filename_reliance(os.path.basename(test_file_path))
                print(f"🏪 Using RELIANCE (YOUSTA) range: {sqfeet_range}")
        elif is_regional:
            sqfeet_range = self.extract_sqfeet_from_filename_regional(os.path.basename(test_file_path))
            print(f"🗺️  Using REGIONAL range: {sqfeet_range}")
        else:
            sqfeet_range = self.extract_sqfeet_from_filename(os.path.basename(test_file_path), mode=mode)
            print(f"📏 Using STANDARD range: {sqfeet_range}")

        if not sqfeet_range or sqfeet_range not in self.sqfeet_ranges:
            return None, f"Could not determine Sq. Feet range from filename or no rules for this range. Available ranges: {list(self.sqfeet_ranges.keys())}"

        try:
            df = pd.read_excel(test_file_path, sheet_name='Extracted Data')

            # Define expected columns based on mode
            if mode == 'amount':
                expected_cols = ['Sr.No', 'Particulars', 'As Per CWI (Amount)', 'Unit']
                qty_col_name = 'As Per CWI (Amount)'
            elif mode == 'rate':
                expected_cols = ['Sr.No', 'Particulars', 'Rate', 'Unit', 'As Per CWI (Qty)', 'As Per CWI (Amount)']
                qty_col_name = 'Rate'
            else:  # quantity mode
                expected_cols = ['Sr.No', 'Particulars', 'As Per CWI (Qty)', 'Unit']
                qty_col_name = 'As Per CWI (Qty)'
                if model_key == 'reliance_qty_yousta':
                    expected_cols.insert(1, 'Services Codes')

            col_mapping = {}

            # Print available columns for debugging
            print(f"Available columns in Excel: {df.columns.tolist()}")
            print(f"Looking for columns: {expected_cols}")

            for expected_col in expected_cols:
                found = False
                for df_col in df.columns:
                    # More aggressive cleaning for matching
                    clean_expected = expected_col.lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
                    clean_df_col = str(df_col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')

                    # Check if they match
                    if clean_expected == clean_df_col or clean_expected in clean_df_col or clean_df_col in clean_expected:
                        col_mapping[expected_col] = df_col
                        found = True
                        print(f"✓ Matched '{expected_col}' to '{df_col}'")
                        break

                if not found:
                    print(f"✗ Could not find column: '{expected_col}'")
                    return None, f"Required column '{expected_col}' could not be found. Available columns: {df.columns.tolist()}"

            df_renamed = df.rename(columns={v: k for k, v in col_mapping.items()})
            results = []

            # ✅ NEW: Get selected region from Flask session (if available)
            selected_region = None
            try:
                from flask import session
                selected_region = session.get('last_region', '')
                if selected_region:
                    print(f"📍 Selected region from session: {selected_region}")
            except:
                print("⚠️  Could not access Flask session")

            for _, row in df_renamed.iterrows():
                # ✅ CRITICAL FIX: Initialize BOTH variables
                service_code_val = None  # For Reliance (service code)
                particular_to_validate = None  # For ALL models (what we validate against)

                # ✅ NEW: For Reliance (both Yousta AND Smart), use Service Code as primary identifier
                if model_key and 'reliance' in model_key:
                    service_code = row.get('Services Codes', '')

                    # Skip if service code is missing
                    if pd.isna(service_code) or str(service_code).strip() == '':
                        print(f"  ⚠️ Skipping row with missing service code")
                        continue

                    # Use ACTUAL Particulars column for display
                    particular_val = row.get('Particulars')
                    service_code_val = str(service_code).strip()
                    particular_to_validate = service_code_val  # ← Validate using service code
                    print(f"\n🔍 Processing Service Code: {service_code_val}")
                
                else:
                    # For other models (HDFC/ICICI), use Particulars
                    if pd.isna(row.get('Particulars')) or str(row.get('Particulars')).strip() == '': 
                        continue
                    particular_val = row.get('Particulars')
                    particular_to_validate = particular_val  # ← Validate using particular text
                    # service_code_val remains None

                qty_val = row.get(qty_col_name, 0)
                unit_val = row.get('Unit', '')

                # Debug output
                print(f"\n🔍 Row {_}: {str(particular_val)[:50]}...")
                print(f"   Column '{qty_col_name}' value: {qty_val} (type: {type(qty_val).__name__})")
                print(f"   Unit: {unit_val}")

                # Infer unit if missing
                if pd.isna(unit_val) or str(unit_val).strip() == '':
                    unit_val = self.infer_unit_from_particular(particular_val)
                    print(f"   Inferred unit: {unit_val}")

                # Extract quantity
                print(f"   Calling extract_quantity...")
                qty = self.extract_quantity(qty_val)
                print(f"   After extraction: {qty}")

                # ✅ For Reliance, always round to 2 decimals
                if model_key == 'reliance_qty_yousta':
                    qty = round(qty, 2)
                else:
                    qty = self.round_quantity_if_number_based(qty, unit_val)

                print(f"   After rounding: {qty}")

                # ✅ SKIP ROWS WITH ZERO QUANTITIES FOR ALL RELIANCE MODELS
                if model_key and 'reliance' in model_key and qty == 0:
                    print(f"  ⚠️  Skipping row with zero quantity: {particular_val}")
                    continue

                # ✅ Skip zero quantities in both qty and amount modes
                if mode == 'amount' and qty == 0:
                    print(f"  ⚠️  Skipping particular with zero amount: {particular_val}")
                    continue
                
                if mode == 'quantity' and qty == 0:
                    print(f"  ⚠️  Skipping particular with zero quantity: {particular_val}")
                    continue

                # ✅ CRITICAL FIX: Pass particular_to_validate (which has the right value)
                validation_result = self.validate_quantity(
                    particular_to_validate,  # ← NOW THIS IS CORRECT FOR ALL MODELS
                    qty, 
                    sqfeet_range, 
                    is_regional=is_regional, 
                    selected_region=selected_region, 
                    model_key=model_key
                )

                # ✅ NEW: Handle regional validation results differently
                if is_regional and isinstance(validation_result[1], dict):
                    # Regional mode returns dictionary of messages
                    is_valid = validation_result[0]
                    regional_msgs = validation_result[1]
                    matched = validation_result[2]
                    sim = validation_result[3]
                    threshold_min = validation_result[4]
                    threshold_max = validation_result[5]
                    weighted_avg_min = validation_result[6]
                    weighted_avg_max = validation_result[7]
                    min_threshold_files = validation_result[8]
                    max_threshold_files = validation_result[9]
                else:
                    # Standard mode returns single message string
                    is_valid, msg, matched, sim, threshold_min, threshold_max, weighted_avg_min, weighted_avg_max, min_threshold_files, max_threshold_files = validation_result
                    regional_msgs = None

                # Build result dict based on mode
                if mode == 'rate':
                    rate_value = qty
                    qty_value = self.extract_quantity(row.get('As Per CWI (Qty)', 0))
                    amount_value = self.extract_quantity(row.get('As Per CWI (Amount)', 0))
                    qty_value = self.round_quantity_if_number_based(qty_value, unit_val)

                    # ✅ NEW: Build result dict with regional columns if applicable
                    if is_regional and regional_msgs:
                        result_dict = {
                            'Sr.No': row.get('Sr.No', ''), 
                            'Original_Particular': str(particular_val),
                            'Rate': rate_value,
                            'Unit': unit_val,
                            'Quantity': qty_value,
                            'Amount': amount_value,
                            'Sq_Feet_Range': sqfeet_range, 
                            'Is_Valid': is_valid,
                            'Validation_Message_East': regional_msgs.get('east', 'N/A'),
                            'Validation_Message_West': regional_msgs.get('west', 'N/A'),
                            'Validation_Message_North': regional_msgs.get('north', 'N/A'),
                            'Validation_Message_South': regional_msgs.get('south', 'N/A'),
                            'Selected_Region': selected_region,
                            'Matched_Particular': matched, 
                            'Similarity_Score': sim,
                            'Threshold_Min': threshold_min,
                            'Threshold_Max': threshold_max,
                            'Weighted_Avg_Min': weighted_avg_min,
                            'Weighted_Avg_Max': weighted_avg_max,
                            'Min_Threshold_Files': min_threshold_files,
                            'Max_Threshold_Files': max_threshold_files
                        }
                    else:
                        result_dict = {
                            'Sr.No': row.get('Sr.No', ''), 
                            'Original_Particular': str(particular_val),
                            'Rate': rate_value,
                            'Unit': unit_val,
                            'Quantity': qty_value,
                            'Amount': amount_value,
                            'Sq_Feet_Range': sqfeet_range, 
                            'Is_Valid': is_valid,
                            'Validation_Message': msg, 
                            'Matched_Particular': matched, 
                            'Similarity_Score': sim,
                            'Threshold_Min': threshold_min,
                            'Threshold_Max': threshold_max,
                            'Weighted_Avg_Min': weighted_avg_min,
                            'Weighted_Avg_Max': weighted_avg_max,
                            'Min_Threshold_Files': min_threshold_files,
                            'Max_Threshold_Files': max_threshold_files
                        }

                elif mode == 'amount':
                    # ✅ NEW: Build result dict with regional columns if applicable
                    if is_regional and regional_msgs:
                        result_dict = {
                            'Sr.No': row.get('Sr.No', ''), 
                            'Original_Particular': str(particular_val),
                            'Amount': qty,
                            'Unit': unit_val,
                            'Sq_Feet_Range': sqfeet_range, 
                            'Is_Valid': is_valid,
                            'Validation_Message_East': regional_msgs.get('east', 'N/A'),
                            'Validation_Message_West': regional_msgs.get('west', 'N/A'),
                            'Validation_Message_North': regional_msgs.get('north', 'N/A'),
                            'Validation_Message_South': regional_msgs.get('south', 'N/A'),
                            'Selected_Region': selected_region,
                            'Matched_Particular': matched, 
                            'Similarity_Score': sim,
                            'Threshold_Min': threshold_min,
                            'Threshold_Max': threshold_max,
                            'Weighted_Avg_Min': weighted_avg_min,
                            'Weighted_Avg_Max': weighted_avg_max,
                            'Min_Threshold_Files': min_threshold_files,
                            'Max_Threshold_Files': max_threshold_files
                        }
                    else:
                        result_dict = {
                            'Sr.No': row.get('Sr.No', ''), 
                            'Original_Particular': str(particular_val),
                            'Amount': qty,
                            'Unit': unit_val,
                            'Sq_Feet_Range': sqfeet_range, 
                            'Is_Valid': is_valid,
                            'Validation_Message': msg, 
                            'Matched_Particular': matched, 
                            'Similarity_Score': sim,
                            'Threshold_Min': threshold_min,
                            'Threshold_Max': threshold_max,
                            'Weighted_Avg_Min': weighted_avg_min,
                            'Weighted_Avg_Max': weighted_avg_max,
                            'Min_Threshold_Files': min_threshold_files,
                            'Max_Threshold_Files': max_threshold_files
                        }

                else:  # quantity mode
                    # ✅ NEW: Build result dict with regional columns if applicable
                    if is_regional and regional_msgs:
                        result_dict = {
                            'Sr.No': row.get('Sr.No', ''), 
                            'Original_Particular': str(particular_val),
                            'Quantity': qty,
                            'Unit': unit_val,
                            'Sq_Feet_Range': sqfeet_range, 
                            'Is_Valid': is_valid,
                            'Validation_Message_East': regional_msgs.get('east', 'N/A'),
                            'Validation_Message_West': regional_msgs.get('west', 'N/A'),
                            'Validation_Message_North': regional_msgs.get('north', 'N/A'),
                            'Validation_Message_South': regional_msgs.get('south', 'N/A'),
                            'Selected_Region': selected_region,
                            'Matched_Particular': matched, 
                            'Similarity_Score': sim,
                            'Threshold_Min': threshold_min,
                            'Threshold_Max': threshold_max,
                            'Weighted_Avg_Min': weighted_avg_min,
                            'Weighted_Avg_Max': weighted_avg_max,
                            'Min_Threshold_Files': min_threshold_files,
                            'Max_Threshold_Files': max_threshold_files
                        }
                    else:
                        # Standard quantity mode result
                        result_dict = {
                            'Sr.No': row.get('Sr.No', ''), 
                            'Original_Particular': str(particular_val),
                            'Quantity': qty,
                            'Unit': unit_val,
                            'Sq_Feet_Range': sqfeet_range, 
                            'Is_Valid': is_valid,
                            'Validation_Message': msg, 
                            'Matched_Particular': matched, 
                            'Similarity_Score': sim,
                            'Threshold_Min': threshold_min,
                            'Threshold_Max': threshold_max,
                            'Weighted_Avg_Min': weighted_avg_min,
                            'Weighted_Avg_Max': weighted_avg_max,
                            'Min_Threshold_Files': min_threshold_files,
                            'Max_Threshold_Files': max_threshold_files
                        }

                    # ✅ ADD SERVICE CODE FOR ALL RELIANCE MODELS
                    if model_key and 'reliance' in model_key:
                        result_dict['Service_Code'] = service_code_val
                        # Add the actual Particulars text from the Excel file
                        actual_particular = row.get('Particulars', 'N/A')
                        result_dict['Actual_Particular'] = str(actual_particular) if not pd.isna(actual_particular) else 'N/A'

                # ✅ CRITICAL: Append result_dict for ALL modes
                results.append(result_dict)

            return results, None

        except Exception as e:
            import traceback
            traceback.print_exc()
            return None, f"An unexpected error occurred while processing the Excel file: {e}"

    def format_indian_currency(self, value):
        """Format number to Indian currency style"""
        try:
            if pd.isna(value) or value == 'NA': return str(value)
            val_float = float(value)
            s = "{:.2f}".format(val_float)
            parts = s.split('.')
            integer_part = parts[0]
            decimal_part = parts[1]
            
            if len(integer_part) > 3:
                last_three = integer_part[-3:]
                rest = integer_part[:-3]
                rest = re.sub(r"\B(?=(\d{2})+(?!\d))", ",", rest)
                formatted_int = rest + "," + last_three
            else:
                formatted_int = integer_part
                
            return f"{formatted_int}.{decimal_part}"
        except:
            return str(value)

# --- CWIGroupedValidator Class for Grouped Per Sq.Feet Validation ---
class CWIGroupedValidator:
    def __init__(self):
        # Store data segregated by sq feet ranges (250 sqft intervals)
        self.sqfeet_ranges = {
            '500-750': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '751-1000': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '1001-1250': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '1251-1500': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '1501-1750': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '1751-2000': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '2001-2250': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '2251-2500': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '2501-2750': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '2751-3000': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '3001-3250': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '3251-3500': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '3501-3750': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '3751-4000': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '4001-4250': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '4251-4500': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '4501-4750': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '4751-5000': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)},
            '5000+': {'grouped_ranges': {}, 'particulars_data': defaultdict(list)}
        }
        self.rate_sqfeet_ranges = {}
        # 2. Amount As per CWI (500 sqft buckets) - From Summary Sheet
        # ✅ FIX: Initialize this attribute here to avoid AttributeError
        self.amount_cwi_ranges = defaultdict(lambda: {
            'grouped_ranges': {}, 
            'particulars_data': defaultdict(list)
        })
        self.amount_sqfeet_ranges = {}
        
        # 3. >>> NEW: Subcategory specific ranges (The fix for your error) <<<
        self.rate_per_sqft_250_ranges = {} 
        self.rate_per_sqft_500_ranges = {}
        self.amount_per_cwi_250_ranges = {}

        # 4. Helpers
        self.vectorizers = {}
        self.particular_vectors = {}
        self.particular_names = {}
        self.training_stats = {}
        self.known_groups = {}
        self.all_groups = set()
        self.all_subcategories = {} # To store the mapping
        self.fallback_group = 'Others/Uncategorized'
        self.new_groups_found = set()
        
        # Standard grouped categories with their keywords
        self.known_groups = {
            'Civil Works': ['civil', 'civl'],
            'Carpentry Works': ['carpentry', 'carpentary', 'carpenter', 'carpentory'],
            'Non-Interior Works': ['non-int', 'non int', 'nonint', 'non interior', 'noninterior'],
            'Electrical Works': ['electrical', 'electric', 'electricl'],
            'Additional Works': ['additional', 'addi', 'additonal', 'aditional'],
            'Reimbursement of Vendor Elements': ['reimbursement', 'reimbrsement', 'reimburse', 'vendor', 'reimbursment'],
            'Electricity of Diesel Expense': ['electricity of diesel', 'diesel expense', 'electricity diesel', 'diesel']
        }
   
    def extract_sqfeet_from_filename(self, filename):
        """Extract square feet range from filename (250 sqft intervals)"""
        filename_lower = filename.lower()

        # Patterns for 250 sqft intervals
        patterns = [
            (r'500\s*-?\s*750', '500-750'),
            (r'751\s*-?\s*1000', '751-1000'),
            (r'1001\s*-?\s*1250', '1001-1250'),
            (r'1251\s*-?\s*1500', '1251-1500'),
            (r'1501\s*-?\s*1750', '1501-1750'),
            (r'1751\s*-?\s*2000', '1751-2000'),
            (r'2001\s*-?\s*2250', '2001-2250'),
            (r'2251\s*-?\s*2500', '2251-2500'),
            (r'2501\s*-?\s*2750', '2501-2750'),
            (r'2751\s*-?\s*3000', '2751-3000'),
            (r'3001\s*-?\s*3250', '3001-3250'),
            (r'3251\s*-?\s*3500', '3251-3500'),
            (r'3501\s*-?\s*3750', '3501-3750'),
            (r'3751\s*-?\s*4000', '3751-4000'),
            (r'4001\s*-?\s*4250', '4001-4250'),
            (r'4251\s*-?\s*4500', '4251-4500'),
            (r'4501\s*-?\s*4750', '4501-4750'),
            (r'4751\s*-?\s*5000', '4751-5000')
        ]

        for pattern, range_key in patterns:
            if re.search(pattern, filename_lower):
                return range_key

        # Check for 5000+ pattern
        if re.search(r'5000\+|above\s*5000|more\s*than\s*5000', filename_lower):
            return '5000+'

        # Numeric logic for 250 sqft intervals
        numbers = re.findall(r'\d+', filename_lower)
        if numbers:
            for num_str in numbers:
                num = int(num_str)
                
                if 500 <= num <= 750:
                    return '500-750'
                elif 751 <= num <= 1000:
                    return '751-1000'
                elif 1001 <= num <= 1250:
                    return '1001-1250'
                elif 1251 <= num <= 1500:
                    return '1251-1500'
                elif 1501 <= num <= 1750:
                    return '1501-1750'
                elif 1751 <= num <= 2000:
                    return '1751-2000'
                elif 2001 <= num <= 2250:
                    return '2001-2250'
                elif 2251 <= num <= 2500:
                    return '2251-2500'
                elif 2501 <= num <= 2750:
                    return '2501-2750'
                elif 2751 <= num <= 3000:
                    return '2751-3000'
                elif 3001 <= num <= 3250:
                    return '3001-3250'
                elif 3251 <= num <= 3500:
                    return '3251-3500'
                elif 3501 <= num <= 3750:
                    return '3501-3750'
                elif 3751 <= num <= 4000:
                    return '3751-4000'
                elif 4001 <= num <= 4250:
                    return '4001-4250'
                elif 4251 <= num <= 4500:
                    return '4251-4500'
                elif 4501 <= num <= 4750:
                    return '4501-4750'
                elif 4751 <= num <= 5000:
                    return '4751-5000'
                elif num > 5000:
                    return '5000+'

        return None
   
    def clean_text(self, text):
        """Clean and normalize text for better matching"""
        if pd.isna(text):
            return ""
        text = str(text).lower()
        text = re.sub(r'[^\w\s:/.-]', ' ', text)
        text = ' '.join(text.split())
        return text
    
    def standardize_group_name(self, particular_text):
        """Standardize particular name - match to known groups or create new group"""
        if pd.isna(particular_text) or str(particular_text).strip() == '':
            return None
    
        cleaned_text = self.clean_text(particular_text)
        original_text = str(particular_text).strip()
    
        # First, check against known groups
        for standard_name, keywords in self.known_groups.items():
            for keyword in keywords:
                if keyword in cleaned_text:
                    self.all_groups.add(standard_name)
                    return standard_name
    
        # If no match found, create a new group name from the original text
        new_group_name = self._create_group_name(original_text)
        self.all_groups.add(new_group_name)
        self.new_groups_found.add(new_group_name)
    
        return new_group_name

    def _create_group_name(self, text):
        """Create a clean group name from text"""
        # Remove 'total of', 'total', prefixes
        text = re.sub(r'^total\s+of\s+', '', text.lower())
        text = re.sub(r'^total\s+', '', text.lower())

        # Capitalize each word
        words = text.split()
        capitalized = ' '.join(word.capitalize() for word in words)

        # Add 'Works' suffix if not present
        if 'work' not in capitalized.lower():
            capitalized += ' Works'
        elif 'works' not in capitalized.lower():
            capitalized = capitalized.replace('Work', 'Works')

        return capitalized
    def format_indian_currency(self, value):
        """Format number to Indian currency style (e.g., 7,00,909.11)"""
        try:
            if pd.isna(value): return "0.00"
            val_float = float(value)
            s = "{:.2f}".format(val_float)
            parts = s.split('.')
            integer_part = parts[0]
            decimal_part = parts[1]
            
            if len(integer_part) > 3:
                last_three = integer_part[-3:]
                rest = integer_part[:-3]
                # Regex to add commas every 2 digits for the Indian style
                rest = re.sub(r"\B(?=(\d{2})+(?!\d))", ",", rest)
                formatted_int = rest + "," + last_three
            else:
                formatted_int = integer_part
                
            return f"{formatted_int}.{decimal_part}"
        except:
            return str(value)
    def extract_amount(self, amount_value):
        """Extract numeric amount from various formats"""
        if pd.isna(amount_value):
            return 0
       
        amount_str = str(amount_value).strip()
        # Remove currency symbols and commas
        amount_str = amount_str.replace('₹', '').replace('Rs', '').replace(',', '').replace('INR', '').strip()
        # Handle negative values
        amount_str = amount_str.replace('(', '-').replace(')', '')
        numbers = re.findall(r'-?\d+\.?\d*', amount_str)
        if numbers:
            return float(numbers[0])
        return 0
    
    def find_similar_group(self, query_particular, sqfeet_range, threshold=0.5):
        """Find most similar group from standard groups within specific sq feet range"""
        if sqfeet_range not in self.vectorizers or sqfeet_range not in self.particular_vectors:
            return None, 0
        
        vectorizer = self.vectorizers[sqfeet_range]
        particular_vectors = self.particular_vectors[sqfeet_range]
        particular_names = self.particular_names[sqfeet_range]
        
        query_cleaned = self.clean_text(query_particular)
        query_vector = vectorizer.transform([query_cleaned])
        
        similarities = cosine_similarity(query_vector, particular_vectors)[0]
        best_match_idx = np.argmax(similarities)
        best_similarity = similarities[best_match_idx]
        
        if best_similarity >= threshold:
            return particular_names[best_match_idx], best_similarity
        
        return None, best_similarity
   
    def validate_grouped_data(self, group_name, cwi_amount, rate, sqfeet_range):
        """Validate CWI amount and rate for a grouped category within specific sq feet range"""
        if sqfeet_range not in self.sqfeet_ranges:
            return {
                'cwi_valid': False,
                'rate_valid': False,
                'message': f"Invalid sq feet range: {sqfeet_range}"
            }
        
        range_data = self.sqfeet_ranges[sqfeet_range]
        
        # Standardize the group name
        standard_group = self.standardize_group_name(group_name)
        
        if standard_group is None:
            # Try similarity matching
            similar_group, similarity = self.find_similar_group(group_name, sqfeet_range)
            if similar_group:
                standard_group = similar_group
            else:
                return {
                    'cwi_valid': False,
                    'rate_valid': False,
                    'message': f"Unknown group: '{group_name}' - not found in standard groups"
                }
        
        # Check if this group exists in training data for this sq feet range
        if standard_group not in range_data['grouped_ranges']:
            return {
                'cwi_valid': False,
                'rate_valid': False,
                'message': f"Group '{standard_group}' not found in training data for {sqfeet_range} range"
            }
        
        group_stats = range_data['grouped_ranges'][standard_group]
        cwi_stats = group_stats['cwi_amount_stats']
        rate_stats = group_stats['rate_stats']
        
        # Validate CWI amount
        cwi_valid = False
        cwi_message = ""
        # --- UPDATED VALIDATION LOGIC WITH FORMATTING ---
        if cwi_amount == 0 and cwi_stats['zero_allowed']:
            cwi_valid = True
            cwi_message = f"Zero CWI amount allowed ({cwi_stats['zero_count']}/{cwi_stats['total_count']} cases)"
        elif cwi_stats['validation_range_lower'] <= cwi_amount <= cwi_stats['validation_range_upper']:
            cwi_valid = True
            cwi_message = f"Within range: ₹{self.format_indian_currency(cwi_stats['validation_range_lower'])} - ₹{self.format_indian_currency(cwi_stats['validation_range_upper'])}"
        else:
            cwi_message = f"Outside range: ₹{self.format_indian_currency(cwi_stats['validation_range_lower'])} - ₹{self.format_indian_currency(cwi_stats['validation_range_upper'])}"
        
        # Validate Rate
        rate_valid = False
        rate_message = ""
        
        if rate == 0 and rate_stats['zero_allowed']:
            rate_valid = True
            rate_message = f"Zero rate allowed ({rate_stats['zero_count']}/{rate_stats['total_count']} cases)"
        elif rate_stats['validation_range_lower'] <= rate <= rate_stats['validation_range_upper']:
            rate_valid = True
            rate_message = f"Within range: ₹{self.format_indian_currency(rate_stats['validation_range_lower'])} - ₹{self.format_indian_currency(rate_stats['validation_range_upper'])}"
        else:
            rate_message = f"Outside range: ₹{self.format_indian_currency(rate_stats['validation_range_lower'])} - ₹{self.format_indian_currency(rate_stats['validation_range_upper'])}"
        
        return {
            'cwi_valid': cwi_valid,
            'rate_valid': rate_valid,
            'standard_group': standard_group,
            'cwi_message': cwi_message,
            'rate_message': rate_message,
            'cwi_stats': cwi_stats,
            'rate_stats': rate_stats
        }
   
    def load_model(self, model_path):
        """Load trained model from file with proper key normalization"""
        try:
            with open(model_path, 'rb') as f:
                model_data = pickle.load(f)

            # --- 1. Load Primary sqfeet_ranges ---
            if 'sqfeet_ranges' in model_data:
                self.sqfeet_ranges = model_data['sqfeet_ranges']
            elif 'rate_sqfeet_ranges' in model_data:
                self.sqfeet_ranges = model_data['rate_sqfeet_ranges']
            elif 'amount_cwi_ranges' in model_data:
                self.sqfeet_ranges = model_data['amount_cwi_ranges']
            else:
                self.sqfeet_ranges = {}

            # 🔥 CRITICAL FIX: Normalize all range data keys
            for range_name, range_data in self.sqfeet_ranges.items():
                # Count which keys exist
                has_quantity = 'quantity_ranges' in range_data and range_data['quantity_ranges']
                has_rate = 'rate_ranges' in range_data and range_data['rate_ranges']
                has_amount = 'amount_ranges' in range_data and range_data['amount_ranges']

                # Determine which key has the actual data
                if has_quantity:
                    source_key = 'quantity_ranges'
                    source_data = range_data['quantity_ranges']
                elif has_rate:
                    source_key = 'rate_ranges'
                    source_data = range_data['rate_ranges']
                elif has_amount:
                    source_key = 'amount_ranges'
                    source_data = range_data['amount_ranges']
                else:
                    # No data in this range
                    continue

                # 🔥 COPY SOURCE DATA TO ALL THREE KEYS
                range_data['quantity_ranges'] = source_data
                range_data['rate_ranges'] = source_data
                range_data['amount_ranges'] = source_data

                print(f"  ✓ Normalized range {range_name}: copied '{source_key}' to all keys ({len(source_data)} items)")

            # --- 2. Load Text Similarity Assets ---
            self.vectorizers = model_data.get('vectorizers', {})
            self.particular_vectors = model_data.get('particular_vectors', {})
            self.particular_names = model_data.get('particular_names', {})

            # --- 3. Load Other Data Structures ---
            self.training_stats = model_data.get('training_stats', {})
            self.rate_sqfeet_ranges = model_data.get('rate_sqfeet_ranges', {})
            self.amount_cwi_ranges = model_data.get('amount_cwi_ranges', defaultdict(lambda: {
                'grouped_ranges': {}, 'particulars_data': defaultdict(list)
            }))
            self.rate_per_sqft_250_ranges = model_data.get('rate_per_sqft_250_ranges', {})
            self.rate_per_sqft_500_ranges = model_data.get('rate_per_sqft_500_ranges', defaultdict(lambda: {
                'grouped_ranges': {}, 'subcategory_data': defaultdict(list)
            }))
            self.amount_per_cwi_250_ranges = model_data.get('amount_per_cwi_250_ranges', defaultdict(lambda: {
                'grouped_ranges': {}, 'subcategory_data': defaultdict(list)
            }))
            self.known_groups = model_data.get('known_groups', {})
            self.fallback_group = model_data.get('fallback_group', 'Others')
            self.all_subcategories = model_data.get('subcategory_map', {})

            print(f"✓ Model loaded successfully: {model_path}")
            print(f"  - Ranges loaded: {list(self.sqfeet_ranges.keys())}")
            print(f"  - Similarity engines active: {len(self.vectorizers) > 0}")
            print(f"  - All keys normalized to: quantity_ranges, rate_ranges, amount_ranges")

            return True

        except Exception as e:
            print(f"❌ Error loading model: {str(e)}")
            import traceback
            traceback.print_exc()
            return False
# --- Flask Application Setup ---
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.secret_key = 'super-secret-key'
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_USE_SIGNER'] = True

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

MODEL_PATHS = {
    'hdfc_qty': "trained_model_sqfeet_zero_aware.pkl",
    'hdfc_amount': "trained_model_sqfeet_zero_aware_logic_amount_HDFC.pkl",  # ADD THIS
    'hdfc_rate': "trained_model_sqfeet_zero_aware_logic_rate_HDFC.pkl",
    'hdfc_grouped': "trained_model_sqfeet_zero_aware_logic_grouped_HDFC.pkl",
    'icici_qty': "trained_model_sqfeet_weighted_avg_ICICI.pkl",
    'icici_amount': "trained_model_sqfeet_zero_aware_logic_amount_ICICI.pkl",
    'icici_rate': "trained_model_sqfeet_zero_aware_logic_rate_ICICI.pkl",
    'icici_grouped': "trained_model_sqfeet_zero_aware_logic_grouped_ICICI.pkl",  # ADD THIS LINE
    'icici_rate_east': "trained_model_sqfeet_zero_aware_logic_rate_EAST.pkl",
    'icici_rate_west': "trained_model_sqfeet_zero_aware_logic_rate_WEST.pkl",
    'icici_rate_north': "trained_model_sqfeet_zero_aware_logic_rate_NORTH.pkl",
    'icici_rate_south': "trained_model_sqfeet_zero_aware_logic_rate_SOUTH.pkl",
    'icici_qty_east': "trained_model_sqfeet_zero_aware_logic_qty_EAST.pkl",  # You'll train these later
    'icici_qty_west': "trained_model_sqfeet_zero_aware_logic_qty_WEST.pkl",
    'icici_qty_north': "trained_model_sqfeet_zero_aware_logic_qty_NORTH.pkl",
    'icici_qty_south': "trained_model_sqfeet_zero_aware_logic_qty_SOUTH.pkl",
    'axis_rate': "trained_model_sqfeet_zero_aware_logic_rate_AXIS.pkl",
    # West region (your current working models)
    'hdfc_grouped_subcategory_west': "trained_model_HDFC_grouped_rate_amount_dumped_West.pkl",
    'icici_grouped_subcategory_west': "trained_model_ICICI_grouped_rate_amount_dumped_West.pkl",
    
    # South region
    'hdfc_grouped_subcategory_south': "trained_model_HDFC_grouped_rate_amount_dumped_South.pkl",
    'icici_grouped_subcategory_south': "trained_model_ICICI_grouped_rate_amount_dumped_South.pkl",
    
    # North region
    'hdfc_grouped_subcategory_north': "trained_model_HDFC_grouped_rate_amount_dumped_North.pkl",
    'icici_grouped_subcategory_north': "trained_model_ICICI_grouped_rate_amount_dumped_North.pkl",
    
    # East region
    'hdfc_grouped_subcategory_east': "trained_model_HDFC_grouped_rate_amount_dumped_East.pkl",
    'icici_grouped_subcategory_east': "trained_model_ICICI_grouped_rate_amount_dumped_East.pkl",
    
    # PAN INDIA
    'hdfc_grouped_subcategory_india': "trained_model_HDFC_grouped_rate_amount_dumped_INDIA.pkl",
    'icici_grouped_subcategory_india': "trained_model_ICICI_grouped_rate_amount_dumped_INDIA.pkl",

    #reliance models
    'reliance_qty_yousta': "trained_model_service_code_pan_india.pkl",
    'reliance_rate_yousta': "trained_model_service_code_pan_india_RATE.pkl",  # Keep on hold for now
    'reliance_qty_smart': "trained_model_service_code_smart_pan_india.pkl",

}

VALIDATORS = {}
GROUPED_VALIDATORS = {}
_loaded_models = {}

def get_validator(model_key):
    """Lazy load validator - only loads when first requested"""
    if model_key in _loaded_models:
        return _loaded_models[model_key]
    
    if model_key not in MODEL_PATHS:
        print(f"❌ Model key '{model_key}' not in MODEL_PATHS")
        return None
    
    model_path = MODEL_PATHS[model_key]
    if not os.path.exists(model_path):
        print(f"❌ Model file not found: {model_path}")
        return None
    
    print(f"⏳ Lazy loading model: {model_key}")
    
    if 'grouped' in model_key:
        validator = CWIGroupedValidator()
        validator.load_model(model_path)
        GROUPED_VALIDATORS[model_key] = validator
    else:
        validator = CWIValidatorTester()
        validator.load_model(model_path)
        VALIDATORS[model_key] = validator
    
    _loaded_models[model_key] = validator
    print(f"✅ {model_key.upper()} loaded successfully.")
    return validator

# Verify model files exist at startup (no loading)
for client, model_path in MODEL_PATHS.items():
    if os.path.exists(model_path):
        print(f"✅ Model file found: {client}")
    else:
        print(f"❌ WARNING: Model file not found: {model_path}")

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_thresholds(message):
    if pd.isna(message): return "NA", "NA"
    # NEW: Updated regex to handle both integer and decimal ranges
    range_match = re.search(r'Range.*?:\s*([\d.]+)\s*-\s*([\d.]+)', str(message))
    if range_match: 
        val1 = range_match.group(1)
        val2 = range_match.group(2)
        # NEW: Return as int if no decimal point, else float
        return (int(val1) if '.' not in val1 else float(val1), 
                int(val2) if '.' not in val2 else float(val2))
    if "Zero quantity allowed" in str(message): return 0, 0
    return "NA", "NA"

def highlight_invalid(row):
    color = '#FFC7CE'
    style = f'background-color: {color}'
    styles = [''] * len(row)
    if not row['Is_Valid']:
        # Highlight the value column
        if 'Amount' in row.index:
            value_col = 'Amount'
        elif 'Rate' in row.index:
            value_col = 'Rate'
        else:
            value_col = 'Quantity'
        styles[row.index.get_loc(value_col)] = style
        styles[row.index.get_loc('Is_Valid')] = style
        # For grouped mode, also highlight the individual validity columns
        if 'Is_Valid_Amount' in row.index and not row.get('Is_Valid_Amount', True):
            styles[row.index.get_loc('Is_Valid_Amount')] = style
            if 'As_Per_CWI_Amount' in row.index:
                styles[row.index.get_loc('As_Per_CWI_Amount')] = style
        if 'Is_Valid_Rate' in row.index and not row.get('Is_Valid_Rate', True):
            styles[row.index.get_loc('Is_Valid_Rate')] = style
            if 'Rate_per_sqfeet' in row.index:
                styles[row.index.get_loc('Rate_per_sqfeet')] = style
    return styles
def validate_grouped_file(validator, filepath):
    """Validate grouped per sq.feet data from Summary Sheet"""
    try:
        df = pd.read_excel(filepath, sheet_name='Summary Sheet')
        
        # Find CWI column
        cwi_col = None
        for col in df.columns:
            if 'cwi' in str(col).lower():
                cwi_col = col
                break
        
        if cwi_col is None:
            return None, "Could not find column containing 'CWI' in Summary Sheet"
        
        # Last column is Rate per sq.feet
        rate_col = df.columns[-1]
        
        # Find total row to stop processing
        total_idx = None
        for idx, row in df.iterrows():
            first_val = str(row.iloc[0]).strip().lower() if not pd.isna(row.iloc[0]) else ''
            second_val = str(row.iloc[1]).strip().lower() if not pd.isna(row.iloc[1]) else ''
            if first_val in ['total', 'grand total'] or second_val in ['total', 'grand total']:
                total_idx = idx
                break
        
        if total_idx is not None:
            df = df.iloc[:total_idx + 1]
        
        # Extract sq feet range from filename
        sqfeet_range = validator.extract_sqfeet_from_filename(os.path.basename(filepath))
        if not sqfeet_range:
            return None, "Could not determine sq feet range from filename"
        
        results = []
        for idx, row in df.iterrows():
            sr_no = row.iloc[0]
            particular = row.iloc[1]
            
            if pd.isna(particular) or str(particular).strip() == '':
                continue
            
            # Extract amounts
            cwi_amount = validator.extract_amount(row[cwi_col])
            rate = validator.extract_amount(row[rate_col])
            
            # Validate
            validation_result = validator.validate_grouped_data(
                particular, cwi_amount, rate, sqfeet_range
            )
            
            # ✅ NEW: Extract weighted averages from stats
            cwi_stats = validation_result.get('cwi_stats', {})
            rate_stats = validation_result.get('rate_stats', {})
            
            # Skip "Others/Uncategorized" rows entirely
            standard_group_name = validation_result.get('standard_group', '')
            if standard_group_name and 'other' in str(standard_group_name).lower():
                continue

            # Determine validity based on weighted min/max ranges
            cwi_w_min = cwi_stats.get('weighted_avg_min', None)
            cwi_w_max = cwi_stats.get('weighted_avg_max', None)
            rate_w_min = rate_stats.get('weighted_avg_min', None)
            rate_w_max = rate_stats.get('weighted_avg_max', None)

            # Threshold (absolute) min/max for tooltip display
            cwi_t_min = cwi_stats.get('overall_min', cwi_stats.get('validation_range_lower', 0))
            cwi_t_max = cwi_stats.get('overall_max', cwi_stats.get('validation_range_upper', 0))
            rate_t_min = rate_stats.get('overall_min', rate_stats.get('validation_range_lower', 0))
            rate_t_max = rate_stats.get('overall_max', rate_stats.get('validation_range_upper', 0))

            # Get files at threshold min/max
            cwi_min_files = cwi_stats.get('min_threshold_files', [])
            cwi_max_files = cwi_stats.get('max_threshold_files', [])
            rate_min_files = rate_stats.get('min_threshold_files', [])
            rate_max_files = rate_stats.get('max_threshold_files', [])

            if cwi_w_min is not None and cwi_w_max is not None:
                is_valid_amount = (cwi_amount == 0 and cwi_stats.get('zero_allowed', False)) or \
                                  (cwi_w_min <= cwi_amount <= cwi_w_max)
            else:
                is_valid_amount = validation_result['cwi_valid']

            if rate_w_min is not None and rate_w_max is not None:
                is_valid_rate = (rate == 0 and rate_stats.get('zero_allowed', False)) or \
                                (rate_w_min <= rate <= rate_w_max)
            else:
                is_valid_rate = validation_result['rate_valid']

            result_dict = {
                'Sr.No': sr_no,
                'Particulars': str(particular),
                'As_Per_CWI_Amount': cwi_amount,
                'Rate_per_sqfeet': rate,
                'Sq_Feet_Range': sqfeet_range,
                'Is_Valid': is_valid_amount and is_valid_rate,
                'Is_Valid_Amount': is_valid_amount,
                'Is_Valid_Rate': is_valid_rate,
                'CWI_Weighted_Min': cwi_w_min if cwi_w_min is not None else 'NA',
                'CWI_Weighted_Max': cwi_w_max if cwi_w_max is not None else 'NA',
                'Rate_Weighted_Min': rate_w_min if rate_w_min is not None else 'NA',
                'Rate_Weighted_Max': rate_w_max if rate_w_max is not None else 'NA',
                'CWI_Threshold_Min': cwi_t_min,
                'CWI_Threshold_Max': cwi_t_max,
                'Rate_Threshold_Min': rate_t_min,
                'Rate_Threshold_Max': rate_t_max,
                'CWI_Min_Files': cwi_min_files,
                'CWI_Max_Files': cwi_max_files,
                'Rate_Min_Files': rate_min_files,
                'Rate_Max_Files': rate_max_files,
            }
            results.append(result_dict)
        
        return results, None
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return None, f"Error processing grouped file: {str(e)}"
def process_comparison(hdfc_filepath, icici_filepath):
    """Process HDFC vs ICICI comparison for grouped validation"""
    try:
        print("\n" + "="*60)
        print("🔄 COMPARISON MODE ACTIVATED")
        print("="*60)
        
        # Load both grouped validators
        hdfc_validator = get_validator('hdfc_grouped')
        icici_validator = get_validator('icici_grouped')
        
        if not hdfc_validator or not icici_validator:
            print("❌ Comparison models not available")
            return None, "Comparison models not available. Please ensure both HDFC and ICICI grouped models are loaded."
        
        print("✅ Both validators loaded successfully")
        
        # Validate both files
        print("🔍 Validating HDFC file...")
        hdfc_results, hdfc_error = validate_grouped_file(hdfc_validator, hdfc_filepath)
        
        print("🔍 Validating ICICI file...")
        icici_results, icici_error = validate_grouped_file(icici_validator, icici_filepath)
        
        if hdfc_error:
            print(f"❌ HDFC validation error: {hdfc_error}")
            return None, f"HDFC validation error: {hdfc_error}"
        
        if icici_error:
            print(f"❌ ICICI validation error: {icici_error}")
            return None, f"ICICI validation error: {icici_error}"
        
        print(f"✅ HDFC: {len(hdfc_results)} records validated")
        print(f"✅ ICICI: {len(icici_results)} records validated")
        
        # Convert to DataFrames
        hdfc_df = pd.DataFrame(hdfc_results)
        icici_df = pd.DataFrame(icici_results)
        
        # Format currency values for HDFC
        print("💰 Formatting HDFC currency values...")
        hdfc_df['As_Per_CWI_Amount'] = hdfc_df['As_Per_CWI_Amount'].apply(
            lambda x: hdfc_validator.format_indian_currency(x)
        )
        hdfc_df['Rate_per_sqfeet'] = hdfc_df['Rate_per_sqfeet'].apply(
            lambda x: hdfc_validator.format_indian_currency(x)
        )
        hdfc_df['CWI_Weighted_Min'] = hdfc_df['CWI_Weighted_Min'].apply(
            lambda x: hdfc_validator.format_indian_currency(x) if x != 'NA' and not pd.isna(x) else 'NA'
        )
        hdfc_df['CWI_Weighted_Max'] = hdfc_df['CWI_Weighted_Max'].apply(
            lambda x: hdfc_validator.format_indian_currency(x) if x != 'NA' and not pd.isna(x) else 'NA'
        )
        hdfc_df['Rate_Weighted_Min'] = hdfc_df['Rate_Weighted_Min'].apply(
            lambda x: hdfc_validator.format_indian_currency(x) if x != 'NA' and not pd.isna(x) else 'NA'
        )
        hdfc_df['Rate_Weighted_Max'] = hdfc_df['Rate_Weighted_Max'].apply(
            lambda x: hdfc_validator.format_indian_currency(x) if x != 'NA' and not pd.isna(x) else 'NA'
        )
        
        # Format currency values for ICICI
        print("💰 Formatting ICICI currency values...")
        icici_df['As_Per_CWI_Amount'] = icici_df['As_Per_CWI_Amount'].apply(
            lambda x: icici_validator.format_indian_currency(x)
        )
        icici_df['Rate_per_sqfeet'] = icici_df['Rate_per_sqfeet'].apply(
            lambda x: icici_validator.format_indian_currency(x)
        )
        icici_df['CWI_Weighted_Min'] = icici_df['CWI_Weighted_Min'].apply(
            lambda x: icici_validator.format_indian_currency(x) if x != 'NA' and not pd.isna(x) else 'NA'
        )
        icici_df['CWI_Weighted_Max'] = icici_df['CWI_Weighted_Max'].apply(
            lambda x: icici_validator.format_indian_currency(x) if x != 'NA' and not pd.isna(x) else 'NA'
        )
        icici_df['Rate_Weighted_Min'] = icici_df['Rate_Weighted_Min'].apply(
            lambda x: icici_validator.format_indian_currency(x) if x != 'NA' and not pd.isna(x) else 'NA'
        )
        icici_df['Rate_Weighted_Max'] = icici_df['Rate_Weighted_Max'].apply(
            lambda x: icici_validator.format_indian_currency(x) if x != 'NA' and not pd.isna(x) else 'NA'
        )
        
        # Format Is_Valid column
        hdfc_df['Is_Valid'] = hdfc_df['Is_Valid'].apply(lambda x: '✅ Valid' if x else '❌ Invalid')
        icici_df['Is_Valid'] = icici_df['Is_Valid'].apply(lambda x: '✅ Valid' if x else '❌ Invalid')
        
        # Save comparison reports
        timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
        hdfc_report = f'hdfc_comparison_{timestamp}.xlsx'
        icici_report = f'icici_comparison_{timestamp}.xlsx'
        
        hdfc_df.to_excel(os.path.join(app.config['OUTPUT_FOLDER'], hdfc_report), index=False)
        icici_df.to_excel(os.path.join(app.config['OUTPUT_FOLDER'], icici_report), index=False)
        
        print(f"✅ Reports saved: {hdfc_report}, {icici_report}")
        print("="*60 + "\n")
        
        # Convert to HTML for display
        hdfc_html = hdfc_df.to_html(classes='table table-striped', index=False, escape=False)
        icici_html = icici_df.to_html(classes='table table-striped', index=False, escape=False)
        
        return {
            'hdfc_table': hdfc_html,
            'icici_table': icici_html,
            'hdfc_report': hdfc_report,
            'icici_report': icici_report
        }, None
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"❌ Comparison error: {str(e)}")
        return None, f'Comparison error: {str(e)}'

def process_qar_analysis(hdfc_filepath, icici_filepath):
    """Process QAR (Qty/Amount/Rate) Analysis between ICICI and HDFC"""
    try:
        print("\n" + "="*60)
        print("🔄 QAR ANALYSIS MODE ACTIVATED")
        print("="*60)
        
        # ✅ FIX: Load ALL three validators for each bank
        hdfc_qty_validator = get_validator('hdfc_qty')
        hdfc_amount_validator = get_validator('hdfc_amount')
        hdfc_rate_validator = get_validator('hdfc_rate')

        icici_qty_validator = get_validator('icici_qty')
        icici_amount_validator = get_validator('icici_amount')
        icici_rate_validator = get_validator('icici_rate')
        
        if not all([hdfc_qty_validator, hdfc_amount_validator, hdfc_rate_validator,
                    icici_qty_validator, icici_amount_validator, icici_rate_validator]):
            return None, "QAR analysis models not available. Please ensure qty, amount, and rate models are loaded for both banks."
        
        print("✅ All QAR validators loaded successfully")
        
        # Read both Excel files
        hdfc_df = pd.read_excel(hdfc_filepath, sheet_name='Extracted Data')
        icici_df = pd.read_excel(icici_filepath, sheet_name='Extracted Data')
        
        # Extract sq feet ranges
        hdfc_sqft_qty = hdfc_qty_validator.extract_sqfeet_from_filename(os.path.basename(hdfc_filepath), mode='quantity')
        hdfc_sqft_amt = hdfc_amount_validator.extract_sqfeet_from_filename(os.path.basename(hdfc_filepath), mode='amount')
        hdfc_sqft_rate = hdfc_rate_validator.extract_sqfeet_from_filename(os.path.basename(hdfc_filepath), mode='rate')
        
        icici_sqft_qty = icici_qty_validator.extract_sqfeet_from_filename(os.path.basename(icici_filepath), mode='quantity')
        icici_sqft_amt = icici_amount_validator.extract_sqfeet_from_filename(os.path.basename(icici_filepath), mode='amount')
        icici_sqft_rate = icici_rate_validator.extract_sqfeet_from_filename(os.path.basename(icici_filepath), mode='rate')
        
        # Find vitrified tiles 800x800 particular
        # ✅ NEW: Find all 10 matching particulars
        particulars_to_find = [
            ('Vitrified Tiles 800x800', find_vitrified_800x800),
            ('Siporex 9 inch', find_siporex_9inch),
            ('R.C.C Cement Concrete', find_rcc_cement_concrete),
            ('Aluminum Framework BWP', find_aluminum_framework_bwp),
            ('Ceramic Tiles 300x600', find_ceramic_tiles_300x600),
            ('Aluminum Composite Panel', find_aluminum_composite_panel),
            ('12mm Plaster 1:4', find_12mm_plaster_1_4),
            ('Service Counter Table', find_service_counter_table),
            ('Full Height Partition BWR', find_full_height_partition_bwr),
            ('Gypsum False Ceiling', find_gypsum_false_ceiling),
            ('Plaster of Paris Punning', find_plaster_of_paris_punning),
            ('Acrylic Emulsion Paint', find_acrylic_emulsion_paint),  # ✅ NEW
            ('Modular Ceiling', find_modular_ceiling) 
        ]
        
        # Around line 1100-1200, replace this section:

        comparison_results = []
        icici_only_results = []

        for particular_name, find_function in particulars_to_find:
            print(f"🔍 Searching for: {particular_name}")

            icici_row = find_function(icici_df, 'icici')
            hdfc_row = find_function(hdfc_df, 'hdfc')

            if icici_row is None:
                print(f"  ⚠️  Not found in ICICI: {particular_name}")
                continue

            print(f"  ✅ Found in ICICI: {icici_row['Particulars']}")

            # Part 1: ICICI vs HDFC Comparison
            if hdfc_row is not None:
                print(f"  ✅ Found in HDFC: {hdfc_row['Particulars']}")

                comparison_result = compare_icici_with_other(
                    icici_row, hdfc_row, 
                    icici_qty_validator, icici_amount_validator, icici_rate_validator,
                    hdfc_qty_validator, hdfc_amount_validator, hdfc_rate_validator,
                    icici_sqft_qty, icici_sqft_amt, icici_sqft_rate,
                    hdfc_sqft_qty, hdfc_sqft_amt, hdfc_sqft_rate
                )
                comparison_result['Particular_Category'] = particular_name
                comparison_results.append(comparison_result)
            else:
                print(f"  ⚠️  Not found in HDFC: {particular_name}")

            # Part 2: ICICI Bank Only Analysis - THIS MUST RUN FOR EVERY ICICI ROW
            icici_only_result = analyze_icici_only(
                icici_row, 
                icici_qty_validator, icici_amount_validator, icici_rate_validator,
                icici_sqft_qty, icici_sqft_amt, icici_sqft_rate
            )
            icici_only_result['Particular_Category'] = particular_name
            icici_only_results.append(icici_only_result)  # ✅ ADD TO LIST, NOT OVERWRITE

        # Final check
        if not comparison_results and not icici_only_results:
            return None, "Could not find any matching particulars in the files"

        return {
            'comparison_results': comparison_results,
            'icici_only_results': icici_only_results,  # ✅ CHANGE FROM icici_only_result
            'hdfc_file': os.path.basename(hdfc_filepath),
            'icici_file': os.path.basename(icici_filepath)
        }, None
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return None, f"QAR Analysis error: {str(e)}"

def find_vitrified_800x800(df, bank_type):
    """Find the vitrified tiles 800x800 particular from dataframe"""
    # Map column names
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    # Search for 800x800 vitrified tiles
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        # Check for 800x800 pattern
        has_size = bool(re.search(r'800\s*[x*×]\s*800', particular))
        has_vitrified = 'vitrified' in particular and 'tiles' in particular
        
        if has_size and has_vitrified:
            return {
                'Sr.No': idx + 1,
                'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                'Qty': row.get(col_mapping.get('Qty', ''), 0),
                'Amount': row.get(col_mapping.get('Amount', ''), 0),
                'Rate': row.get(col_mapping.get('Rate', ''), 0),
                'Unit': row.get(col_mapping.get('Unit', ''), '')
            }
    
    return None

def find_siporex_9inch(df, bank_type):
    """Find Siporex 9 inch particular from dataframe"""
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        if bank_type == 'icici':
            # Check for 'siporex' AND ('9inch' OR '9 inch')
            has_siporex = 'siporex' in particular
            has_9inch = bool(re.search(r'9\s*inch', particular))
            
            if has_siporex and has_9inch:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
        else:  # hdfc
            # Check for '150mm' AND 'siporex'
            has_150mm = bool(re.search(r'150\s*mm', particular))
            has_siporex = 'siporex' in particular
            
            if has_150mm and has_siporex:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
    
    return None

def find_rcc_cement_concrete(df, bank_type):
    """Find R.C.C Work with providing and casting"""
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        if bank_type == 'icici':
            # Check for 'r.c.c' or 'rcc' AND 'work' AND 'providing' AND 'casting'
            has_rcc = bool(re.search(r'r\.?\s*c\.?\s*c', particular))
            has_work = 'work' in particular
            has_providing = 'providing' in particular
            has_casting = 'casting' in particular
            
            if has_rcc and has_work and has_providing and has_casting:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
        else:  # hdfc
            # Check for 'r.c.c' or 'rcc' AND 'work' AND 'providing' AND ('casting' OR '&')
            has_rcc = bool(re.search(r'r\.?\s*c\.?\s*c', particular))
            has_work = 'work' in particular
            has_providing = 'providing' in particular
            # Accept both 'casting' and '&' (ampersand often used as "and")
            has_casting_or_ampersand = 'casting' in particular or '&' in particular
            
            if has_rcc and has_work and has_providing and has_casting_or_ampersand:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
    
    return None

def find_aluminum_framework_bwp(df, bank_type):
    """Find Aluminum framework with BWP plywood"""
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        if bank_type == 'icici':
            # Check for 'wall' or 'column', 'cladding', '2 x 1.5', 'aluminum'/'aluminium'
            has_wall_column = 'wall' in particular or 'column' in particular
            has_cladding = 'cladding' in particular
            has_2x15 = bool(re.search(r'2\s*[x×]\s*1\.?5', particular))
            has_aluminum = 'aluminum' in particular or 'aluminium' in particular
            
            if has_wall_column and has_cladding and has_2x15 and has_aluminum:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
        else:  # hdfc
            # Check for 'aluminium'/'aluminum', 'frame', 'bwp', 'plywood'
            has_aluminum = 'aluminum' in particular or 'aluminium' in particular
            has_frame = 'frame' in particular
            has_bwp = 'bwp' in particular
            has_plywood = 'plywood' in particular
            
            if has_aluminum and has_frame and has_bwp and has_plywood:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
    
    return None

def find_ceramic_tiles_300x600(df, bank_type):
    """Find ceramic tiles 300x600mm"""
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        if bank_type == 'icici':
            # Check for 'ceramic tiles' AND '300x600' or '300 x 600'
            has_ceramic = 'ceramic' in particular and 'tiles' in particular
            has_300x600 = bool(re.search(r'300\s*[x×]\s*600', particular))
            
            if has_ceramic and has_300x600:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
        else:  # hdfc
            # Check for '600mm x 600mm' or '600 x 600' AND 'vitrified'
            has_600x600 = bool(re.search(r'600\s*(?:mm)?\s*[x×]\s*600', particular))
            has_vitrified = 'vitrified' in particular
            
            if has_600x600 and has_vitrified:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
    
    return None

def find_aluminum_composite_panel(df, bank_type):
    """Find Aluminum Composite Panel"""
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        if bank_type == 'icici':
            # ✅ NEW: Check for 'exterior grade' AND '3mm' AND 'thick' AND 'aluminum' AND 'composite' AND 'panel'
            # ❌ IGNORE if contains 'silver grey' or 'silver' AND 'grey'
            has_exterior = 'exterior' in particular and 'grade' in particular
            has_3mm = '3mm' in particular or '3 mm' in particular
            has_thick = 'thick' in particular
            has_aluminum = 'aluminum' in particular or 'aluminium' in particular
            has_composite = 'composite' in particular
            has_panel = 'panel' in particular
            
            # Check if it contains "silver grey" or "silver" + "grey"
            has_silver_grey = ('silver' in particular and 'grey' in particular) or 'silver grey' in particular or 'silvergrey' in particular
            
            # Match if all required keywords present AND does NOT have silver grey
            if has_exterior and has_3mm and has_thick and has_aluminum and has_composite and has_panel and not has_silver_grey:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
        else:  # hdfc
            # Check for 'providing' AND 'fixing' AND 'tricolour'
            has_providing = 'providing' in particular
            has_fixing = 'fixing' in particular
            has_tricolour = 'tricolour' in particular or 'tricolor' in particular
            
            if has_providing and has_fixing and has_tricolour:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
    
    return None

def find_12mm_plaster_1_4(df, bank_type):
    """Find 12mm thick plaster 1:4"""
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        if bank_type == 'icici':
            # Check for '12mm' OR '12 mm' AND 'plaster' AND '1:4'
            has_12mm = bool(re.search(r'12\s*mm', particular))
            has_plaster = 'plaster' in particular
            has_1_4 = '1:4' in particular or '1 : 4' in particular
            
            if has_12mm and has_plaster and has_1_4:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
        else:  # hdfc
            # Check for '1:4' AND 'cement plaster' AND '12mm' OR '12 mm'
            has_1_4 = '1:4' in particular or '1 : 4' in particular
            has_cement_plaster = 'cement' in particular and 'plaster' in particular
            has_12mm = bool(re.search(r'12\s*mm', particular))
            
            if has_1_4 and has_cement_plaster and has_12mm:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
    
    return None

def find_service_counter_table(df, bank_type):
    """Find modular service counter table"""
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        if bank_type == 'icici':
            # Check for 'modular service counter table' AND '0.75 inch'
            has_modular = 'modular' in particular
            has_service_counter = 'service' in particular and 'counter' in particular
            has_table = 'table' in particular
            has_075 = bool(re.search(r'0\.75\s*inch', particular))
            
            if has_modular and has_service_counter and has_table and has_075:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
        else:  # hdfc
            # Check for "branch manager's table" OR "branch manager table" AND '1350 x 675' or '1350x675'
            has_branch_manager = 'branch' in particular and 'manager' in particular and 'table' in particular
            has_1350x675 = bool(re.search(r'1350\s*[x×]\s*675', particular))
            
            if has_branch_manager and has_1350x675:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
    
    return None

def find_full_height_partition_bwr(df, bank_type):
    """Find full height partitions with BWR plywood"""
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        if bank_type == 'icici':
            # Check for 'full height partitions' AND '2 x 1.5'
            has_full_height = 'full' in particular and 'height' in particular
            has_partitions = 'partition' in particular
            has_2x15 = bool(re.search(r'2\s*[x×]\s*1\.?5', particular))
            
            if has_full_height and has_partitions and has_2x15:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
        else:  # hdfc
            # Check for 'full height' AND 'solid partition' AND 'bwr plywood'
            has_full_height = 'full' in particular and 'height' in particular
            has_solid_partition = 'solid' in particular and 'partition' in particular
            has_bwr = 'bwr' in particular
            has_plywood = 'plywood' in particular
            
            if has_full_height and has_solid_partition and has_bwr and has_plywood:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
    
    return None

def find_gypsum_false_ceiling(df, bank_type):
    """Find Gypsum False Ceiling"""
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        if bank_type == 'icici':
            # Check for 'gypsum' AND 'false ceiling'
            has_gypsum = 'gypsum' in particular
            has_false_ceiling = 'false' in particular and 'ceiling' in particular
            
            if has_gypsum and has_false_ceiling:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
        else:  # hdfc
            # Check for 'gypsum' AND 'board' AND 'ceiling'
            has_gypsum = 'gypsum' in particular
            has_board = 'board' in particular
            has_ceiling = 'ceiling' in particular
            
            if has_gypsum and has_board and has_ceiling:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
    
    return None

def find_plaster_of_paris_punning(df, bank_type):
    """Find Plaster of Paris punning"""
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        # Check for 'plaster of paris' AND 'punning'
        has_pop = 'plaster' in particular and 'paris' in particular
        has_punning = 'punning' in particular
        
        if has_pop and has_punning:
            return {
                'Sr.No': idx + 1,
                'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                'Qty': row.get(col_mapping.get('Qty', ''), 0),
                'Amount': row.get(col_mapping.get('Amount', ''), 0),
                'Rate': row.get(col_mapping.get('Rate', ''), 0),
                'Unit': row.get(col_mapping.get('Unit', ''), '')
            }
    
    return None

def find_acrylic_emulsion_paint(df, bank_type):
    """Find acrylic emulsion paint particular from dataframe"""
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        if bank_type == 'icici':
            # Check for 'acrylic' AND 'emulsion' AND 'paint'
            has_acrylic = 'acrylic' in particular
            has_emulsion = 'emulsion' in particular
            has_paint = 'paint' in particular
            
            if has_acrylic and has_emulsion and has_paint:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
        else:  # hdfc
            # Check for 'applying' AND 'acrylic' AND 'emulsion' AND 'paint'
            has_applying = 'applying' in particular
            has_acrylic = 'acrylic' in particular
            has_emulsion = 'emulsion' in particular
            has_paint = 'paint' in particular
            
            if has_applying and has_acrylic and has_emulsion and has_paint:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
    
    return None

def find_modular_ceiling(df, bank_type):
    """Find modular ceiling particular from dataframe"""
    col_mapping = {}
    for col in df.columns:
        clean_col = str(col).lower().replace('.', '').replace(' ', '').replace('(', '').replace(')', '').replace('-', '')
        if 'particular' in clean_col:
            col_mapping['Particulars'] = col
        elif 'qty' in clean_col and 'cwi' in clean_col:
            col_mapping['Qty'] = col
        elif 'amount' in clean_col and 'cwi' in clean_col:
            col_mapping['Amount'] = col
        elif 'rate' in clean_col:
            col_mapping['Rate'] = col
        elif 'unit' in clean_col:
            col_mapping['Unit'] = col
    
    for idx, row in df.iterrows():
        particular = str(row.get(col_mapping.get('Particulars', ''), '')).lower()
        
        if bank_type == 'icici':
            # Check for 'modular' AND 'ceiling' (case-insensitive)
            has_modular = 'modular' in particular
            has_ceiling = 'ceiling' in particular
            
            if has_modular and has_ceiling:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
        else:  # hdfc
            # Check for 'grid' AND 'modular' AND 'false' AND 'ceiling'
            has_grid = 'grid' in particular
            has_modular = 'modular' in particular
            has_false = 'false' in particular
            has_ceiling = 'ceiling' in particular
            
            if has_grid and has_modular and has_false and has_ceiling:
                return {
                    'Sr.No': idx + 1,
                    'Particulars': str(row.get(col_mapping.get('Particulars', ''), '')),
                    'Qty': row.get(col_mapping.get('Qty', ''), 0),
                    'Amount': row.get(col_mapping.get('Amount', ''), 0),
                    'Rate': row.get(col_mapping.get('Rate', ''), 0),
                    'Unit': row.get(col_mapping.get('Unit', ''), '')
                }
    
    return None


def compare_icici_with_other(icici_row, hdfc_row, 
                             icici_qty_val, icici_amt_val, icici_rate_val,
                             hdfc_qty_val, hdfc_amt_val, hdfc_rate_val,
                             icici_sqft_qty, icici_sqft_amt, icici_sqft_rate,
                             hdfc_sqft_qty, hdfc_sqft_amt, hdfc_sqft_rate):
    """Compare ICICI with HDFC for the particular using separate validators"""
    
    result = {
        'Sr.No': 1,
        'Particular_Category': 'N/A', 
        'ICICI_Particulars': icici_row['Particulars'],
        'ICICI_Qty': icici_qty_val.extract_quantity(icici_row['Qty']),
        'ICICI_Amount': icici_amt_val.extract_amount(icici_row['Amount']),
        'ICICI_Rate': icici_rate_val.extract_amount(icici_row['Rate']),
        'Other_Bank_Particulars': hdfc_row['Particulars']
    }
    
    hdfc_raw_particular = hdfc_row['Particulars']
    hdfc_cleaned = hdfc_qty_val.clean_text(hdfc_raw_particular)
    
    # ===== QUANTITY VALIDATION =====
    if hdfc_sqft_qty in hdfc_qty_val.sqfeet_ranges:
        range_data = hdfc_qty_val.sqfeet_ranges[hdfc_sqft_qty]
        range_key = 'quantity_ranges' if 'quantity_ranges' in range_data else 'rate_ranges' if 'rate_ranges' in range_data else 'amount_ranges'

        qty_stats = None
        if range_key in range_data:
            if hdfc_cleaned in range_data[range_key]:
                qty_stats = range_data[range_key][hdfc_cleaned]
            else:
                sim_name, sim_score = hdfc_qty_val.find_similar_particular(hdfc_raw_particular, hdfc_sqft_qty)
                if sim_name:
                    qty_stats = range_data[range_key][sim_name]

        if qty_stats:
            # 1. Calculate Weighted Stats
            weighted_min = qty_stats.get('weighted_avg_min', 0)
            weighted_max = qty_stats.get('weighted_avg_max', 0)
            
            # 2. Calculate Threshold (Overall) Stats
            overall_min = qty_stats.get('overall_min', 0)
            overall_max = qty_stats.get('overall_max', 0)

            # 3. Format Strings
            weighted_str = f"{weighted_min:.2f} - {weighted_max:.2f}"
            threshold_str = f"{overall_min:.2f} - {overall_max:.2f}"

            if result['ICICI_Qty'] == 0:
                # ✅ Zero Case: Valid, show ranges
                result['Is_Valid_Qty_Other'] = True
                zero_count = qty_stats.get('zero_count', 0)
                total = qty_stats.get('total_count', 1)
                result['Validation_Message_Qty'] = f"✅ Zero allowed ({zero_count}/{total} cases)"
                result['Weighted_Avg_Range_Qty'] = weighted_str
                result['Threshold_Range_Qty'] = threshold_str 
            else:
                if weighted_min == 0 and weighted_max == 0:
                    result['Is_Valid_Qty_Other'] = 'N/A'
                    result['Weighted_Avg_Range_Qty'] = 'N/A'
                    result['Threshold_Range_Qty'] = 'N/A'
                    result['Validation_Message_Qty'] = 'N/A - Zero data only'
                else:
                    result['Is_Valid_Qty_Other'] = weighted_min <= result['ICICI_Qty'] <= weighted_max
                    result['Weighted_Avg_Range_Qty'] = weighted_str
                    result['Threshold_Range_Qty'] = threshold_str 

                    if result['Is_Valid_Qty_Other']:
                        result['Validation_Message_Qty'] = f"✅ Within range: {weighted_str}"
                    else:
                        result['Validation_Message_Qty'] = f"❌ Outside range: {weighted_str}"
        else:
            result['Is_Valid_Qty_Other'] = 'N/A'
            result['Weighted_Avg_Range_Qty'] = 'N/A'
            result['Threshold_Range_Qty'] = 'N/A'
            result['Validation_Message_Qty'] = 'N/A - No match'
    else:
        result['Is_Valid_Qty_Other'] = 'N/A'
        result['Weighted_Avg_Range_Qty'] = 'N/A'
        result['Threshold_Range_Qty'] = 'N/A'
        result['Validation_Message_Qty'] = 'N/A - Range error'

    # ===== AMOUNT VALIDATION =====
    if hdfc_sqft_amt in hdfc_amt_val.sqfeet_ranges:
        range_data = hdfc_amt_val.sqfeet_ranges[hdfc_sqft_amt]
        range_key = 'amount_ranges' if 'amount_ranges' in range_data else 'rate_ranges' if 'rate_ranges' in range_data else 'quantity_ranges'

        amt_stats = None
        if range_key in range_data:
            if hdfc_cleaned in range_data[range_key]:
                amt_stats = range_data[range_key][hdfc_cleaned]
            else:
                sim_name, sim_score = hdfc_amt_val.find_similar_particular(hdfc_raw_particular, hdfc_sqft_amt)
                if sim_name:
                    amt_stats = range_data[range_key][sim_name]

        if amt_stats:
            # 1. Calculate Weighted Stats
            weighted_min = amt_stats.get('weighted_avg_min', 0)
            weighted_max = amt_stats.get('weighted_avg_max', 0)
            
            # 2. Calculate Threshold (Overall) Stats
            overall_min = amt_stats.get('overall_min', 0)
            overall_max = amt_stats.get('overall_max', 0)

            # 3. Format Strings
            weighted_str = f"₹{hdfc_amt_val.format_indian_currency(weighted_min)} - ₹{hdfc_amt_val.format_indian_currency(weighted_max)}"
            threshold_str = f"₹{hdfc_amt_val.format_indian_currency(overall_min)} - ₹{hdfc_amt_val.format_indian_currency(overall_max)}"

            if result['ICICI_Amount'] == 0:
                # ✅ Zero Case
                result['Is_Valid_Amount_Other'] = True
                zero_count = amt_stats.get('zero_count', 0)
                total = amt_stats.get('total_count', 1)
                result['Validation_Message_Amount'] = f"✅ Zero allowed ({zero_count}/{total} cases)"
                result['Weighted_Avg_Range_Amount'] = weighted_str
                result['Threshold_Range_Amount'] = threshold_str
            else:
                if weighted_min == 0 and weighted_max == 0:
                    result['Is_Valid_Amount_Other'] = 'N/A'
                    result['Weighted_Avg_Range_Amount'] = 'N/A'
                    result['Threshold_Range_Amount'] = 'N/A'
                    result['Validation_Message_Amount'] = 'N/A - Zero data'
                else:
                    result['Is_Valid_Amount_Other'] = weighted_min <= result['ICICI_Amount'] <= weighted_max
                    result['Weighted_Avg_Range_Amount'] = weighted_str
                    result['Threshold_Range_Amount'] = threshold_str

                    if result['Is_Valid_Amount_Other']:
                        result['Validation_Message_Amount'] = f"✅ Valid"
                    else:
                        result['Validation_Message_Amount'] = f"❌ Invalid"
        else:
            result['Is_Valid_Amount_Other'] = 'N/A'
            result['Weighted_Avg_Range_Amount'] = 'N/A'
            result['Threshold_Range_Amount'] = 'N/A'
            result['Validation_Message_Amount'] = 'N/A - No match'
    else:
        result['Is_Valid_Amount_Other'] = 'N/A'
        result['Weighted_Avg_Range_Amount'] = 'N/A'
        result['Threshold_Range_Amount'] = 'N/A'
        result['Validation_Message_Amount'] = 'N/A - Range error'
    
    # ===== RATE VALIDATION =====
    if hdfc_sqft_rate in hdfc_rate_val.sqfeet_ranges:
        range_data = hdfc_rate_val.sqfeet_ranges[hdfc_sqft_rate]
        range_key = 'rate_ranges' if 'rate_ranges' in range_data else 'amount_ranges' if 'amount_ranges' in range_data else 'quantity_ranges'

        rate_stats = None
        if range_key in range_data:
            if hdfc_cleaned in range_data[range_key]:
                rate_stats = range_data[range_key][hdfc_cleaned]
            else:
                sim_name, sim_score = hdfc_rate_val.find_similar_particular(hdfc_raw_particular, hdfc_sqft_rate)
                if sim_name:
                    rate_stats = range_data[range_key][sim_name]

        if rate_stats:
            # 1. Calculate Weighted Stats
            weighted_min = rate_stats.get('weighted_avg_min', 0)
            weighted_max = rate_stats.get('weighted_avg_max', 0)
            
            # 2. Calculate Threshold (Overall) Stats
            overall_min = rate_stats.get('overall_min', 0)
            overall_max = rate_stats.get('overall_max', 0)

            # 3. Format Strings
            weighted_str = f"₹{hdfc_rate_val.format_indian_currency(weighted_min)} - ₹{hdfc_rate_val.format_indian_currency(weighted_max)}"
            threshold_str = f"₹{hdfc_rate_val.format_indian_currency(overall_min)} - ₹{hdfc_rate_val.format_indian_currency(overall_max)}"

            if result['ICICI_Rate'] == 0:
                # ✅ Zero Case
                result['Is_Valid_Rate_Other'] = True
                zero_count = rate_stats.get('zero_count', 0)
                total = rate_stats.get('total_count', 1)
                result['Validation_Message_Rate'] = f"✅ Zero allowed ({zero_count}/{total} cases)"
                result['Weighted_Avg_Range_Rate'] = weighted_str
                result['Threshold_Range_Rate'] = threshold_str
            else:
                if weighted_min == 0 and weighted_max == 0:
                    result['Is_Valid_Rate_Other'] = 'N/A'
                    result['Weighted_Avg_Range_Rate'] = 'N/A'
                    result['Threshold_Range_Rate'] = 'N/A'
                    result['Validation_Message_Rate'] = 'N/A - Zero data'
                else:
                    result['Is_Valid_Rate_Other'] = weighted_min <= result['ICICI_Rate'] <= weighted_max
                    result['Weighted_Avg_Range_Rate'] = weighted_str
                    result['Threshold_Range_Rate'] = threshold_str

                    if result['Is_Valid_Rate_Other']:
                        result['Validation_Message_Rate'] = f"✅ Valid"
                    else:
                        result['Validation_Message_Rate'] = f"❌ Invalid"
        else:
            result['Is_Valid_Rate_Other'] = 'N/A'
            result['Weighted_Avg_Range_Rate'] = 'N/A'
            result['Threshold_Range_Rate'] = 'N/A'
            result['Validation_Message_Rate'] = 'N/A - No match'
    else:
        result['Is_Valid_Rate_Other'] = 'N/A'
        result['Weighted_Avg_Range_Rate'] = 'N/A'
        result['Threshold_Range_Rate'] = 'N/A'
        result['Validation_Message_Rate'] = 'N/A - Range error'

    return result
def analyze_icici_only(icici_row, 
                       icici_qty_val, icici_amt_val, icici_rate_val,
                       icici_sqft_qty, icici_sqft_amt, icici_sqft_rate):
    """Analyze ICICI data against its own models (qty, amount, rate separately)"""
    
    result = {
        'Sr.No': 1,
        'Particular_Category': 'N/A', 
        'Particulars': icici_row['Particulars'],
        'Qty': icici_qty_val.extract_quantity(icici_row['Qty']),
        'Amount': icici_amt_val.extract_amount(icici_row['Amount']),
        'Rate': icici_rate_val.extract_amount(icici_row['Rate'])
    }
    
    icici_raw_particular = icici_row['Particulars']
    icici_cleaned = icici_qty_val.clean_text(icici_raw_particular)
    
    # ===== QUANTITY (using QTY model) =====
    if icici_sqft_qty in icici_qty_val.sqfeet_ranges:
        range_data = icici_qty_val.sqfeet_ranges[icici_sqft_qty]
        range_key = 'quantity_ranges' if 'quantity_ranges' in range_data else 'rate_ranges' if 'rate_ranges' in range_data else 'amount_ranges'

        qty_stats = None
        if range_key in range_data:
            if icici_cleaned in range_data[range_key]:
                qty_stats = range_data[range_key][icici_cleaned]
            else:
                sim_name, sim_score = icici_qty_val.find_similar_particular(icici_raw_particular, icici_sqft_qty)
                if sim_name:
                    qty_stats = range_data[range_key][sim_name]

        if qty_stats:
            # 1. Calculate stats FIRST
            weighted_min = qty_stats.get('weighted_avg_min', qty_stats.get('overall_mean', 0))
            weighted_max = qty_stats.get('weighted_avg_max', qty_stats.get('overall_mean', 0))
            
            # Format Strings
            weighted_str = f"{weighted_min:.2f} - {weighted_max:.2f}"
            threshold_str = f"{qty_stats.get('overall_min', 0):.2f} - {qty_stats.get('overall_max', 0):.2f}"

            if result['Qty'] == 0:
                # ✅ Zero Case: Show Ranges
                result['Is_Valid_Qty'] = True
                zero_count = qty_stats.get('zero_count', 0)
                total = qty_stats.get('total_count', 1)
                result['Validation_Message_Qty'] = f"✅ Zero allowed ({zero_count}/{total} cases)"
                result['Threshold_Range_Qty'] = threshold_str 
                result['Weighted_Avg_Range_Qty'] = weighted_str 
            else:
                if weighted_min == 0 and weighted_max == 0:
                    result['Is_Valid_Qty'] = 'N/A'
                    result['Threshold_Range_Qty'] = 'N/A'
                    result['Weighted_Avg_Range_Qty'] = 'N/A'
                    result['Validation_Message_Qty'] = 'N/A - Zero data'
                else:
                    result['Is_Valid_Qty'] = weighted_min <= result['Qty'] <= weighted_max
                    result['Threshold_Range_Qty'] = threshold_str
                    result['Weighted_Avg_Range_Qty'] = weighted_str

                    # ✅ FIX: Added range string to message
                    if result['Is_Valid_Qty']:
                        result['Validation_Message_Qty'] = f"✅ Within range: {weighted_str}"
                    else:
                        result['Validation_Message_Qty'] = f"❌ Outside range: {weighted_str}"
        else:
            result['Is_Valid_Qty'] = 'N/A'
            result['Threshold_Range_Qty'] = 'N/A'
            result['Weighted_Avg_Range_Qty'] = 'N/A'
            result['Validation_Message_Qty'] = 'N/A - No match'
    else:
        result['Is_Valid_Qty'] = 'N/A'
        result['Threshold_Range_Qty'] = 'N/A'
        result['Weighted_Avg_Range_Qty'] = 'N/A'
        result['Validation_Message_Qty'] = 'N/A - Range error'
    
    # ===== AMOUNT (using AMOUNT model) =====
    if icici_sqft_amt in icici_amt_val.sqfeet_ranges:
        range_data = icici_amt_val.sqfeet_ranges[icici_sqft_amt]
        range_key = 'amount_ranges' if 'amount_ranges' in range_data else 'rate_ranges' if 'rate_ranges' in range_data else 'quantity_ranges'

        amt_stats = None
        if range_key in range_data:
            if icici_cleaned in range_data[range_key]:
                amt_stats = range_data[range_key][icici_cleaned]
            else:
                sim_name, sim_score = icici_amt_val.find_similar_particular(icici_raw_particular, icici_sqft_amt)
                if sim_name:
                    amt_stats = range_data[range_key][sim_name]

        if amt_stats:
            # 1. Calculate stats FIRST
            weighted_min = amt_stats.get('weighted_avg_min', amt_stats.get('overall_mean', 0))
            weighted_max = amt_stats.get('weighted_avg_max', amt_stats.get('overall_mean', 0))
            
            # Format Strings
            weighted_str = f"₹{icici_amt_val.format_indian_currency(weighted_min)} - ₹{icici_amt_val.format_indian_currency(weighted_max)}"
            threshold_str = f"₹{icici_amt_val.format_indian_currency(amt_stats.get('overall_min', 0))} - ₹{icici_amt_val.format_indian_currency(amt_stats.get('overall_max', 0))}"

            if result['Amount'] == 0:
                # ✅ Zero Case: Show Ranges
                result['Is_Valid_Amount'] = True
                result['Validation_Message_Amount'] = f"✅ Zero allowed"
                result['Threshold_Range_Amount'] = threshold_str 
                result['Weighted_Avg_Range_Amount'] = weighted_str
            else:
                if weighted_min == 0 and weighted_max == 0:
                    result['Is_Valid_Amount'] = 'N/A'
                    result['Threshold_Range_Amount'] = 'N/A'
                    result['Weighted_Avg_Range_Amount'] = 'N/A'
                    result['Validation_Message_Amount'] = 'N/A - Zero data'
                else:
                    result['Is_Valid_Amount'] = weighted_min <= result['Amount'] <= weighted_max
                    result['Threshold_Range_Amount'] = threshold_str
                    result['Weighted_Avg_Range_Amount'] = weighted_str

                    # ✅ FIX: Added range string to message
                    if result['Is_Valid_Amount']:
                        result['Validation_Message_Amount'] = f"✅ Within range: {weighted_str}"
                    else:
                        result['Validation_Message_Amount'] = f"❌ Outside range: {weighted_str}"
        else:
            result['Is_Valid_Amount'] = 'N/A'
            result['Threshold_Range_Amount'] = 'N/A'
            result['Weighted_Avg_Range_Amount'] = 'N/A'
            result['Validation_Message_Amount'] = 'N/A - No match'
    else:
        result['Is_Valid_Amount'] = 'N/A'
        result['Threshold_Range_Amount'] = 'N/A'
        result['Weighted_Avg_Range_Amount'] = 'N/A'
        result['Validation_Message_Amount'] = 'N/A - Range error'
    
    # ===== RATE (using RATE model) =====
    if icici_sqft_rate in icici_rate_val.sqfeet_ranges:
        range_data = icici_rate_val.sqfeet_ranges[icici_sqft_rate]
        range_key = 'rate_ranges' if 'rate_ranges' in range_data else 'amount_ranges' if 'amount_ranges' in range_data else 'quantity_ranges'
        
        rate_stats = None
        if range_key in range_data:
            if icici_cleaned in range_data[range_key]:
                rate_stats = range_data[range_key][icici_cleaned]
            else:
                sim_name, sim_score = icici_rate_val.find_similar_particular(icici_raw_particular, icici_sqft_rate)
                if sim_name:
                    rate_stats = range_data[range_key][sim_name]
            
        if rate_stats:
            # 1. Calculate stats FIRST
            weighted_min = rate_stats.get('weighted_avg_min', rate_stats.get('overall_mean', 0))
            weighted_max = rate_stats.get('weighted_avg_max', rate_stats.get('overall_mean', 0))
            
            # Format Strings
            weighted_str = f"₹{icici_rate_val.format_indian_currency(weighted_min)} - ₹{icici_rate_val.format_indian_currency(weighted_max)}"
            threshold_str = f"₹{icici_rate_val.format_indian_currency(rate_stats.get('overall_min', 0))} - ₹{icici_rate_val.format_indian_currency(rate_stats.get('overall_max', 0))}"

            if result['Rate'] == 0:
                # ✅ Zero Case: Show Ranges
                result['Is_Valid_Rate'] = True
                result['Validation_Message_Rate'] = f"✅ Zero allowed"
                result['Threshold_Range_Rate'] = threshold_str 
                result['Weighted_Avg_Range_Rate'] = weighted_str
            else:
                if weighted_min == 0 and weighted_max == 0:
                    result['Is_Valid_Rate'] = 'N/A'
                    result['Threshold_Range_Rate'] = 'N/A'
                    result['Weighted_Avg_Range_Rate'] = 'N/A'
                    result['Validation_Message_Rate'] = 'N/A - Zero data'
                else:
                    result['Is_Valid_Rate'] = weighted_min <= result['Rate'] <= weighted_max
                    result['Threshold_Range_Rate'] = threshold_str
                    result['Weighted_Avg_Range_Rate'] = weighted_str
                    
                    # ✅ FIX: Added range string to message
                    if result['Is_Valid_Rate']:
                        result['Validation_Message_Rate'] = f"✅ Within range: {weighted_str}"
                    else:
                        result['Validation_Message_Rate'] = f"❌ Outside range: {weighted_str}"
        else:
            result['Is_Valid_Rate'] = 'N/A'
            result['Threshold_Range_Rate'] = 'N/A'
            result['Weighted_Avg_Range_Rate'] = 'N/A'
            result['Validation_Message_Rate'] = 'N/A - No match'
    else:
        result['Is_Valid_Rate'] = 'N/A'
        result['Threshold_Range_Rate'] = 'N/A'
        result['Weighted_Avg_Range_Rate'] = 'N/A'
        result['Validation_Message_Rate'] = 'N/A - Range error'
    
    return result
@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if not MODEL_PATHS:
        flash("No validation models are configured.")
        return redirect(url_for('upload_file'))

    if request.method == 'POST':
        print("\n" + "="*60)
        print("🔵 UPLOAD REQUEST RECEIVED")
        print("="*60)
        
        # ✅ STEP 1: Check selected clients FIRST
        selected_clients = request.form.get('selected_clients', '')
        print(f"📋 Selected clients: {selected_clients}")

        # Get selected region (new)
        selected_region = request.form.get('selected_region', '')
        print(f"📍 Selected region: {selected_region}")
        
        # ✅ STEP 2: Detect comparison mode BEFORE any file checks
        is_comparison_mode = 'HDFC Bank' in selected_clients and 'ICICI Bank' in selected_clients
        print(f"🔄 Comparison mode: {is_comparison_mode}")
        
        # ✅ STEP 3: Handle comparison mode separately
        if is_comparison_mode:
            print("🔄 COMPARISON MODE DETECTED")
            
            # Check for dual files
            if 'hdfc_file' not in request.files or 'icici_file' not in request.files:
                print("❌ Both files required for comparison")
                flash('Both HDFC and ICICI files are required for comparison')
                return redirect(request.url)
            
            hdfc_file = request.files['hdfc_file']
            icici_file = request.files['icici_file']
            
            print(f"📁 HDFC file received: {hdfc_file.filename}")
            print(f"📁 ICICI file received: {icici_file.filename}")
            
            if hdfc_file.filename == '' or icici_file.filename == '':
                print("❌ File selection incomplete")
                flash('Please select both files')
                return redirect(request.url)
            
            if not (allowed_file(hdfc_file.filename) and allowed_file(icici_file.filename)):
                print("❌ Invalid file types")
                flash('Both files must be Excel format (.xlsx or .xls)')
                return redirect(request.url)
            
            # Save both files
            hdfc_filename = secure_filename(hdfc_file.filename)
            icici_filename = secure_filename(icici_file.filename)
            hdfc_filepath = os.path.join(app.config['UPLOAD_FOLDER'], hdfc_filename)
            icici_filepath = os.path.join(app.config['UPLOAD_FOLDER'], icici_filename)
            
            print(f"💾 Saving HDFC file: {hdfc_filename}")
            print(f"💾 Saving ICICI file: {icici_filename}")
            
            try:
                hdfc_file.save(hdfc_filepath)
                icici_file.save(icici_filepath)
                print("✅ Both files saved successfully")
            except Exception as e:
                print(f"❌ Error saving files: {str(e)}")
                flash(f'Error saving files: {str(e)}')
                return redirect(request.url)

            
            
            # ✅ NEW: Process QAR Analysis FIRST
            print("🔍 Starting QAR Analysis...")
            qar_data, qar_error = process_qar_analysis(hdfc_filepath, icici_filepath)
            
            if qar_error:
                print(f"❌ QAR Analysis error: {qar_error}")
                flash(qar_error)
                return redirect(request.url)
            
            print("✅ QAR Analysis completed successfully")
            
            # Also process grouped comparison (existing functionality)
            print("🔍 Starting grouped comparison processing...")
            comparison_data, error_msg = process_comparison(hdfc_filepath, icici_filepath)
            
            if error_msg:
                print(f"❌ Comparison error: {error_msg}")
                flash(error_msg)
                return redirect(request.url)
            
            print("✅ Grouped comparison completed successfully")
            print("="*60 + "\n")
            
            # ✅ NEW: Render QAR analysis template instead of comparison template
            return render_template('qar_analysis.html',
                                 qar_data=qar_data,
                                 comparison_data=comparison_data)
        
        # ✅ STEP 4: Single file upload (only runs if NOT comparison mode)
        print("📄 SINGLE FILE MODE")
        
        if 'file' not in request.files:
            print("❌ No file part in request")
            flash('No file part')
            return redirect(request.url)
        
        file = request.files['file']
        print(f"📁 File received: {file.filename}")
        
        if file.filename == '':
            print("❌ No file selected")
            flash('No selected file')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            print(f"✅ File type allowed: {file.filename}")
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            print(f"📂 Target filepath: {filepath}")
            
            # Remove existing file if it exists
            if os.path.exists(filepath):
                print(f"⚠️  File already exists, attempting to remove...")
                try:
                    os.remove(filepath)
                    print(f"✅ Old file removed successfully")
                except Exception as e:
                    print(f"❌ Error removing old file: {str(e)}")
                    flash(f'Error removing old file: {str(e)}')
                    return redirect(request.url)
            
            # Save new file
            print(f"💾 Saving file...")
            try:
                file.save(filepath)
                print(f"✅ File saved successfully to: {filepath}")
            except PermissionError as e:
                print(f"❌ Permission error: {str(e)}")
                flash('⚠️ File is currently open. Please close it and try again.')
                return redirect(request.url)
            except Exception as e:
                print(f"❌ Unexpected error saving file: {str(e)}")
                flash(f'Error saving file: {str(e)}')
                return redirect(request.url)
            
            # Get selected client
            selected_clients = request.form.get('selected_clients', 'HDFC Bank')
            print(f"🏦 Selected clients from form: {selected_clients}")
            
            client_name = selected_clients.split(',')[0] if selected_clients else 'HDFC Bank'
            validation_mode = request.form.get('validation_mode', 'quantity')
            session['last_uploaded_file'] = filepath
            session['last_client'] = client_name
            session['last_region'] = selected_region
            print(f"📊 Validation mode: {validation_mode}")

            is_regional = False

            # Build model key based on client and mode
            if client_name == 'HDFC Bank':
                if validation_mode == 'quantity':
                    model_key = 'hdfc_qty'
                elif validation_mode == 'amount':
                    model_key = 'hdfc_amount'
                elif validation_mode == 'rate':
                    model_key = 'hdfc_rate'
                elif validation_mode == 'grouped':
                    model_key = 'hdfc_grouped'
                else:
                    model_key = 'hdfc_qty'

            elif client_name == 'ICICI Bank':
                # ✅ NEW: Check if region is selected for ICICI
                is_regional = False  
                if selected_region and selected_region in ['east', 'west', 'north', 'south']:
                    is_regional = True
                    print(f"🗺️ REGIONAL MODE DETECTED: {selected_region.upper()}")

                    if validation_mode != 'rate':
                        print(f"⚠️ Mode '{validation_mode}' requested, but Regional Quantity model is not ready. Switching to 'rate'.")
                        validation_mode = 'rate'

                    model_key = f'icici_rate_{selected_region}'
                        
                else:
                    # Use default ICICI models (no region selected)
                    if validation_mode == 'quantity':
                        model_key = 'icici_qty'
                    elif validation_mode == 'amount':
                        model_key = 'icici_amount'
                    elif validation_mode == 'rate':
                        model_key = 'icici_rate'
                    elif validation_mode == 'grouped':
                        model_key = 'icici_grouped'
                    else:
                        model_key = 'icici_qty'

            # Inside the upload_file() route, after the ICICI Bank section:

            elif client_name == 'Reliance Retail':
                # Get sub-client selection
                sub_client = request.form.get('reliance_sub_client', 'yousta')

                if sub_client == 'yousta':
                    if validation_mode == 'quantity':
                        model_key = 'reliance_qty_yousta'
                    elif validation_mode == 'rate':
                        model_key = 'reliance_rate_yousta'
                    else:
                        model_key = 'reliance_qty_yousta'
                elif sub_client == 'smart':
                    if validation_mode == 'quantity':
                        model_key = 'reliance_qty_smart'
                    elif validation_mode == 'rate':
                        flash('Rate mode not available for Smart yet.')
                        return redirect(request.url)
                    else:
                        model_key = 'reliance_qty_smart'
                else:
                    # Other sub-clients not ready yet
                    flash(f'{sub_client.capitalize()} model not available yet.')
                    return redirect(request.url)

                print(f"🏦 Using Reliance model: {model_key}")
            else:
                model_key = 'hdfc_qty'

            print(f"🏦 Using model key: {model_key}")

            # Handle grouped mode differently
            if validation_mode == 'grouped':
                VALIDATOR = get_validator(model_key)
                if VALIDATOR is None:
                    print(f"❌ Grouped validator not available")
                    flash(f'Grouped validation model is not available.')
                    return redirect(request.url)
                
                print(f"✅ Using grouped validator")
                results_data, error_msg = validate_grouped_file(VALIDATOR, filepath)
            else:
                VALIDATOR = get_validator(model_key)
                if VALIDATOR is None:
                    print(f"❌ Validator not available for model: {model_key}")
                    flash(f'Validation model for {client_name} - {validation_mode} is not available.')
                    return redirect(request.url)
                
                print(f"✅ Using {model_key} validator")
                selected_region = request.form.get('selected_region', '')
                
                print(f"📂 Validating file: {os.path.basename(filepath)}")
                print(f"📊 Validation mode: {validation_mode}")
                print(f"📍 Selected region: {selected_region}")
                print(f"🎯 Model key: {model_key}")
                print(f"🌍 Is regional: {is_regional}")

                if hasattr(VALIDATOR, 'sqfeet_ranges'):
                    print(f"🗂️  Model contains these ranges: {list(VALIDATOR.sqfeet_ranges.keys())}")

                results_data, error_msg = VALIDATOR.validate_test_file(filepath, mode=validation_mode, is_regional=is_regional,model_key=model_key)
            
            if error_msg:
                print(f"❌ Validation error: {error_msg}")
                flash(error_msg)
                return redirect(request.url)
            
            if results_data is None:
                print(f"❌ Validation returned None (no results)")
                flash("Validation failed - no results returned")
                return redirect(request.url)
            
            print(f"✅ Validation successful! Results count: {len(results_data)}")

            # Process results
            print(f"📊 Processing results...")
            results_df = pd.DataFrame(results_data)

            if validation_mode not in ['quantity', 'grouped'] and 'Threshold_Min' not in results_df.columns:
                results_df['Threshold_Min'], results_df['Threshold_Max'] = zip(
                    *results_df['Validation_Message'].map(extract_thresholds)
                )

            # Save reports
            print(f"💾 Saving reports...")
            full_report_filename = f"validated_{filename}"
            full_report_filepath = os.path.join(app.config['OUTPUT_FOLDER'], full_report_filename)
            styled_df = results_df.style.apply(highlight_invalid, axis=1)
            styled_df.to_excel(full_report_filepath, engine='xlsxwriter', index=False)
            print(f"✅ Full report saved: {full_report_filename}")
            
            invalid_df = results_df[results_df['Is_Valid'] == False].copy()
            invalid_report_filename = f"invalid_rows_{filename}"
            if not invalid_df.empty:
                invalid_report_filepath = os.path.join(app.config['OUTPUT_FOLDER'], invalid_report_filename)
                invalid_df.to_excel(invalid_report_filepath, index=False)
                print(f"✅ Invalid report saved: {invalid_report_filename}")

            # Prepare display
            print(f"🖥️  Preparing web display...")
            display_df = results_df.copy()

            # Format display based on validation_mode
            if validation_mode == 'grouped':
                display_df['As_Per_CWI_Amount'] = display_df['As_Per_CWI_Amount'].apply(
                    lambda x: VALIDATOR.format_indian_currency(x))
                display_df['Rate_per_sqfeet'] = display_df['Rate_per_sqfeet'].apply(
                    lambda x: VALIDATOR.format_indian_currency(x))

                def fmt_currency(val):
                    if val == 'NA' or pd.isna(val): return 'NA'
                    return f"Rs. {VALIDATOR.format_indian_currency(val)}"

                # Combine min/max into single range columns
                def make_range(row, min_col, max_col):
                    mn = row.get(min_col, 'NA')
                    mx = row.get(max_col, 'NA')
                    if mn == 'NA' or mx == 'NA' or pd.isna(mn) or pd.isna(mx):
                        return 'NA'
                    return f"Rs. {VALIDATOR.format_indian_currency(mn)} - Rs. {VALIDATOR.format_indian_currency(mx)}"

                display_df['CWI_Weighted_Range'] = display_df.apply(
                    lambda row: make_range(row, 'CWI_Weighted_Min', 'CWI_Weighted_Max'), axis=1)
                display_df['Rate_Weighted_Range'] = display_df.apply(
                    lambda row: make_range(row, 'Rate_Weighted_Min', 'Rate_Weighted_Max'), axis=1)

                # Clickable threshold columns with file tooltips
                def make_threshold_clickable(row, val_col, files_col):
                    val = row.get(val_col, 'NA')
                    if val == 'NA' or pd.isna(val): return 'NA'
                    display_val = f"Rs. {VALIDATOR.format_indian_currency(val)}"
                    files_data = row.get(files_col, [])
                    if not files_data or not isinstance(files_data, list):
                        return display_val
                    tooltip_lines = []
                    for i, f in enumerate(files_data[:5]):
                        if isinstance(f, dict):
                            fname = f.get('file', f.get('original_particular', 'Unknown'))
                        else:
                            fname = str(f)
                        tooltip_lines.append(f"{i+1}. {fname}")
                    if len(files_data) > 5:
                        tooltip_lines.append(f"... and {len(files_data)-5} more")
                    tooltip_text = html.escape("\\n".join(tooltip_lines))
                    return (f'<span class="threshold-clickable" title="{tooltip_text}" '
                            f'style="cursor:pointer;color:#007bff;text-decoration:underline;">'
                            f'{display_val}</span>')

                display_df['Threshold_Min_Amount'] = display_df.apply(
                    lambda row: make_threshold_clickable(row, 'CWI_Threshold_Min', 'CWI_Min_Files'), axis=1)
                display_df['Threshold_Max_Amount'] = display_df.apply(
                    lambda row: make_threshold_clickable(row, 'CWI_Threshold_Max', 'CWI_Max_Files'), axis=1)
                display_df['Threshold_Min_Rate'] = display_df.apply(
                    lambda row: make_threshold_clickable(row, 'Rate_Threshold_Min', 'Rate_Min_Files'), axis=1)
                display_df['Threshold_Max_Rate'] = display_df.apply(
                    lambda row: make_threshold_clickable(row, 'Rate_Threshold_Max', 'Rate_Max_Files'), axis=1)

                # Format Is_Valid_Amount and Is_Valid_Rate with icons
                if 'Is_Valid_Amount' in display_df.columns:
                    display_df['Is_Valid_Amount'] = display_df['Is_Valid_Amount'].apply(
                        lambda x: '✅ Valid' if x else '❌ Invalid')
                if 'Is_Valid_Rate' in display_df.columns:
                    display_df['Is_Valid_Rate'] = display_df['Is_Valid_Rate'].apply(
                        lambda x: '✅ Valid' if x else '❌ Invalid')

            elif validation_mode == 'rate':
                display_df['Rate_Raw'] = display_df['Rate'].copy()
                display_df['Amount_Raw'] = display_df['Amount'].copy()
                display_df['Rate'] = display_df.apply(lambda row: VALIDATOR.format_indian_currency(row['Rate']), axis=1)
                display_df['Amount'] = display_df.apply(lambda row: VALIDATOR.format_indian_currency(row['Amount']), axis=1)
                display_df['Quantity'] = display_df.apply(lambda row: int(round(row['Quantity'])) if VALIDATOR.is_number_based_unit(row['Unit']) else row['Quantity'], axis=1)

                if 'Weighted_Avg_Min' in display_df.columns:
                    def format_weighted_avg(row, col_name):
                        val = row.get(col_name)
                        if pd.isna(val) or val is None or val == 'NA': return 'NA'
                        return VALIDATOR.format_indian_currency(val)

                    display_df['Weighted_Avg_Min'] = display_df.apply(lambda row: format_weighted_avg(row, 'Weighted_Avg_Min'), axis=1)
                    display_df['Weighted_Avg_Max'] = display_df.apply(lambda row: format_weighted_avg(row, 'Weighted_Avg_Max'), axis=1)

                if 'Threshold_Min' in display_df.columns and 'Threshold_Max' in display_df.columns:
                    def format_threshold_clickable(row, col_name, files_col_name):
                        val = row.get(col_name)
                        if val == 'NA' or pd.isna(val): return 'NA'
                        display_val = VALIDATOR.format_indian_currency(val)
                        files_data = row.get(files_col_name, [])
                        if not files_data: return str(display_val)
                        tooltip_lines = [f"{i+1}. {f.get('file', 'Unknown')}: {f.get('original_particular', 'N/A')}" for i, f in enumerate(files_data[:5])]
                        if len(files_data) > 5: tooltip_lines.append(f"... and {len(files_data) - 5} more files")
                        tooltip_text = html.escape("\\n".join(tooltip_lines))
                        return f'<span class="threshold-clickable" title="{tooltip_text}" style="cursor: pointer; color: #007bff; text-decoration: underline;">{display_val}</span>'

                    display_df['Threshold_Min'] = display_df.apply(lambda row: format_threshold_clickable(row, 'Threshold_Min', 'Min_Threshold_Files'), axis=1)
                    display_df['Threshold_Max'] = display_df.apply(lambda row: format_threshold_clickable(row, 'Threshold_Max', 'Max_Threshold_Files'), axis=1)

            elif validation_mode == 'amount':
                display_df['Amount_Raw'] = display_df['Amount'].copy()
                display_df['Amount'] = display_df.apply(lambda row: VALIDATOR.format_indian_currency(int(round(row['Amount'])) if VALIDATOR.is_number_based_unit(row['Unit']) else row['Amount']), axis=1)

                if 'Weighted_Avg_Min' in display_df.columns:
                    def format_weighted_avg(row, col_name):
                        val = row.get(col_name)
                        if pd.isna(val) or val is None or val == 'NA': return 'NA'
                        return VALIDATOR.format_indian_currency(val)

                    display_df['Weighted_Avg_Min'] = display_df.apply(lambda row: format_weighted_avg(row, 'Weighted_Avg_Min'), axis=1)
                    display_df['Weighted_Avg_Max'] = display_df.apply(lambda row: format_weighted_avg(row, 'Weighted_Avg_Max'), axis=1)

                if 'Threshold_Min' in display_df.columns and 'Threshold_Max' in display_df.columns:
                    def format_threshold_clickable(row, col_name, files_col_name):
                        val = row.get(col_name)
                        if val == 'NA' or pd.isna(val): return 'NA'
                        display_val = VALIDATOR.format_indian_currency(val)
                        files_data = row.get(files_col_name, [])
                        if not files_data: return str(display_val)
                        tooltip_lines = [f"{i+1}. {f.get('file', 'Unknown')}: {f.get('original_particular', 'N/A')}" for i, f in enumerate(files_data[:5])]
                        if len(files_data) > 5: tooltip_lines.append(f"... and {len(files_data) - 5} more files")
                        tooltip_text = html.escape("\\n".join(tooltip_lines))
                        return f'<span class="threshold-clickable" title="{tooltip_text}" style="cursor: pointer; color: #007bff; text-decoration: underline;">{display_val}</span>'

                    display_df['Threshold_Min'] = display_df.apply(lambda row: format_threshold_clickable(row, 'Threshold_Min', 'Min_Threshold_Files'), axis=1)
                    display_df['Threshold_Max'] = display_df.apply(lambda row: format_threshold_clickable(row, 'Threshold_Max', 'Max_Threshold_Files'), axis=1)

            else: # quantity mode
                display_df['Quantity'] = display_df.apply(lambda row: int(round(row['Quantity'])) if VALIDATOR.is_number_based_unit(row['Unit']) else row['Quantity'], axis=1)

                if 'Weighted_Avg_Min' in display_df.columns:
                    def format_weighted_avg(row, col_name):
                        val = row.get(col_name)
                        if pd.isna(val) or val is None or val == 'NA': return 'NA'
                        return int(round(val)) if VALIDATOR.is_number_based_unit(row.get('Unit', '')) else f"{val:.2f}"

                    display_df['Weighted_Avg_Min'] = display_df.apply(lambda row: format_weighted_avg(row, 'Weighted_Avg_Min'), axis=1)
                    display_df['Weighted_Avg_Max'] = display_df.apply(lambda row: format_weighted_avg(row, 'Weighted_Avg_Max'), axis=1)

                if 'Threshold_Min' in display_df.columns and 'Threshold_Max' in display_df.columns:
                    def format_threshold_clickable(row, col_name, files_col_name):
                        val = row.get(col_name)
                        if val == 'NA' or pd.isna(val): return 'NA'
                        display_val = int(round(float(val))) if VALIDATOR.is_number_based_unit(row.get('Unit', '')) else f"{float(val):.2f}"
                        files_data = row.get(files_col_name, [])
                        if not files_data: return str(display_val)
                        tooltip_lines = [f"{i+1}. {f.get('file', 'Unknown')}: {f.get('original_particular', 'N/A')}" for i, f in enumerate(files_data[:5])]
                        if len(files_data) > 5: tooltip_lines.append(f"... and {len(files_data) - 5} more files")
                        tooltip_text = html.escape("\\n".join(tooltip_lines))
                        return f'<span class="threshold-clickable" title="{tooltip_text}" style="cursor: pointer; color: #007bff; text-decoration: underline;">{display_val}</span>'

                    display_df['Threshold_Min'] = display_df.apply(lambda row: format_threshold_clickable(row, 'Threshold_Min', 'Min_Threshold_Files'), axis=1)
                    display_df['Threshold_Max'] = display_df.apply(lambda row: format_threshold_clickable(row, 'Threshold_Max', 'Max_Threshold_Files'), axis=1)

            # Add hover tooltip for Original_Particular column
            if 'Original_Particular' in display_df.columns:
                display_df['Original_Particular'] = display_df['Original_Particular'].apply(lambda x: f'<span title="{html.escape(str(x))}">{html.escape(str(x))}</span>')    

            # Analysis button
            if validation_mode != 'grouped':
                def create_analysis_button(row):
                    if not row['Is_Valid'] and row['Threshold_Min'] != 'NA':
                        safe_particular = html.escape(row["Original_Particular"])
                        value = row.get('Rate_Raw', 0) if validation_mode == 'rate' else row.get('Amount_Raw', 0) if validation_mode == 'amount' else row.get('Quantity', 0)
                        return (f'<button class="btn btn-sm btn-info btn-analysis" data-particular="{safe_particular}" data-quantity="{value}" '
                                f'data-range="{row["Sq_Feet_Range"]}" data-client="{client_name}" data-mode="{validation_mode}" data-region="{selected_region}">'
                                f'View Analysis</button>')
                    return ''
                display_df['Analysis'] = display_df.apply(create_analysis_button, axis=1)

            display_df['Is_Valid'] = display_df['Is_Valid'].apply(lambda x: '✅ Valid' if x else '❌ Invalid')
            
            # Dynamic column order
            if validation_mode == 'grouped':
                column_order = ['Sr.No', 'Particulars', 'As_Per_CWI_Amount', 'Rate_per_sqfeet', 
                                'Sq_Feet_Range', 'Is_Valid_Amount', 'Is_Valid_Rate', 
                                'CWI_Weighted_Range', 'Rate_Weighted_Range',
                                'Threshold_Min_Amount', 'Threshold_Max_Amount',
                                'Threshold_Min_Rate', 'Threshold_Max_Rate']
            elif validation_mode == 'amount':
                column_order = ['Sr.No', 'Original_Particular', 'Amount', 'Unit', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message', 'Matched_Particular', 'Similarity_Score','Weighted_Avg_Min', 'Weighted_Avg_Max', 'Threshold_Min', 'Threshold_Max', 'Analysis']
            elif validation_mode == 'rate':
                if client_name == 'Reliance Retail':
        # ✅ NEW: Add Service_Code as second column for Reliance
                    column_order = ['Sr.No', 'Service_Code', 'Original_Particular', 'Rate', 'Unit', 'Quantity', 'Amount', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message', 'Matched_Particular', 'Similarity_Score', 'Threshold_Min', 'Threshold_Max', 'Weighted_Avg_Min', 'Weighted_Avg_Max', 'Analysis']
                else:
                    column_order = ['Sr.No', 'Original_Particular', 'Rate', 'Unit', 'Quantity', 'Amount', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message', 'Matched_Particular', 'Similarity_Score', 'Threshold_Min', 'Threshold_Max', 'Weighted_Avg_Min', 'Weighted_Avg_Max', 'Analysis']
                if is_regional:
                    column_order = ['Sr.No', 'Original_Particular', 'Rate', 'Unit', 'Quantity', 'Amount', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message_East', 'Validation_Message_West', 'Validation_Message_North', 'Validation_Message_South', 'Matched_Particular', 'Similarity_Score', 'Threshold_Min', 'Threshold_Max', 'Weighted_Avg_Min', 'Weighted_Avg_Max', 'Analysis']
                else:
                    column_order = ['Sr.No', 'Original_Particular', 'Rate', 'Unit', 'Quantity', 'Amount', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message', 'Matched_Particular', 'Similarity_Score', 'Threshold_Min', 'Threshold_Max', 'Weighted_Avg_Min', 'Weighted_Avg_Max', 'Analysis']
            else: # quantity
                if client_name == 'Reliance Retail':
                    # ✅ NEW: Add Service_Code as second column for Reliance
                    column_order = ['Sr.No', 'Service_Code', 'Original_Particular', 'Quantity', 'Unit', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message', 'Matched_Particular', 'Similarity_Score', 'Threshold_Min', 'Threshold_Max', 'Weighted_Avg_Min', 'Weighted_Avg_Max', 'Analysis']
                else:
                    column_order = ['Sr.No', 'Original_Particular', 'Quantity', 'Unit', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message', 'Matched_Particular', 'Similarity_Score', 'Threshold_Min', 'Threshold_Max', 'Weighted_Avg_Min', 'Weighted_Avg_Max', 'Analysis']
                    
            
            display_df = display_df[column_order]
            
            # >>>>>>>>>> FILTER LOGIC STARTS HERE <<<<<<<<<<
            # Filter out rows where all range columns are N/A
            check_cols = ['Weighted_Avg_Min', 'Weighted_Avg_Max', 'Threshold_Min', 'Threshold_Max']
            # Only check columns that actually exist in the dataframe
            existing_check_cols = [c for c in check_cols if c in display_df.columns]
            
            if existing_check_cols:
                def is_row_all_na(row):
                    # Returns True if ALL checked columns are 'NA'
                    for col in existing_check_cols:
                        val = str(row.get(col, '')).strip().upper()
                        # If we find any value that is NOT in the NA list, then the row is kept
                        if val not in ['NA', 'N/A', 'NAN', 'NONE', '']:
                            return False 
                    return True

                # Keep rows where NOT all columns are NA
                #display_df = display_df[~display_df.apply(is_row_all_na, axis=1)]
            # >>>>>>>>>> FILTER LOGIC ENDS HERE <<<<<<<<<<

            # Generate HTML table
            results_html = display_df.to_html(classes='table table-striped table-hover', justify='left', index=False, escape=False)

            if is_regional and selected_region:
                region_col_name = f'Validation_Message_{selected_region.capitalize()}'
                results_html = results_html.replace(f'<th>{region_col_name}</th>', f'<th class="selected-region-header">{region_col_name} <i class="fas fa-check-circle"></i></th>')

            if is_regional and selected_region:
                region_col_name = f'Validation_Message_{selected_region.capitalize()}'
                results_html = results_html.replace(f'<th>{region_col_name}</th>', f'<th class="selected-region-header">{region_col_name} ✓</th>')

            total = len(results_df)
            valid = results_df['Is_Valid'].sum()
            invalid = total - valid
            accuracy = (valid / total * 100) if total > 0 else 0
            
            print(f"📈 Statistics: Total={total}, Valid={valid}, Invalid={invalid}, Accuracy={accuracy:.2f}%")
            print(f"✅ Rendering results page...")
            print("="*60 + "\n")

            return render_template('results.html', 
                    table_html=results_html, 
                    full_report_filename=full_report_filename,
                    invalid_report_filename=invalid_report_filename,
                    total=total, valid=valid, invalid=invalid, 
                    accuracy=f"{accuracy:.2f}%",
                    client=client_name,
                    validation_mode=validation_mode,
                    region=selected_region) 
        else:
            print(f"❌ Invalid file type: {file.filename}")
            flash('Invalid file type. Please upload an .xlsx or .xls file.')
            return redirect(request.url)

    return render_template('index.html')
@app.route('/outputs/<filename>')
def download_file(filename):
    return send_from_directory(app.config['OUTPUT_FOLDER'], filename, as_attachment=True)
    
@app.route('/error')
def error_page():
    return render_template('error.html', message="An unknown error occurred.")

# --- FINAL /graph ROUTE with ALL fixes ---
@app.route('/graph')
def get_graph():
    try:
        # --- 1. Get data from request ---
        chart_type = request.args.get('type')
        sq_range = request.args.get('range')
        particular = request.args.get('particular')
        user_value = float(request.args.get('value', 0))
        client = request.args.get('client', 'HDFC Bank')
        validation_mode = request.args.get('mode', 'quantity')
        
        # ✅ NEW: Get region parameter for regional models
        selected_region = request.args.get('region', '')

        print(f"🔍 Graph request: {chart_type}, range={sq_range}, mode={validation_mode}, region={selected_region}")

        # Build model key
        if client == 'HDFC Bank':
            if validation_mode == 'quantity':
                model_key = 'hdfc_qty'
            elif validation_mode == 'amount':
                model_key = 'hdfc_amount'
            elif validation_mode == 'rate':
                model_key = 'hdfc_rate'
            else:
                model_key = 'hdfc_qty'
        elif client == 'ICICI Bank':
            # ✅ NEW: Check for regional mode
            if selected_region and selected_region in ['east', 'west', 'north', 'south']:
                # Use regional model
                if validation_mode == 'rate':
                    model_key = f'icici_rate_{selected_region}'
                elif validation_mode == 'quantity':
                    model_key = f'icici_qty_{selected_region}'
                else:
                    model_key = f'icici_rate_{selected_region}'  # Default to rate for regional
                print(f"🗺️ Using REGIONAL model: {model_key}")
            else:
                # Use standard models
                if validation_mode == 'quantity':
                    model_key = 'icici_qty'
                elif validation_mode == 'amount':
                    model_key = 'icici_amount'
                elif validation_mode == 'rate':
                    model_key = 'icici_rate'
                else:
                    model_key = 'icici_qty'
        else:
            model_key = 'hdfc_qty'

        print(f"🔑 Using model: {model_key}")

        validator = get_validator(model_key)
        if validator is None:
            return jsonify({'error': f'Validation model not available for {model_key}'})

        # FIXED: Determine which key to use
        if 'rate_ranges' in validator.sqfeet_ranges.get(sq_range, {}):
            range_key = 'rate_ranges'
        elif 'amount_ranges' in validator.sqfeet_ranges.get(sq_range, {}):
            range_key = 'amount_ranges'
        else:
            range_key = 'quantity_ranges'

        cleaned_particular = validator.clean_text(particular)
        matched_particular = None
        
        # FIXED: Use range_key instead of hardcoded 'quantity_ranges'
        if cleaned_particular in validator.sqfeet_ranges[sq_range][range_key]:
            matched_particular = cleaned_particular
        else:
            sim_p, _ = validator.find_similar_particular(particular, sq_range)
            if sim_p: 
                matched_particular = sim_p

        if not matched_particular:
            return jsonify({'error': 'Could not find statistics for this item.'})

        # FIXED: Use range_key instead of hardcoded 'quantity_ranges'
        stats = validator.sqfeet_ranges[sq_range][range_key][matched_particular]
        mean = stats.get('overall_mean', 0)
        std = stats.get('overall_std', 0)

        # ✅ FIX: Use THRESHOLD values (absolute min/max) for quantity mode
        # For quantity mode, use threshold values (overall_min/max)
        # For amount/rate modes, use validation_range which is correct
        # REPLACE WITH:
        if validation_mode == 'quantity':
            lower_bound = stats.get('weighted_avg_min', stats.get('overall_mean', 0))
            upper_bound = stats.get('weighted_avg_max', stats.get('overall_mean', 0))  
        else:
            # ✅ USE WEIGHTED AVERAGES FOR AMOUNT AND RATE
            lower_bound = stats.get('weighted_avg_min', stats.get('overall_mean', 0))
            upper_bound = stats.get('weighted_avg_max', stats.get('overall_mean', 0))
        # --- 3. Generate the correct plot based on chart_type ---
        fig, ax = plt.subplots(figsize=(8, 4))
        explanation = ""

        if chart_type == 'bell_curve':
            if std == 0: std = mean * 0.1 if mean > 0 else 1
            x = np.linspace(mean - 4*std, mean + 4*std, 1000)
            y = norm.pdf(x, mean, std)
            ax.plot(x, y, label='Normal Distribution')
            ax.axvline(lower_bound, color='green', linestyle='--', label=f'Valid Min: {lower_bound:.2f}')
            ax.axvline(upper_bound, color='green', linestyle='--')
            ax.axvline(mean, color='orange', linestyle='-', label=f'Average: {mean:.2f}')
            ax.plot(user_value, 0, 'rX', markersize=12, label=f'Your Value: {user_value:.2f}')
            ax.fill_between(x, y, where=((x >= lower_bound) & (x <= upper_bound)), color='green', alpha=0.2)
            ax.set_title('Bell Curve Analysis')
            ax.set_xlabel('Quantity')
            ax.set_ylabel('Likelihood')
            # ✅ FIX: Update explanation to show THRESHOLD range for quantity mode
            if validation_mode == 'quantity':
                # ✅ UNIFIED EXPLANATION FOR ALL MODES
                explanation = (f"This bell curve shows the expected distribution. The average is <b>{mean:.2f}</b>, and the weighted average range is from <b>{lower_bound:.2f}</b> to <b>{upper_bound:.2f}</b>. "
                               f"Your value of <b style='color:red;'>{user_value:.2f}</b> is outside this range.")
            elif validation_mode in ['amount', 'rate']:
                # ✅ NEW: Different explanation for weighted averages
                explanation = (f"This bell curve shows the expected distribution. The average is <b>{mean:.2f}</b>, and the weighted average range (validated range from historical data) is from <b>{lower_bound:.2f}</b> to <b>{upper_bound:.2f}</b>. "
                               f"Your value of <b style='color:red;'>{user_value:.2f}</b> is outside this range.")
            else:
                explanation = (f"This bell curve shows the expected distribution. The average is <b>{mean:.2f}</b>, and the valid range is from <b>{lower_bound:.2f}</b> to <b>{upper_bound:.2f}</b>. "
                               f"Your value of <b style='color:red;'>{user_value:.2f}</b> is outside this range.")

        elif chart_type == 'bar_plot':
            ax.bar(['Min', 'Max'], [lower_bound, upper_bound], color='green', alpha=0.6, label='Acceptable Range')
            ax.bar(['Your Value'], [user_value], color='red', label='Your Value')
            ax.set_title('Bar Plot Comparison')
            ax.set_ylabel('Quantity')
            if validation_mode == 'quantity':
                # ✅ UNIFIED EXPLANATION FOR ALL MODES
                explanation = (f"This bar plot compares your value against the weighted average min/max. "
                               f"Your value is <b style='color:red;'>{user_value:.2f}</b>, while the weighted average range is <b>{lower_bound:.2f}</b> - <b>{upper_bound:.2f}</b>.")
            elif validation_mode in ['amount', 'rate']:
                # ✅ UNIFIED EXPLANATION FOR ALL MODES
                explanation = (f"This bar plot compares your value against the weighted average min/max. "
                               f"Your value is <b style='color:red;'>{user_value:.2f}</b>, while the weighted average range is <b>{lower_bound:.2f}</b> - <b>{upper_bound:.2f}</b>.")
            else:
                # ✅ UNIFIED EXPLANATION FOR ALL MODES
                explanation = (f"This bar plot compares your value against the weighted average min/max. "
                               f"Your value is <b style='color:red;'>{user_value:.2f}</b>, while the weighted average range is <b>{lower_bound:.2f}</b> - <b>{upper_bound:.2f}</b>.")

        elif chart_type == 'box_plot':
            fig, ax = plt.subplots(figsize=(9, 5)) 
            if std == 0: std = mean * 0.1 if mean > 0 else 1
            np.random.seed(0) 
            sample_data = np.random.normal(loc=mean, scale=std, size=100)
            # Add the user's value to the data to see it as an outlier if it is one
            data_with_user_value = np.append(sample_data, user_value)
            
            boxplot = ax.boxplot(data_with_user_value, vert=False, widths=0.4, patch_artist=True,
                                 boxprops=dict(facecolor='lightblue'),
                                 medianprops=dict(color='red', linewidth=2))

            ax.plot(user_value, 1, 'rX', markersize=12, label=f'Your Value: {user_value:.2f}')
            ax.set_title('Box Plot Analysis (Historical Data Distribution)')
            ax.set_xlabel('Quantity')
            ax.grid(True, axis='x', linestyle='--', alpha=0.7) 
            ax.set_yticklabels([]) 
            ax.set_ylim(0.5, 1.5)
            explanation = (f"This Box Plot shows the statistical spread of historical data. The 'box' represents the middle 50% of the data. "
                           f"The red line inside the box is the median. The whiskers show the expected range, and circles are outliers. "
                           f"Your value of <b style='color:red;'>{user_value:.2f}</b> (the red 'X') is compared against this distribution.")

        else: # Default to threshold_line
            fig, ax = plt.subplots(figsize=(9, 3)) 
            ax.hlines(y=1, xmin=lower_bound, xmax=upper_bound, linewidth=15, color='green', alpha=0.5)
            
            is_valid = lower_bound <= user_value <= upper_bound
            marker_color = 'blue' if is_valid else 'red'
            ax.plot(user_value, 1, 'X', markersize=15, color=marker_color, label=f'Your Value: {user_value:.2f}')

            ax.text(lower_bound, 0.97, f'Min\n{lower_bound:.2f}', ha='center', va='top')
            ax.text(upper_bound, 0.97, f'Max\n{upper_bound:.2f}', ha='center', va='top')
            ax.text(user_value, 1.03, f'Your Value\n{user_value:.2f}', ha='center', va='bottom', color=marker_color, weight='bold')

            padding = (upper_bound - lower_bound) * 0.5 if (upper_bound > lower_bound) else upper_bound or 1
            ax.set_xlim(min(lower_bound, user_value) - padding, max(upper_bound, user_value) + padding)
            ax.set_ylim(0.9, 1.1) 
            ax.grid(True, axis='x', linestyle='--', alpha=0.7)
            ax.get_yaxis().set_visible(False)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['left'].set_visible(False)
            ax.set_title('Threshold Line Chart')
            plt.tight_layout()
            if validation_mode == 'quantity':
                # ✅ UNIFIED EXPLANATION FOR ALL MODES
                explanation = (f"The green line shows the weighted average range from <b>{lower_bound:.2f}</b> to <b>{upper_bound:.2f}</b>. "
                               f"Your value of <b style='color:{marker_color};'>{user_value:.2f}</b> is shown by the 'X'.")
            elif validation_mode in ['amount', 'rate']:
                # ✅ UNIFIED EXPLANATION FOR ALL MODES
                explanation = (f"The green line shows the weighted average range from <b>{lower_bound:.2f}</b> to <b>{upper_bound:.2f}</b>. "
                               f"Your value of <b style='color:{marker_color};'>{user_value:.2f}</b> is shown by the 'X'.")
            else:
                # ✅ UNIFIED EXPLANATION FOR ALL MODES
                explanation = (f"The green line shows the weighted average range from <b>{lower_bound:.2f}</b> to <b>{upper_bound:.2f}</b>. "
                               f"Your value of <b style='color:{marker_color};'>{user_value:.2f}</b> is shown by the 'X'.")
        
        if chart_type != 'threshold_line':
            ax.legend()
        
        # --- 4. Save plot to buffer and encode ---
        buf = io.BytesIO()
        fig.savefig(buf, format='png')
        buf.seek(0)
        plt.close(fig)
        gc.collect()
        image_base64 = base64.b64encode(buf.read()).decode('utf-8')

        return jsonify({'image': image_base64, 'explanation': explanation})

    except Exception as e:
        print(f"Error in graph generation: {e}")
        return jsonify({'error': 'An error occurred while generating the chart.'})

@app.route('/rate_comparison_analysis')
def rate_comparison_analysis():
    """Display ICICI vs HDFC vs AXIS rate ranges comparison without file upload"""
    try:
        print("\n" + "="*60)
        print("🔍 RATE COMPARISON ANALYSIS - NO FILE UPLOAD")
        print("="*60)
        
        # Load validators
        icici_validator = get_validator('icici_rate')
        hdfc_validator = get_validator('hdfc_rate')
        axis_validator = get_validator('axis_rate')
        
        if not icici_validator or not hdfc_validator or not axis_validator:
            flash("Rate models not available. Please ensure all models are loaded.")
            return redirect(url_for('upload_file'))
        
        print("✅ All validators loaded successfully")

        results = []
        sr_no = 1

        # ICICI branch: 1250sqft range
        icici_sqft = '1001-1250'
        # HDFC branch: 1270sqft range  
        hdfc_sqft = '1251-1500'
        # AXIS: similar range
        axis_sqft = '1251-1500'


        def match_particular_by_keywords(particular_text, keywords, exclude_keywords=None):
            """Check if particular matches all keywords and doesn't contain exclude keywords"""
            particular_lower = particular_text.lower()
            
            # Check if all keywords are present
            if not all(keyword.lower() in particular_lower for keyword in keywords):
                return False
            
            # Check if any exclude keywords are present
            if exclude_keywords:
                if any(exclude_keyword.lower() in particular_lower for exclude_keyword in exclude_keywords):
                    return False
            
            return True
        
       # Define the 13 particulars with their exact search terms
        # Define the 13 particulars with their exact search terms
        # Define the 13 particulars with their exact search terms
        # Define the 13 particulars with EXACT match phrases for Axis Bank
        particulars_mapping = {
            'Vitrified Tiles 800x800': {
                'icici_keywords': ['800', '800', 'vitrified', 'tiles'],
                'hdfc_keywords': ['800mm', '800'],
                'axis_exact_match': None,  # Skip search
                'axis_display': 'N/A - Not found for this bank'
            },
            'Siporex 9 inch': {
                'icici_keywords': ['9', 'inch', 'siporex'],
                'hdfc_keywords': ['150mm', 'siporex'],
                # Matches: "Do-do as 2 .1 above but with siporex / light weight block wall"
                'axis_exact_match': 'light weight block wall',  
                'axis_display': None
            },
            'R.C.C Cement Concrete': {
                'icici_keywords': ['r', 'c', 'c', 'providing', 'casting'],
                'hdfc_keywords': ['r', 'c', 'c', 'providing', 'casting'],
                # Matches: "R.C.C work :Providng & casting insitu"
                'axis_exact_match': 'reinforced cement concrete for wall',
                'axis_display': None
            },
            'Aluminum Framework BWP': {
                'icici_keywords': ['wall', 'column', 'cladding', 'aluminum'],
                'hdfc_keywords': ['aluminium', 'frame', 'bwp', 'plywood'],
                'axis_exact_match': None, # Skip search
                'axis_display': 'N/A - Not found for this bank'
            },
            'Ceramic Tiles 300x600': {
                'icici_keywords': ['ceramic', 'tiles', '300', '600'],
                'hdfc_keywords': ['600mm', '600'],
                'axis_exact_match': None, # Skip search
                'axis_display': 'N/A - Not found for this bank'
            },
            'Aluminium Composite Panel': {
                'icici_keywords': ['exterior', 'grade', '3mm', 'thick', 'aluminum', 'composite', 'panel'],
                'icici_exclude': ['silver', 'grey'],
                'hdfc_keywords': ['providing', 'fixing', 'tricolour'],
                # Matches: "Aluminum Composite panelling sheets of ALUCOBOND MAKE only"
                'axis_exact_match': 'Aluminum Composite panelling sheets of ALUCOBOND MAKE only',
                'axis_display': None
            },
            '12mm Plaster 1:4': {
                'icici_keywords': ['providing and applying in one coat'],
                'hdfc_keywords': ['1 4 cement plaster'],
                # Matches: "P/A 1:4 cement plaster with waterproofing compound"
                'axis_exact_match': 'waterproofing compound to brick',
                'axis_display': None
            },
            'Service Counter Table': {
                'icici_keywords': ['providing and fixing modular service counter table as per detail'],
                'hdfc_keywords': ['branch', 'manager', 'table', '1350', '675'],
                # Matches: "table with top made up of 19mm thk ply with veneer"
                'axis_exact_match': 'table with top made up of 19mm thk ply with veneer',
                'axis_display': None
            },
            'Full Height Partition BWR': {
                'icici_keywords': ['laminate and complete with required accessories'],
                'hdfc_keywords': ['full', 'height', 'solid', 'partition', 'bwr', 'plywood'],
                # Matches: "P/F partitions made out of Aluminium cross section"
                'axis_exact_match': 'Aluminium cross section',
                'axis_display': None
            },
            'Gypsum False Ceiling': {
                'icici_keywords': ['gypsum', 'false', 'ceiling'],
                'icici_exclude': ['vertical', 'drops', '150'],
                'hdfc_keywords': ['gypsum', 'board', 'ceiling'],
                # Matches: "Rate quoted to include cost of providing support framework"
                'axis_exact_match': 'Rate quoted to include cost of providing support framework',
                'axis_display': None
            },
            'Plaster of Paris Punning': {
                'icici_keywords': ['wall', 'punning'],
                'hdfc_keywords': ['plaster', 'paris', 'punning'],
                # Matches: "WALL PUNNING WITHOUT PAINTING"
                'axis_exact_match': 'WALL PUNNING WITHOUT PAINTING',
                'axis_display': None
            },
            'Acrylic Emulsion Paint': {
                'icici_keywords': ['luster', 'paint', 'walls'],
                'hdfc_keywords': ['applying', 'acrylic', 'emulsion', 'paint'],
                # Matches: "LUSTER Paint for Walls only in the customer area"
                'axis_exact_match': 'LUSTER Paint for Walls only',
                'axis_display': None
            },
            'Modular Ceiling': {
                'icici_keywords': ['modular', 'grid', 'ceiling'],
                'icici_exclude': ['vertical', 'drops', '150'],
                'hdfc_keywords': ['grid', 'modular', 'false', 'ceiling'],
                # ✅ UPDATED: New search phrase
                'axis_exact_match': 'Rate quoted to include cost of providing support framework formed of',
                # ✅ NEW: Explicitly exclude 150mm items for Axis
                'axis_exclude': ['150mm', '150 mm', 'vertical drops'], 
                'axis_display': None
            }
        
        
        }       

        def match_particular_by_keywords(particular_text, keywords, exclude_keywords=None):
            """Check if particular matches all keywords and doesn't contain exclude keywords"""
            particular_lower = particular_text.lower()

            # Check if all keywords are present
            if not all(keyword.lower() in particular_lower for keyword in keywords):
                return False

            # Check if any exclude keywords are present
            if exclude_keywords:
                if any(exclude_keyword.lower() in particular_lower for exclude_keyword in exclude_keywords):
                    return False

            return True

        # Process each category
        # Process each category
        # Process each category
        for category, mapping in particulars_mapping.items():
            print(f"\n🔍 Processing: {category}")

            result = {
                'Sr_No': sr_no,
                'Particular_Category': category,
                'ICICI_Particulars': 'N/A',
                'ICICI_Rate_Range': 'N/A - N/A',
                'Bank1_Particulars': 'N/A',
                'Bank1_Rate_Range': 'N/A - N/A',
                'Bank2_Particulars': mapping.get('axis_display', 'N/A'),
                'Bank2_Rate_Range': 'N/A - N/A',
                'Variation_Bank1': 0,
                'Variation_Bank2': 0
            }

            # === ICICI DATA ===
            if icici_sqft in icici_validator.sqfeet_ranges:
                range_data = icici_validator.sqfeet_ranges[icici_sqft]
                rate_ranges = range_data.get('rate_ranges', {})

                icici_keywords = mapping.get('icici_keywords', [])
                icici_exclude = mapping.get('icici_exclude', [])

                # ✅ DEBUG: Print first 5 particulars to see what's in the model
                print(f"  📊 ICICI Model has {len(rate_ranges)} particulars in range {icici_sqft}")
                print(f"  🔍 Looking for keywords: {icici_keywords}")

                found_match = False
                for particular, stats in rate_ranges.items():
                    # ✅ DEBUG: Print what we're checking
                    if any(kw in particular for kw in icici_keywords[:2]):  # Print if it has any of first 2 keywords
                        print(f"    🔎 Checking: {particular[:100]}...")
                        print(f"    ✓ Has all keywords? {match_particular_by_keywords(particular, icici_keywords, icici_exclude)}")

                    if match_particular_by_keywords(particular, icici_keywords, icici_exclude):
                        original_particular = stats.get('original_particular', particular)
                        result['ICICI_Particulars'] = original_particular

                        weighted_min = stats.get('weighted_avg_min', stats.get('overall_mean', 0))
                        weighted_max = stats.get('weighted_avg_max', stats.get('overall_mean', 0))
                        result['ICICI_Rate_Range'] = f"₹{icici_validator.format_indian_currency(weighted_min)} - ₹{icici_validator.format_indian_currency(weighted_max)}"
                        print(f"  ✅ ICICI found: {original_particular[:80]}...")
                        found_match = True
                        break

                if not found_match:
                    print(f"  ❌ ICICI: No match found for {category}")

                for particular, stats in rate_ranges.items():
                    if match_particular_by_keywords(particular, icici_keywords, icici_exclude):
                        original_particular = stats.get('original_particular', particular)
                        result['ICICI_Particulars'] = original_particular

                        weighted_min = stats.get('weighted_avg_min', stats.get('overall_mean', 0))
                        weighted_max = stats.get('weighted_avg_max', stats.get('overall_mean', 0))
                        result['ICICI_Rate_Range'] = f"₹{icici_validator.format_indian_currency(weighted_min)} - ₹{icici_validator.format_indian_currency(weighted_max)}"
                        print(f"  ✅ ICICI found: {result['ICICI_Rate_Range']}")
                        break  # ✅ THIS IS CORRECT - break after finding match in ICICI

            # === HDFC DATA ===
            if hdfc_sqft in hdfc_validator.sqfeet_ranges:
                range_data = hdfc_validator.sqfeet_ranges[hdfc_sqft]
                rate_ranges = range_data.get('rate_ranges', {})

                hdfc_keywords = mapping.get('hdfc_keywords', [])

                for particular, stats in rate_ranges.items():
                    if match_particular_by_keywords(particular, hdfc_keywords):
                        original_particular = stats.get('original_particular', particular)
                        result['Bank1_Particulars'] = original_particular

                        weighted_min = stats.get('weighted_avg_min', stats.get('overall_mean', 0))
                        weighted_max = stats.get('weighted_avg_max', stats.get('overall_mean', 0))
                        result['Bank1_Rate_Range'] = f"₹{hdfc_validator.format_indian_currency(weighted_min)} - ₹{hdfc_validator.format_indian_currency(weighted_max)}"
                        print(f"  ✅ HDFC found: {result['Bank1_Rate_Range']}")
                        break  # ✅ THIS IS CORRECT - break after finding match in HDFC

            # === AXIS DATA (Exact Phrase Match) ===
            # === AXIS DATA (Exact Phrase Match) ===
            # Initialize with default N/A or specific display message
            result['Bank2_Particulars'] = mapping.get('axis_display', 'N/A - Not found for this bank')
            result['Bank2_Rate_Range'] = 'N/A - N/A'

            axis_exact_match = mapping.get('axis_exact_match')
            axis_exclude = mapping.get('axis_exclude', []) # Get exclusion list

            # Only search if there is an exact match phrase provided
            if axis_exact_match and axis_sqft in axis_validator.sqfeet_ranges:
                range_data = axis_validator.sqfeet_ranges[axis_sqft]
                rate_ranges = range_data.get('rate_ranges', {})

                for particular, stats in rate_ranges.items():
                    particular_lower = particular.lower().strip()
                    
                    # 1. EXCLUSION CHECK: Skip if it contains any excluded words (like "150mm")
                    if any(ex.lower() in particular_lower for ex in axis_exclude):
                        continue

                    # 2. EXACT PHRASE MATCH CHECK
                    if axis_exact_match.lower().strip() in particular_lower:
                        
                        original_particular = stats.get('original_particular', particular)
                        result['Bank2_Particulars'] = original_particular

                        weighted_min = stats.get('weighted_avg_min', stats.get('overall_mean', 0))
                        weighted_max = stats.get('weighted_avg_max', stats.get('overall_mean', 0))
                        result['Bank2_Rate_Range'] = f"₹{axis_validator.format_indian_currency(weighted_min)} - ₹{axis_validator.format_indian_currency(weighted_max)}"
                        
                        print(f"  ✅ AXIS found (Exact): {result['Bank2_Rate_Range']}")
                        break # Stop after finding the first valid match

            # === CALCULATE VARIATIONS ===
            try:
                def extract_range_avg(range_str):
                    if 'N/A' in range_str:
                        return None
                    clean_str = range_str.replace('₹', '').replace(',', '')
                    numbers = re.findall(r'\d+\.?\d*', clean_str)
                    if len(numbers) >= 2:
                        return (float(numbers[0]) + float(numbers[1])) / 2
                    return None

                icici_avg = extract_range_avg(result['ICICI_Rate_Range'])
                bank1_avg = extract_range_avg(result['Bank1_Rate_Range'])
                bank2_avg = extract_range_avg(result['Bank2_Rate_Range'])

                if icici_avg and bank1_avg:
                    variation = ((icici_avg - bank1_avg) / bank1_avg) * 100
                    result['Variation_Bank1'] = round(variation, 2)
                else:
                    result['Variation_Bank1'] = 0

                if icici_avg and bank2_avg:
                    variation = ((icici_avg - bank2_avg) / bank2_avg) * 100
                    result['Variation_Bank2'] = round(variation, 2)
                else:
                    result['Variation_Bank2'] = 0
            except Exception as e:
                print(f"  ⚠️  Error calculating variation: {e}")
                result['Variation_Bank1'] = 0
                result['Variation_Bank2'] = 0

            # ✅ IMPORTANT: Append result AFTER processing all banks for this category
            results.append(result)
            sr_no += 1
            print(f"  ✅ Added category {sr_no - 1}: {category}")

        # ✅ Final check
        print(f"\n✅ Processed {len(results)} particulars")
        print("="*60 + "\n")

        # Debug: Print what we found
        for r in results:
            print(f"{r['Sr_No']}. {r['Particular_Category']}: ICICI={r['ICICI_Particulars'][:50]}..., HDFC={r['Bank1_Particulars'][:50]}...")

        return render_template('rate_comparison_analysis.html', 
                             rate_data=results,
                             total_items=len(results))
        
    except Exception as e:
        print(f"❌ Error in rate comparison: {str(e)}")
        import traceback
        traceback.print_exc()
        flash(f'Error: {str(e)}')
        return redirect(url_for('upload_file'))

@app.route('/dump_analysis')
def dump_analysis():
    try:
        region = request.args.get('region', 'West')
        mode = request.args.get('mode', 'rate')
        
        # Load Grouped Validators based on selected region
        region_lower = region.lower()

        if region_lower == 'pan india':
            icici_model_key = 'icici_grouped_subcategory_india'
            hdfc_model_key = 'hdfc_grouped_subcategory_india'
        else:
            icici_model_key = f'icici_grouped_subcategory_{region_lower}'
            hdfc_model_key = f'hdfc_grouped_subcategory_{region_lower}'

        icici_validator = get_validator(icici_model_key)
        hdfc_validator = get_validator(hdfc_model_key)

        if not icici_validator or not hdfc_validator:
            flash(f"Grouped models not loaded for {region} region.")
            return redirect(url_for('upload_file'))
        
        # 1. Define Ranges & Data Source Selection
        # Both rate and amount now use 250 sqft intervals
        ranges = ['500-750', '751-1000', '1001-1250', '1251-1500', '1501-1750', 
                  '1751-2000', '2001-2250', '2251-2500', '2501-2750', '2751-3000',
                  '3001-3250', '3251-3500', '3501-3750', '3751-4000', '4001-4250', 
                  '4251-4500', '4501-4750', '4751-5000', '5000+']
        
        # Select correct data source based on mode
        if mode == 'amount':
            icici_data = icici_validator.amount_cwi_ranges
            hdfc_data = hdfc_validator.amount_cwi_ranges
        else:
            icici_data = icici_validator.rate_sqfeet_ranges
            hdfc_data = hdfc_validator.rate_sqfeet_ranges

        # 2. Strict Category List - ✅ FIX: Added "TOTAL"
        categories = [
            "CIVIL & RELATED WORKS",
            "POP & FALSE CEILING WORKS",
            "CARPENTRY AND INTERIOR WORKS",
            "PAINTING WORKS",
            "ROLLING SHUTTER AND MS WORK",
            "ELECTRIFICATION AND ALLIED WORKS",
            "ADDITIONAL WORKS",
            "TOTAL"  # ✅ Added TOTAL as requested
        ]

        report_data = {}

        # 3. Helper with Fuzzy Key Matching
        def get_all_stats(data_source, requested_range_key, category, mode_type):
            try:
                # --- FUZZY MATCH LOGIC ---
                matched_key = None
                if requested_range_key in data_source:
                    matched_key = requested_range_key
                else:
                    req_start = int(requested_range_key.split('-')[0])
                    for actual_key in data_source.keys():
                        try:
                            act_start = int(actual_key.split('-')[0])
                            if abs(req_start - act_start) <= 1:
                                matched_key = actual_key
                                break
                        except: continue
                
                if not matched_key: return None
                    
                range_content = data_source[matched_key]
                
                if category not in range_content['grouped_ranges']: return None
                
                stats = range_content['grouped_ranges'][category]
                
                # ✅ FIX: Select correct stats dictionary based on mode
                if mode_type == 'amount':
                    inner_stats = stats.get('amount_stats', {})
                    if not inner_stats:
                        inner_stats = stats.get('rate_stats', {})
                else:
                    inner_stats = stats.get('rate_stats', {})
                
                if not inner_stats: return None
                
                # --- GET VALUES ---
                w_min = inner_stats.get('weighted_min', 0)
                w_max = inner_stats.get('weighted_max', 0)
                
                # Try to get Actual Minimum found in files
                abs_min = inner_stats.get('abs_min_val', 0)
                if abs_min == 0: abs_min = inner_stats.get('overall_min', inner_stats.get('min', 0))

                abs_max = inner_stats.get('abs_max_val', 0)
                if abs_max == 0: abs_max = inner_stats.get('overall_max', inner_stats.get('max', 0))

                t_min = abs_min if abs_min > 0 else inner_stats.get('threshold_min', 0)
                t_max = abs_max if abs_max > 0 else inner_stats.get('threshold_max', 0)
                
                if abs_min > 0 and w_min < abs_min: w_min = abs_min

                min_files = inner_stats.get('min_threshold_files', [])
                if not min_files: min_files = inner_stats.get('abs_min_file', []) 

                max_files = inner_stats.get('max_threshold_files', [])
                if not max_files: max_files = inner_stats.get('abs_max_file', []) 
                
                if w_min == 0 and w_max == 0:
                    mean_val = inner_stats.get('mean', 0)
                    if mean_val > 0:
                        w_min = mean_val
                        w_max = mean_val
                    else:
                        return None
                
                return {
                    'weighted_min': w_min, 'weighted_max': w_max,
                    'threshold_min': t_min, 'threshold_max': t_max,
                    'min_files': min_files, 'max_files': max_files
                }
            except Exception as e:
                print(f"❌ Exception in get_all_stats: {e}")
                return None

        # 4. Generate Report Data
        for rng in ranges:
            report_data[rng] = []
            sr_no = 1
            
            for cat in categories:
                icici_stats = get_all_stats(icici_data, rng, cat, mode)
                hdfc_stats = get_all_stats(hdfc_data, rng, cat, mode)

                if icici_stats:
                    icici_range_str = f"Rs. {icici_validator.format_indian_currency(icici_stats['weighted_min'])} - {icici_validator.format_indian_currency(icici_stats['weighted_max'])}"
                    icici_threshold_min = icici_validator.format_indian_currency(icici_stats['threshold_min'])
                    icici_threshold_max = icici_validator.format_indian_currency(icici_stats['threshold_max'])
                else:
                    icici_range_str = "N/A"
                    icici_threshold_min = "N/A"
                    icici_threshold_max = "N/A"
                
                if hdfc_stats:
                    hdfc_range_str = f"Rs. {hdfc_validator.format_indian_currency(hdfc_stats['weighted_min'])} - {hdfc_validator.format_indian_currency(hdfc_stats['weighted_max'])}"
                    hdfc_threshold_min = hdfc_validator.format_indian_currency(hdfc_stats['threshold_min'])
                    hdfc_threshold_max = hdfc_validator.format_indian_currency(hdfc_stats['threshold_max'])
                else:
                    hdfc_range_str = "N/A"
                    hdfc_threshold_min = "N/A"
                    hdfc_threshold_max = "N/A"

                # Tooltips
                def format_files(files):
                    if not files: return "No files recorded"
                    if isinstance(files, str): return files 
                    lines = []
                    try:
                        for i, f in enumerate(files[:5]):
                            if isinstance(f, dict):
                                file_name = f.get('file', 'Unknown')
                                particular = f.get('original_particular', f.get('particular', 'N/A'))
                                lines.append(f"{i+1}. {file_name}: {particular}")
                            else:
                                lines.append(f"{i+1}. {str(f)}")
                        if len(files) > 5: lines.append(f"... and {len(files) - 5} more files")
                    except: return "File data format error"
                    return "\n".join(lines)
                
                icici_min_files_tooltip = format_files(icici_stats['min_files']) if icici_stats else ""
                icici_max_files_tooltip = format_files(icici_stats['max_files']) if icici_stats else ""
                hdfc_min_files_tooltip = format_files(hdfc_stats['min_files']) if hdfc_stats else ""
                hdfc_max_files_tooltip = format_files(hdfc_stats['max_files']) if hdfc_stats else ""

                # Variation
                variation = "0%"
                status = "Neutral"
                
                if icici_stats and hdfc_stats:
                    icici_mid = (icici_stats['weighted_min'] + icici_stats['weighted_max']) / 2
                    hdfc_mid = (hdfc_stats['weighted_min'] + hdfc_stats['weighted_max']) / 2
                    if hdfc_mid > 0:
                        var_val = ((icici_mid - hdfc_mid) / hdfc_mid) * 100
                        variation = f"{var_val:.2f}%"
                        if var_val > 0: status = "Higher"
                        elif var_val < 0: status = "Lower"

                row = {
                    'Sr_No': sr_no,
                    'Particulars': cat,
                    'ICICI_Value': icici_range_str,
                    'HDFC_Value': hdfc_range_str,
                    'Threshold_Min_ICICI': icici_threshold_min,
                    'Threshold_Max_ICICI': icici_threshold_max,
                    'Threshold_Min_Bank1': hdfc_threshold_min,
                    'Threshold_Max_Bank1': hdfc_threshold_max,
                    'ICICI_Min_Files': icici_min_files_tooltip,
                    'ICICI_Max_Files': icici_max_files_tooltip,
                    'Bank1_Min_Files': hdfc_min_files_tooltip,
                    'Bank1_Max_Files': hdfc_max_files_tooltip,
                    'Variation': variation,
                    'Status': status
                }
                report_data[rng].append(row)
                sr_no += 1

        return render_template('dump_analysis.html', 
                             report_data=report_data, 
                             ranges=ranges, 
                             current_mode=mode, 
                             current_region=region)

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Error generating dump analysis: {str(e)}")
        return redirect(url_for('upload_file'))

def classify_work_category(min_files_text, max_files_text):
    """
    Classify work category based on keywords in Min_Files and Max_Files text.
    Returns appropriate work category string.
    """
    # Combine both texts for searching
    combined_text = f"{min_files_text} {max_files_text}".lower()
    
    # Define category patterns (order matters - check specific ones first)
    category_patterns = [
        # Pattern 1: Civil and Structure
        {
            'keywords': ['c&s', 'civil & structure', 'civil and structure', 'structure', 'struct'],
            'category': 'Civil and Structure'
        },
        # Pattern 2: Fire Protection System - FPS
        {
            'keywords': ['fire protection system', 'fps', 'fire protection'],
            'category': 'Fire Protection System - FPS'
        },
        # Pattern 3: Fire Tube Suppression
        {
            'keywords': ['fire tube suppression', 'fire tube'],
            'category': 'Fire Tube Suppression'
        },
        # Pattern 4: Fire Suppression
        {
            'keywords': ['fire suppression', 'fire supression', 'suppres', 'suppr'],
            'category': 'Fire Suppression'
        },
        # Pattern 5: Civil Works
        {
            'keywords': ['civil', 'cvl', 'clv', ' cl '],
            'category': 'Civil Works'
        },
        # Pattern 6: HVAC
        {
            'keywords': ['hvac'],
            'category': 'HVAC'
        },
        # Pattern 7: Electrical Works
        {
            'keywords': ['electrical', 'ele', 'electric'],
            'category': 'Electrical Works'
        },
        # Pattern 8: Plumbing Works
        {
            'keywords': ['plumbing', 'plmb', 'plum', 'pmb'],
            'category': 'Plumbing Works'
        },
        # Pattern 9: ODO
        {
            'keywords': ['odo'],
            'category': 'ODO'
        },
        # Pattern 10: NSO Smart
        {
            'keywords': ['nso smart', 'nso'],
            'category': 'NSO Smart'
        }
    ]
    
    # Check each pattern
    for pattern in category_patterns:
        for keyword in pattern['keywords']:
            if keyword in combined_text:
                return pattern['category']
    
    # If no match found, return "Others"
    return 'Others'

@app.route('/reliance_dump_analysis')
def reliance_dump_analysis():
    """Display Reliance Yousta service code dump analysis without file upload"""
    try:
        # Load Reliance validator
        reliance_validator = get_validator('reliance_qty_yousta')
        
        if not reliance_validator:
            flash("Reliance Yousta model not available.")
            return redirect(url_for('upload_file'))
        
        # Get all ranges from the model
        ranges = list(reliance_validator.sqfeet_ranges.keys())
        
        report_data = []
        excel_data = []  # Separate list for Excel with file names
        sr_no = 1
        
        # Iterate through each range
        for range_name in ranges:
            range_data = reliance_validator.sqfeet_ranges[range_name]
            
            # Get quantity_ranges (which contains service codes)
            if 'quantity_ranges' not in range_data:
                continue
            
            quantity_ranges = range_data['quantity_ranges']
            
            # Sort service codes in ascending order
            sorted_service_codes = sorted(quantity_ranges.keys())
            
            for service_code in sorted_service_codes:
                stats = quantity_ranges[service_code]
                
                # ✅ FIX: Get particular from the list stored in training
                particulars_list = stats.get('particulars', [])
                if particulars_list and len(particulars_list) > 0:
                    # Pick the first particular from the list
                    particular = particulars_list[0]
                else:
                    particular = 'N/A'
                
                num_files = stats.get('total_count', 0)
                w_min = stats.get('weighted_avg_min', 0)
                w_max = stats.get('weighted_avg_max', 0)
                t_min = stats.get('threshold_min', stats.get('min', 0))
                t_max = stats.get('threshold_max', stats.get('max', 0))
                
                # Get file data for threshold min/max
                min_files = stats.get('min_files', [])
                max_files = stats.get('max_files', [])
                
                # ✅ Extract file names for EXCEL only
                def extract_file_names_for_excel(files_list):
                    if not files_list:
                        return 'N/A'
                    names = []
                    for f in files_list[:10]:  # Limit to 10 files
                        if isinstance(f, dict):
                            file_name = f.get('file', 'Unknown')
                            names.append(file_name)
                        else:
                            names.append(str(f))
                    return ', '.join(names)
                
                min_file_names = extract_file_names_for_excel(min_files)
                max_file_names = extract_file_names_for_excel(max_files)
                
                # Format tooltips for WEBPAGE (hover only)
                def format_files_tooltip(files):
                    if not files: return "No files recorded"
                    lines = []
                    for i, f in enumerate(files[:5]):
                        if isinstance(f, dict):
                            file_name = f.get('file', 'Unknown')
                            lines.append(f"{i+1}. {file_name}")
                        else:
                            lines.append(f"{i+1}. {str(f)}")
                    if len(files) > 5: 
                        lines.append(f"... and {len(files) - 5} more files")
                    return "\n".join(lines)
                
                min_tooltip = format_files_tooltip(min_files)
                max_tooltip = format_files_tooltip(max_files)
                
                # Round to 2 decimals for Reliance
                w_min = round(w_min, 2)
                w_max = round(w_max, 2)
                t_min = round(t_min, 2)
                t_max = round(t_max, 2)
                
                # ✅ WEB DISPLAY DATA (no file name columns)
                web_row = {
                    'Sr_No': sr_no,
                    'Service_Code': service_code,
                    'Particulars': particular,
                    'Area_Range': range_name,
                    'Num_Files': num_files,
                    'Weighted_Min': w_min,
                    'Weighted_Max': w_max,
                    'Weighted_Range': f"{w_min} - {w_max}",
                    'Threshold_Min': t_min,
                    'Threshold_Max': t_max,
                    'Min_Tooltip': min_tooltip,
                    'Max_Tooltip': max_tooltip
                }
                report_data.append(web_row)
                
                # ✅ Classify work category based on file names
                work_category = classify_work_category(min_file_names, max_file_names)
                
                # ✅ EXCEL DATA (with file name columns AND work category)
                excel_row = {
                    'Sr_No': sr_no,
                    'Service_Code': service_code,
                    'Particulars': particular,
                    'Work_Category': work_category,  # ✅ NEW COLUMN
                    'Area_Range': range_name,
                    'Num_Files': num_files,
                    'Weighted_Min': w_min,
                    'Weighted_Max': w_max,
                    'Weighted_Range': f"{w_min} - {w_max}",
                    'Threshold_Min': t_min,
                    'Threshold_Max': t_max,
                    'Min_Files': min_file_names,
                    'Max_Files': max_file_names
                }
                excel_data.append(excel_row)
                               
                sr_no += 1
        
        # ✅ Create Excel report with file names and styling
        timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f'reliance_yousta_dump_{timestamp}.xlsx'
        excel_filepath = os.path.join(app.config['OUTPUT_FOLDER'], excel_filename)
        
        # Create DataFrame from EXCEL data (includes file names and work category)
        df = pd.DataFrame(excel_data)
        
        # Reorder columns for Excel (Work_Category after Particulars)
        excel_columns = ['Sr_No', 'Service_Code', 'Particulars', 'Work_Category', 'Area_Range', 
                        'Num_Files', 'Weighted_Min', 'Weighted_Max', 'Weighted_Range', 
                        'Threshold_Min', 'Threshold_Max', 'Min_Files', 'Max_Files']
        df = df[excel_columns]
        
        # Save to Excel with styling
        with pd.ExcelWriter(excel_filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Service Codes Analysis')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Service Codes Analysis']
            
            # Import styling modules
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Header styling - make headers bold with blue background
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths for better readability
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 to prevent extremely wide columns
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze the header row
            worksheet.freeze_panes = 'A2'
        
        print(f"✅ Excel report saved: {excel_filename}")
        
        return render_template('reliance_dump_analysis.html', 
                             report_data=report_data,  # ✅ Web data (no file columns)
                             total_items=len(report_data),
                             excel_report=excel_filename)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Error: {str(e)}")
        return redirect(url_for('upload_file'))


@app.route('/reliance_smart_dump_analysis')
def reliance_smart_dump_analysis():
    """Display Reliance Smart service code dump analysis without file upload"""
    try:
        # Load Smart validator
        smart_validator = get_validator('reliance_qty_smart')
        
        if not smart_validator:
            flash("Reliance Smart model not available.")
            return redirect(url_for('upload_file'))
        
        # Get all ranges from the model
        ranges = list(smart_validator.sqfeet_ranges.keys())
        
        report_data = []
        excel_data = []
        sr_no = 1
        
        # Iterate through each range
        for range_name in ranges:
            range_data = smart_validator.sqfeet_ranges[range_name]
            
            # Get quantity_ranges (which contains service codes)
            if 'quantity_ranges' not in range_data:
                continue
            
            quantity_ranges = range_data['quantity_ranges']
            
            # Sort service codes in ascending order
            sorted_service_codes = sorted(quantity_ranges.keys())
            
            for service_code in sorted_service_codes:
                stats = quantity_ranges[service_code]
                
                # Get particular from the list stored in training
                particulars_list = stats.get('particulars', [])
                if particulars_list and len(particulars_list) > 0:
                    particular = particulars_list[0]
                else:
                    particular = 'N/A'
                
                num_files = stats.get('total_count', 0)
                w_min = stats.get('weighted_avg_min', 0)
                w_max = stats.get('weighted_avg_max', 0)
                t_min = stats.get('threshold_min', stats.get('min', 0))
                t_max = stats.get('threshold_max', stats.get('max', 0))
                
                # Get file data for threshold min/max
                min_files = stats.get('min_files', [])
                max_files = stats.get('max_files', [])
                
                # Extract file names for EXCEL only
                def extract_file_names_for_excel(files_list):
                    if not files_list:
                        return 'N/A'
                    names = []
                    for f in files_list[:10]:
                        if isinstance(f, dict):
                            file_name = f.get('file', 'Unknown')
                            names.append(file_name)
                        else:
                            names.append(str(f))
                    return ', '.join(names)
                
                min_file_names = extract_file_names_for_excel(min_files)
                max_file_names = extract_file_names_for_excel(max_files)
                
                # Format tooltips for WEBPAGE (hover only)
                def format_files_tooltip(files):
                    if not files: return "No files recorded"
                    lines = []
                    for i, f in enumerate(files[:5]):
                        if isinstance(f, dict):
                            file_name = f.get('file', 'Unknown')
                            lines.append(f"{i+1}. {file_name}")
                        else:
                            lines.append(f"{i+1}. {str(f)}")
                    if len(files) > 5: 
                        lines.append(f"... and {len(files) - 5} more files")
                    return "\n".join(lines)
                
                min_tooltip = format_files_tooltip(min_files)
                max_tooltip = format_files_tooltip(max_files)
                
                # Round to 2 decimals for Smart
                w_min = round(w_min, 2)
                w_max = round(w_max, 2)
                t_min = round(t_min, 2)
                t_max = round(t_max, 2)
                
                # WEB DISPLAY DATA (no file name columns)
                web_row = {
                    'Sr_No': sr_no,
                    'Service_Code': service_code,
                    'Particulars': particular,
                    'Area_Range': range_name,
                    'Num_Files': num_files,
                    'Weighted_Min': w_min,
                    'Weighted_Max': w_max,
                    'Weighted_Range': f"{w_min} - {w_max}",
                    'Threshold_Min': t_min,
                    'Threshold_Max': t_max,
                    'Min_Tooltip': min_tooltip,
                    'Max_Tooltip': max_tooltip
                }
                report_data.append(web_row)
                
                # ✅ Classify work category based on file names
                work_category = classify_work_category(min_file_names, max_file_names)
                
                # ✅ EXCEL DATA (with file name columns AND work category)
                excel_row = {
                    'Sr_No': sr_no,
                    'Service_Code': service_code,
                    'Particulars': particular,
                    'Work_Category': work_category,  # ✅ NEW COLUMN
                    'Area_Range': range_name,
                    'Num_Files': num_files,
                    'Weighted_Min': w_min,
                    'Weighted_Max': w_max,
                    'Weighted_Range': f"{w_min} - {w_max}",
                    'Threshold_Min': t_min,
                    'Threshold_Max': t_max,
                    'Min_Files': min_file_names,
                    'Max_Files': max_file_names
                }
                excel_data.append(excel_row)
                               
                sr_no += 1
        
        # Create Excel report with file names and styling
        timestamp = pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f'reliance_smart_dump_{timestamp}.xlsx'
        excel_filepath = os.path.join(app.config['OUTPUT_FOLDER'], excel_filename)
        
        # Create DataFrame from EXCEL data (includes file names and work category)
        df = pd.DataFrame(excel_data)
        
        # Reorder columns for Excel (Work_Category after Particulars)
        excel_columns = ['Sr_No', 'Service_Code', 'Particulars', 'Work_Category', 'Area_Range', 
                        'Num_Files', 'Weighted_Min', 'Weighted_Max', 'Weighted_Range', 
                        'Threshold_Min', 'Threshold_Max', 'Min_Files', 'Max_Files']
        df = df[excel_columns]
        
        # Save to Excel with styling
        with pd.ExcelWriter(excel_filepath, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Service Codes Analysis')
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Service Codes Analysis']
            
            # Import styling modules
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Header styling - make headers bold with blue background
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths for better readability
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 to prevent extremely wide columns
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze the header row
            worksheet.freeze_panes = 'A2'
        
        print(f"✅ Excel report saved: {excel_filename}")
        
        return render_template('reliance_dump_analysis_smart.html', 
                             report_data=report_data,
                             total_items=len(report_data),
                             excel_report=excel_filename)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Error: {str(e)}")
        return redirect(url_for('upload_file'))

# --- HELPER: Calculate Stats for Sub-Categories on the Fly ---
def calculate_subcategory_stats(data_list, validator):
    """
    Takes a list of dictionaries [{'val': 100, 'file': '...'}, ...] 
    and calculates the Weighted Range (Mean ± 1SD), ignoring 0s.
    """
    if not data_list:
        return "N/A"
    
    # Extract values, filtering out 0s
    # The new training code stores 'val' as the TOTAL aggregated rate for that subcat in that file
    values = [float(d['val']) for d in data_list if float(d['val']) > 0]
    
    if not values:
        return "N/A"
    
    # Convert to numpy array
    arr = np.array(values)
    
    mean_val = np.mean(arr)
    std_val = np.std(arr)
    actual_min = np.min(arr)
    
    # Weighted Range Calculation (Clamped to Actual Min)
    calc_min = mean_val - std_val
    weighted_min = max(actual_min, calc_min)
    weighted_max = mean_val + std_val
    
    return f"Rs. {validator.format_indian_currency(weighted_min)} - {validator.format_indian_currency(weighted_max)}"

@app.route('/subcategory_details')
def subcategory_details():
    try:
        # 1. Get Parameters
        range_name = request.args.get('range')
        category_name = request.args.get('category')
        region = request.args.get('region', 'West')
        mode = request.args.get('mode', 'rate')
        
        # 2. Load Validators
        region_lower = region.lower()
        if region_lower == 'pan india':
            icici_key = 'icici_grouped_subcategory_india'
            hdfc_key = 'hdfc_grouped_subcategory_india'
        else:
            icici_key = f'icici_grouped_subcategory_{region_lower}'
            hdfc_key = f'hdfc_grouped_subcategory_{region_lower}'
            
        icici_validator = get_validator(icici_key)
        hdfc_validator = get_validator(hdfc_key)
        
        if not icici_validator or not hdfc_validator:
            flash("Validators not loaded.")
            return redirect(url_for('dump_analysis'))

        # 3. Select Data Source based on Mode
        if mode == 'amount':
            icici_source = icici_validator.amount_per_cwi_250_ranges 
            hdfc_source = hdfc_validator.amount_per_cwi_250_ranges
            range_label = "Amount (250 sq.ft Bucket)"
        else:
            icici_source = icici_validator.rate_per_sqft_250_ranges
            hdfc_source = hdfc_validator.rate_per_sqft_250_ranges
            range_label = "Rate/Sq.Ft (250 sq.ft Bucket)"

        # 🔍 DEBUG: Print what's actually in the model
        print(f"\n{'='*80}")
        print(f"DEBUG SUBCATEGORY DETAILS")
        print(f"{'='*80}")
        print(f"Mode: {mode}")
        print(f"Range requested: {range_name}")
        print(f"Category: {category_name}")
        print(f"\n--- ICICI Source Structure ---")
        print(f"Available ranges in ICICI: {list(icici_source.keys())}")
        
        if range_name in icici_source:
            print(f"\n✓ Range '{range_name}' found in ICICI")
            print(f"Keys in this range: {list(icici_source[range_name].keys())}")
            
            if 'grouped_ranges' in icici_source[range_name]:
                print(f"Groups in grouped_ranges: {list(icici_source[range_name]['grouped_ranges'].keys())}")
                
                if category_name in icici_source[range_name]['grouped_ranges']:
                    print(f"\n✓ Category '{category_name}' found in grouped_ranges")
                    print(f"Keys in category: {list(icici_source[range_name]['grouped_ranges'][category_name].keys())}")
                    
                    if 'sub_stats' in icici_source[range_name]['grouped_ranges'][category_name]:
                        sub_stats = icici_source[range_name]['grouped_ranges'][category_name]['sub_stats']
                        print(f"✓ sub_stats found! Subcategories: {list(sub_stats.keys())[:5]}...")
                        print(f"Total subcategories: {len(sub_stats)}")
                    else:
                        print("✗ 'sub_stats' key not found in category")
                else:
                    print(f"✗ Category '{category_name}' not found in grouped_ranges")
            
            if 'subcategory_data' in icici_source[range_name]:
                print(f"\nGroups in subcategory_data: {list(icici_source[range_name]['subcategory_data'].keys())}")
                
                if category_name in icici_source[range_name]['subcategory_data']:
                    print(f"✓ Category '{category_name}' found in subcategory_data")
                    subcats = icici_source[range_name]['subcategory_data'][category_name]
                    print(f"Subcategories: {list(subcats.keys())[:5]}...")
                    print(f"Total subcategories: {len(subcats)}")
                else:
                    print(f"✗ Category '{category_name}' not found in subcategory_data")
        else:
            print(f"✗ Range '{range_name}' not found in ICICI")
        
        print(f"{'='*80}\n")

        

        # 4. Helper to find fuzzy key
        def find_key(source, req_key):
            if req_key in source: return req_key
            try:
                req_start = int(req_key.split('-')[0])
                for act_key in source.keys():
                    try:
                        act_start = int(act_key.split('-')[0])
                        if abs(req_start - act_start) <= 5: return act_key
                    except: continue
            except: pass
            return None

        icici_range_key = find_key(icici_source, range_name)
        hdfc_range_key = find_key(hdfc_source, range_name)

        # 5. Extract Sub-Category Data
        icici_subcats = {}
        if icici_range_key:
            if icici_range_key in icici_source:
                if mode == 'amount':
                    # AMOUNT MODE: Data is in grouped_ranges -> group -> sub_stats (already calculated stats)
                    grouped_data = icici_source[icici_range_key].get('grouped_ranges', {})
                    if category_name in grouped_data:
                        icici_subcats = grouped_data[category_name].get('sub_stats', {})
                else:
                    # RATE MODE: Data is in subcategory_data -> group -> subcat (list of values)
                    subcat_data = icici_source[icici_range_key].get('subcategory_data', {})
                    if category_name in subcat_data:
                        icici_subcats = subcat_data[category_name]

        hdfc_subcats = {}
        if hdfc_range_key:
            if hdfc_range_key in hdfc_source:
                if mode == 'amount':
                    # AMOUNT MODE: Data is in grouped_ranges -> group -> sub_stats (already calculated stats)
                    grouped_data = hdfc_source[hdfc_range_key].get('grouped_ranges', {})
                    if category_name in grouped_data:
                        hdfc_subcats = grouped_data[category_name].get('sub_stats', {})
                else:
                    # RATE MODE: Data is in subcategory_data -> group -> subcat (list of values)
                    subcat_data = hdfc_source[hdfc_range_key].get('subcategory_data', {})
                    if category_name in subcat_data:
                        hdfc_subcats = hdfc_source[hdfc_range_key]['subcategory_data'][category_name]

        # 6. Align and Categorize WITH DEDUPLICATION
        all_sub_names = set(icici_subcats.keys()) | set(hdfc_subcats.keys())

        common_rows = []
        extra_icici_rows = []
        extra_hdfc_rows = []

        # ✅ STEP 1: First pass - identify what goes in "Common"
        common_names = set()
        for sub_name in sorted(all_sub_names):
            in_icici = sub_name in icici_subcats
            in_hdfc = sub_name in hdfc_subcats

            if in_icici and in_hdfc:
                common_names.add(sub_name.lower().strip())  # Store normalized version

        # ✅ STEP 2: Second pass - categorize with smart deduplication
        for sub_name in sorted(all_sub_names):
            i_data = icici_subcats.get(sub_name, [] if mode == 'rate' else {})
            h_data = hdfc_subcats.get(sub_name, [] if mode == 'rate' else {})

            # Calculate ranges based on mode
            if mode == 'amount':
                if i_data and isinstance(i_data, dict):
                    i_min = i_data.get('weighted_min', 0)
                    i_max = i_data.get('weighted_max', 0)
                    i_range = f"Rs. {icici_validator.format_indian_currency(i_min)} - {icici_validator.format_indian_currency(i_max)}" if i_min > 0 else "N/A"
                else:
                    i_range = "N/A"

                if h_data and isinstance(h_data, dict):
                    h_min = h_data.get('weighted_min', 0)
                    h_max = h_data.get('weighted_max', 0)
                    h_range = f"Rs. {hdfc_validator.format_indian_currency(h_min)} - {hdfc_validator.format_indian_currency(h_max)}" if h_min > 0 else "N/A"
                else:
                    h_range = "N/A"
            else:
                i_range = calculate_subcategory_stats(i_data, icici_validator)
                h_range = calculate_subcategory_stats(h_data, hdfc_validator)

            row = {
                'name': sub_name,
                'icici_range': i_range,
                'hdfc_range': h_range
            }

            in_icici = sub_name in icici_subcats
            in_hdfc = sub_name in hdfc_subcats

            # ✅ NEW LOGIC: Check if similar name exists in common (fuzzy match)
            def is_duplicate_of_common(name):
                """Check if this name is a variation of something already in common"""
                normalized = name.lower().strip().replace(' ', '').replace('/', '').replace('-', '')
                
                # ✅ NEW: Special handling for Panelling/Paneling variations
                # Normalize double-L to single-L
                if 'panelling' in normalized:
                    normalized = normalized.replace('panelling', 'paneling')
                
                for common_name in common_names:
                    common_normalized = common_name.lower().replace(' ', '').replace('/', '').replace('-', '')
                    
                    # ✅ NEW: Also normalize common_name for panelling
                    if 'panelling' in common_normalized:
                        common_normalized = common_normalized.replace('panelling', 'paneling')
                    
                    # If exactly same, treat as duplicate
                    if normalized == common_normalized:
                        return True
                    
                    # Check if one contains the other
                    if normalized in common_normalized or common_normalized in normalized:
                        return True
                    
                    # ✅ NEW: Special case - if both contain "panel" and "work", treat as duplicate
                    if 'panel' in normalized and 'work' in normalized and 'panel' in common_normalized and 'work' in common_normalized:
                        return True
                
                return False

            if in_icici and in_hdfc:
                # Both banks have it - goes to Common
                common_rows.append(row)
            elif in_icici:
                # Only ICICI has it
                # ✅ CHECK: Is this a duplicate/variation of something in Common?
                if not is_duplicate_of_common(sub_name):
                    extra_icici_rows.append(row)
                else:
                    print(f"  🔄 Skipping duplicate in ICICI: {sub_name} (already in Common)")
            elif in_hdfc:
                # Only HDFC has it
                # ✅ CHECK: Is this a duplicate/variation of something in Common?
                if not is_duplicate_of_common(sub_name):
                    extra_hdfc_rows.append(row)
                else:
                    print(f"  🔄 Skipping duplicate in HDFC: {sub_name} (already in Common)")

        return render_template(
            'subcategory_details.html',
            range_name=range_name,
            category_name=category_name,
            region=region,
            mode=mode,
            range_label=range_label,
            common_rows=common_rows,
            extra_icici_rows=extra_icici_rows,
            extra_hdfc_rows=extra_hdfc_rows
        )

    except Exception as e:
        print(f"Error in subcategory details: {e}")
        import traceback
        traceback.print_exc()
        flash("Error loading details.")
        return redirect(url_for('dump_analysis'))

@app.route('/total_breakdown')
def total_breakdown():
    try:
        region = request.args.get('region', 'West')
        mode = request.args.get('mode', 'rate')
        range_name = request.args.get('range')
        
        # Load validators
        region_lower = region.lower()
        if region_lower == 'pan india':
            icici_key = 'icici_grouped_subcategory_india'
            hdfc_key = 'hdfc_grouped_subcategory_india'
        else:
            icici_key = f'icici_grouped_subcategory_{region_lower}'
            hdfc_key = f'hdfc_grouped_subcategory_{region_lower}'
            
        icici_validator = get_validator(icici_key)
        hdfc_validator = get_validator(hdfc_key)
        
        if not icici_validator or not hdfc_validator:
            flash("Validators not loaded.")
            return redirect(url_for('dump_analysis'))

        # Select data source
        if mode == 'amount':
            icici_data = icici_validator.amount_cwi_ranges
            hdfc_data = hdfc_validator.amount_cwi_ranges
        else:
            icici_data = icici_validator.rate_sqfeet_ranges
            hdfc_data = hdfc_validator.rate_sqfeet_ranges

        # Find range (fuzzy match)
        def find_key(source, req_key):
            if req_key in source: return req_key
            try:
                req_start = int(req_key.split('-')[0])
                for act_key in source.keys():
                    try:
                        act_start = int(act_key.split('-')[0])
                        if abs(req_start - act_start) <= 5: return act_key
                    except: continue
            except: pass
            return None

        icici_range_key = find_key(icici_data, range_name)
        hdfc_range_key = find_key(hdfc_data, range_name)

        if not icici_range_key or not hdfc_range_key:
            flash("Range not found in model.")
            return redirect(url_for('dump_analysis'))

        # Categories to include (exclude TOTAL)
        categories = [
            "CIVIL & RELATED WORKS",
            "POP & FALSE CEILING WORKS",
            "CARPENTRY AND INTERIOR WORKS",
            "PAINTING WORKS",
            "ROLLING SHUTTER AND MS WORK",
            "ELECTRIFICATION AND ALLIED WORKS",
            "ADDITIONAL WORKS"
        ]

        # Extract values
        breakdown_data = []
        icici_total = 0
        hdfc_total = 0

        for cat in categories:
            icici_mid = 0
            hdfc_mid = 0

            # Get ICICI value
            if icici_range_key in icici_data:
                range_content = icici_data[icici_range_key]
                if cat in range_content.get('grouped_ranges', {}):
                    stats = range_content['grouped_ranges'][cat]
                    if mode == 'amount':
                        # Amount data is stored under 'amount_stats' key inside rate_sqfeet_ranges
                        inner_stats = stats.get('amount_stats', {})
                        # Fallback to rate_stats if amount_stats not present
                        if not inner_stats:
                            inner_stats = stats.get('rate_stats', {})
                    else:
                        inner_stats = stats.get('rate_stats', {})
                    
                    w_min = inner_stats.get('weighted_min', 0)
                    w_max = inner_stats.get('weighted_max', 0)
                    icici_mid = (w_min + w_max) / 2 if (w_min + w_max) > 0 else 0

            # Get HDFC value
            if hdfc_range_key in hdfc_data:
                range_content = hdfc_data[hdfc_range_key]
                if cat in range_content.get('grouped_ranges', {}):
                    stats = range_content['grouped_ranges'][cat]
                    if mode == 'amount':
                        inner_stats = stats.get('amount_stats', {})
                    else:
                        inner_stats = stats.get('rate_stats', {})
                    
                    w_min = inner_stats.get('weighted_min', 0)
                    w_max = inner_stats.get('weighted_max', 0)
                    hdfc_mid = (w_min + w_max) / 2 if (w_min + w_max) > 0 else 0

            icici_total += icici_mid
            hdfc_total += hdfc_mid

            breakdown_data.append({
                'category': cat,
                'icici_value': icici_mid,
                'hdfc_value': hdfc_mid
            })

        # Calculate percentages
        for row in breakdown_data:
            row['icici_percent'] = (row['icici_value'] / icici_total * 100) if icici_total > 0 else 0
            row['hdfc_percent'] = (row['hdfc_value'] / hdfc_total * 100) if hdfc_total > 0 else 0
            row['icici_display'] = icici_validator.format_indian_currency(row['icici_value'])
            row['hdfc_display'] = hdfc_validator.format_indian_currency(row['hdfc_value'])

        return render_template('total_breakdown.html',
                             breakdown_data=breakdown_data,
                             range_name=range_name,
                             region=region,
                             mode=mode,
                             icici_total=icici_validator.format_indian_currency(icici_total),
                             hdfc_total=hdfc_validator.format_indian_currency(hdfc_total))

    except Exception as e:
        import traceback
        traceback.print_exc()
        flash(f"Error: {str(e)}")
        return redirect(url_for('dump_analysis'))
@app.route('/switch_mode/<mode>')
def switch_mode(mode):
    """Switch validation mode and re-process the last uploaded file"""
    try:
        print("\n" + "="*60)
        print(f"🔄 MODE SWITCH REQUEST: {mode}")
        print("="*60)
        
        # Get stored file info from session
        filepath = session.get('last_uploaded_file')
        client_name = session.get('last_client', 'HDFC Bank')
        selected_region = session.get('last_region', '')

        is_regional = False
        
        print(f"📂 Session filepath: {filepath}")
        print(f"🏦 Session client: {client_name}")
        
        if not filepath:
            print("❌ No filepath in session!")
            flash('No file found in session. Please upload a new file.')
            return redirect(url_for('upload_file'))
            
        if not os.path.exists(filepath):
            print(f"❌ File does not exist at: {filepath}")
            flash('Uploaded file no longer exists. Please upload a new file.')
            return redirect(url_for('upload_file'))
        
        print(f"✅ File exists: {filepath}")
        
        # Build model key based on client and mode
        if client_name == 'HDFC Bank':
            if mode == 'quantity':
                model_key = 'hdfc_qty'
            elif mode == 'amount':
                model_key = 'hdfc_amount'
            elif mode == 'rate':
                model_key = 'hdfc_rate'
            elif mode == 'grouped':
                model_key = 'hdfc_grouped'
            else:
                model_key = 'hdfc_qty'

        elif client_name == 'ICICI Bank':
            # ✅ Check if region is selected for ICICI
            is_regional = False
            if selected_region and selected_region in ['east', 'west', 'north', 'south']:
                is_regional = True
                
                # ⚠️ PER USER REQUEST: Regional Quantity models are not trained yet.
                if mode != 'rate':
                    print(f"⚠️ Mode '{mode}' requested, but Regional Quantity model is not ready. Switching to 'rate'.")
                    mode = 'rate' 

                model_key = f'icici_rate_{selected_region}'

            else:
                # Use default ICICI models (no region selected)
                if mode == 'quantity':
                    model_key = 'icici_qty'
                elif mode == 'amount':
                    model_key = 'icici_amount'
                elif mode == 'rate':
                    model_key = 'icici_rate'
                elif mode == 'grouped':
                    model_key = 'icici_grouped'
                else:
                    model_key = 'icici_qty'
        else:
            model_key = 'hdfc_qty'
        
        print(f"🔑 Using model key: {model_key}")
        
        # Handle grouped mode
        if mode == 'grouped':
            VALIDATOR = get_validator(model_key)
            if VALIDATOR is None:
                print(f"❌ Grouped model not loaded")
                flash(f'Grouped validation model is not available.')
                return redirect(url_for('upload_file'))

            print(f"✅ Grouped validator loaded")
            
            print(f"🔍 Starting grouped validation...")
            results_data, error_msg = validate_grouped_file(VALIDATOR, filepath)
        else:
            VALIDATOR = get_validator(model_key)
            if VALIDATOR is None:
                print(f"❌ Model not loaded: {model_key}")
                flash(f'Validation model for {client_name} - {mode} is not available.')
                return redirect(url_for('upload_file'))

            print(f"✅ Validator loaded: {model_key}")
            
            # Validate file with mode parameter
            print(f"🔍 Starting validation with {mode} mode...")
            results_data, error_msg = VALIDATOR.validate_test_file(filepath, mode=mode, is_regional=is_regional,model_key=model_key)
        
        if error_msg:
            print(f"❌ Validation error: {error_msg}")
            flash(error_msg)
            return redirect(url_for('upload_file'))
        
        if results_data is None:
            print("❌ No results returned")
            flash("Validation failed - no results returned")
            return redirect(url_for('upload_file'))
        
        print(f"✅ Validation successful! Results: {len(results_data)} rows")
        
        # Process results
        results_df = pd.DataFrame(results_data)

        if mode not in ['quantity', 'amount', 'grouped'] and 'Threshold_Min' not in results_df.columns:
            results_df['Threshold_Min'], results_df['Threshold_Max'] = zip(
                *results_df['Validation_Message'].map(extract_thresholds)
            )
        
        # Save reports
        filename = os.path.basename(filepath)
        full_report_filename = f"validated_{mode}_{filename}"
        full_report_filepath = os.path.join(app.config['OUTPUT_FOLDER'], full_report_filename)
        
        if mode != 'grouped':
            styled_df = results_df.style.apply(highlight_invalid, axis=1)
            styled_df.to_excel(full_report_filepath, engine='xlsxwriter', index=False)
        else:
            results_df.to_excel(full_report_filepath, index=False)
        
        print(f"✅ Report saved: {full_report_filename}")
        
        invalid_df = results_df[results_df['Is_Valid'] == False].copy()
        invalid_report_filename = f"invalid_{mode}_{filename}"
        if not invalid_df.empty:
            invalid_report_filepath = os.path.join(app.config['OUTPUT_FOLDER'], invalid_report_filename)
            invalid_df.to_excel(invalid_report_filepath, index=False)
            print(f"✅ Invalid report saved: {invalid_report_filename}")
        
        # Prepare display
        display_df = results_df.copy()
        
        if mode == 'grouped':
            display_df['As_Per_CWI_Amount'] = display_df['As_Per_CWI_Amount'].apply(
                lambda x: VALIDATOR.format_indian_currency(x))
            display_df['Rate_per_sqfeet'] = display_df['Rate_per_sqfeet'].apply(
                lambda x: VALIDATOR.format_indian_currency(x))

            def make_range_sw(row, min_col, max_col):
                mn = row.get(min_col, 'NA')
                mx = row.get(max_col, 'NA')
                if mn == 'NA' or mx == 'NA' or pd.isna(mn) or pd.isna(mx):
                    return 'NA'
                return f"Rs. {VALIDATOR.format_indian_currency(mn)} - Rs. {VALIDATOR.format_indian_currency(mx)}"

            display_df['CWI_Weighted_Range'] = display_df.apply(
                lambda row: make_range_sw(row, 'CWI_Weighted_Min', 'CWI_Weighted_Max'), axis=1)
            display_df['Rate_Weighted_Range'] = display_df.apply(
                lambda row: make_range_sw(row, 'Rate_Weighted_Min', 'Rate_Weighted_Max'), axis=1)

            def make_threshold_clickable_sw(row, val_col, files_col):
                val = row.get(val_col, 'NA')
                if val == 'NA' or pd.isna(val): return 'NA'
                display_val = f"Rs. {VALIDATOR.format_indian_currency(val)}"
                files_data = row.get(files_col, [])
                if not files_data or not isinstance(files_data, list):
                    return display_val
                tooltip_lines = []
                for i, f in enumerate(files_data[:5]):
                    if isinstance(f, dict):
                        fname = f.get('file', f.get('original_particular', 'Unknown'))
                    else:
                        fname = str(f)
                    tooltip_lines.append(f"{i+1}. {fname}")
                if len(files_data) > 5:
                    tooltip_lines.append(f"... and {len(files_data)-5} more")
                tooltip_text = html.escape("\\n".join(tooltip_lines))
                return (f'<span class="threshold-clickable" title="{tooltip_text}" '
                        f'style="cursor:pointer;color:#007bff;text-decoration:underline;">'
                        f'{display_val}</span>')

            display_df['Threshold_Min_Amount'] = display_df.apply(
                lambda row: make_threshold_clickable_sw(row, 'CWI_Threshold_Min', 'CWI_Min_Files'), axis=1)
            display_df['Threshold_Max_Amount'] = display_df.apply(
                lambda row: make_threshold_clickable_sw(row, 'CWI_Threshold_Max', 'CWI_Max_Files'), axis=1)
            display_df['Threshold_Min_Rate'] = display_df.apply(
                lambda row: make_threshold_clickable_sw(row, 'Rate_Threshold_Min', 'Rate_Min_Files'), axis=1)
            display_df['Threshold_Max_Rate'] = display_df.apply(
                lambda row: make_threshold_clickable_sw(row, 'Rate_Threshold_Max', 'Rate_Max_Files'), axis=1)

            if 'Is_Valid_Amount' in display_df.columns:
                display_df['Is_Valid_Amount'] = display_df['Is_Valid_Amount'].apply(
                    lambda x: '✅ Valid' if x else '❌ Invalid')
            if 'Is_Valid_Rate' in display_df.columns:
                display_df['Is_Valid_Rate'] = display_df['Is_Valid_Rate'].apply(
                    lambda x: '✅ Valid' if x else '❌ Invalid')
            
        elif mode == 'rate':
            display_df['Rate_Raw'] = display_df['Rate'].copy()
            display_df['Amount_Raw'] = display_df['Amount'].copy()
            display_df['Rate'] = display_df.apply(lambda row: VALIDATOR.format_indian_currency(row['Rate']), axis=1)
            display_df['Amount'] = display_df.apply(lambda row: VALIDATOR.format_indian_currency(row['Amount']), axis=1)
            display_df['Quantity'] = display_df.apply(lambda row: int(round(row['Quantity'])) if VALIDATOR.is_number_based_unit(row['Unit']) else row['Quantity'], axis=1)
        
            if 'Weighted_Avg_Min' in display_df.columns:
                def format_weighted_avg(row, col_name):
                    val = row.get(col_name)
                    if pd.isna(val) or val is None or val == 'NA': return 'NA'
                    return VALIDATOR.format_indian_currency(val)
                
                display_df['Weighted_Avg_Min'] = display_df.apply(lambda row: format_weighted_avg(row, 'Weighted_Avg_Min'), axis=1)
                display_df['Weighted_Avg_Max'] = display_df.apply(lambda row: format_weighted_avg(row, 'Weighted_Avg_Max'), axis=1)
        
            if 'Threshold_Min' in display_df.columns and 'Threshold_Max' in display_df.columns:
                def format_threshold_clickable(row, col_name, files_col_name):
                    val = row.get(col_name)
                    if val == 'NA' or pd.isna(val): return 'NA'
                    display_val = VALIDATOR.format_indian_currency(val)
                    files_data = row.get(files_col_name, [])
                    if not files_data: return str(display_val)
                    tooltip_lines = [f"{i+1}. {f.get('file', 'Unknown')}: {f.get('original_particular', 'N/A')}" for i, f in enumerate(files_data[:5])]
                    if len(files_data) > 5: tooltip_lines.append(f"... and {len(files_data) - 5} more files")
                    tooltip_text = html.escape("\\n".join(tooltip_lines))
                    return f'<span class="threshold-clickable" title="{tooltip_text}" style="cursor: pointer; color: #007bff; text-decoration: underline;">{display_val}</span>'
                
                display_df['Threshold_Min'] = display_df.apply(lambda row: format_threshold_clickable(row, 'Threshold_Min', 'Min_Threshold_Files'), axis=1)
                display_df['Threshold_Max'] = display_df.apply(lambda row: format_threshold_clickable(row, 'Threshold_Max', 'Max_Threshold_Files'), axis=1)
            
        elif mode == 'amount':
            display_df['Amount_Raw'] = display_df['Amount'].copy()
            display_df['Amount'] = display_df.apply(lambda row: VALIDATOR.format_indian_currency(int(round(row['Amount'])) if VALIDATOR.is_number_based_unit(row['Unit']) else row['Amount']), axis=1)

            if 'Weighted_Avg_Min' in display_df.columns:
                def format_weighted_avg(row, col_name):
                    val = row.get(col_name)
                    if pd.isna(val) or val is None or val == 'NA': return 'NA'
                    return VALIDATOR.format_indian_currency(val)

                display_df['Weighted_Avg_Min'] = display_df.apply(lambda row: format_weighted_avg(row, 'Weighted_Avg_Min'), axis=1)
                display_df['Weighted_Avg_Max'] = display_df.apply(lambda row: format_weighted_avg(row, 'Weighted_Avg_Max'), axis=1)

            if 'Threshold_Min' in display_df.columns and 'Threshold_Max' in display_df.columns:
                def format_threshold_clickable(row, col_name, files_col_name):
                    val = row.get(col_name)
                    if val == 'NA' or pd.isna(val): return 'NA'
                    display_val = VALIDATOR.format_indian_currency(val)
                    files_data = row.get(files_col_name, [])
                    if not files_data: return str(display_val)
                    tooltip_lines = [f"{i+1}. {f.get('file', 'Unknown')}: {f.get('original_particular', 'N/A')}" for i, f in enumerate(files_data[:5])]
                    if len(files_data) > 5: tooltip_lines.append(f"... and {len(files_data) - 5} more files")
                    tooltip_text = html.escape("\\n".join(tooltip_lines))
                    return f'<span class="threshold-clickable" title="{tooltip_text}" style="cursor: pointer; color: #007bff; text-decoration: underline;">{display_val}</span>'

                display_df['Threshold_Min'] = display_df.apply(lambda row: format_threshold_clickable(row, 'Threshold_Min', 'Min_Threshold_Files'), axis=1)
                display_df['Threshold_Max'] = display_df.apply(lambda row: format_threshold_clickable(row, 'Threshold_Max', 'Max_Threshold_Files'), axis=1)
        
        else: # quantity mode
            display_df['Quantity'] = display_df.apply(lambda row: int(round(row['Quantity'])) if VALIDATOR.is_number_based_unit(row['Unit']) else row['Quantity'], axis=1)
            
            if 'Weighted_Avg_Min' in display_df.columns:
                def format_weighted_avg(row, col_name):
                    val = row.get(col_name)
                    if pd.isna(val) or val is None or val == 'NA': return 'NA'
                    return int(round(val)) if VALIDATOR.is_number_based_unit(row.get('Unit', '')) else f"{val:.2f}"

                display_df['Weighted_Avg_Min'] = display_df.apply(lambda row: format_weighted_avg(row, 'Weighted_Avg_Min'), axis=1)
                display_df['Weighted_Avg_Max'] = display_df.apply(lambda row: format_weighted_avg(row, 'Weighted_Avg_Max'), axis=1)

            if 'Threshold_Min' in display_df.columns and 'Threshold_Max' in display_df.columns:
                def format_threshold_clickable(row, col_name, files_col_name):
                    val = row.get(col_name)
                    if val == 'NA' or pd.isna(val): return 'NA'
                    display_val = int(round(float(val))) if VALIDATOR.is_number_based_unit(row.get('Unit', '')) else f"{float(val):.2f}"
                    files_data = row.get(files_col_name, [])
                    if not files_data: return str(display_val)
                    tooltip_lines = [f"{i+1}. {f.get('file', 'Unknown')}: {f.get('original_particular', 'N/A')}" for i, f in enumerate(files_data[:5])]
                    if len(files_data) > 5: tooltip_lines.append(f"... and {len(files_data) - 5} more files")
                    tooltip_text = html.escape("\\n".join(tooltip_lines))
                    return f'<span class="threshold-clickable" title="{tooltip_text}" style="cursor: pointer; color: #007bff; text-decoration: underline;">{display_val}</span>'

                display_df['Threshold_Min'] = display_df.apply(lambda row: format_threshold_clickable(row, 'Threshold_Min', 'Min_Threshold_Files'), axis=1)
                display_df['Threshold_Max'] = display_df.apply(lambda row: format_threshold_clickable(row, 'Threshold_Max', 'Max_Threshold_Files'), axis=1)

        if 'Original_Particular' in display_df.columns:
            display_df['Original_Particular'] = display_df['Original_Particular'].apply(lambda x: f'<span title="{html.escape(str(x))}">{html.escape(str(x))}</span>')    

        if mode != 'grouped':
            def create_analysis_button(row):
                if not row['Is_Valid'] and row['Threshold_Min'] != 'NA':
                    safe_particular = html.escape(row["Original_Particular"])
                    value = row.get('Rate_Raw', 0) if mode == 'rate' else row.get('Amount_Raw', 0) if mode == 'amount' else row.get('Quantity', 0)
                    return (f'<button class="btn btn-sm btn-info btn-analysis" data-particular="{safe_particular}" data-quantity="{value}" '
                            f'data-range="{row["Sq_Feet_Range"]}" data-client="{client_name}" data-mode="{mode}">'
                            f'View Analysis</button>')
                return ''
            display_df['Analysis'] = display_df.apply(create_analysis_button, axis=1)
        
        display_df['Is_Valid'] = display_df['Is_Valid'].apply(lambda x: '✅ Valid' if x else '❌ Invalid')
        
        # Dynamic column order based on mode
        if mode == 'grouped':
            column_order = ['Sr.No', 'Particulars', 'As_Per_CWI_Amount', 'Rate_per_sqfeet',
                            'Sq_Feet_Range', 'Is_Valid_Amount', 'Is_Valid_Rate',
                            'CWI_Weighted_Range', 'Rate_Weighted_Range',
                            'Threshold_Min_Amount', 'Threshold_Max_Amount',
                            'Threshold_Min_Rate', 'Threshold_Max_Rate']
        elif mode == 'amount':
            column_order = ['Sr.No', 'Original_Particular', 'Amount', 'Unit', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message', 'Matched_Particular', 'Similarity_Score','Weighted_Avg_Min', 'Weighted_Avg_Max', 'Threshold_Min', 'Threshold_Max', 'Analysis']
        elif mode == 'rate':
            if is_regional:
                column_order = ['Sr.No', 'Original_Particular', 'Rate', 'Unit', 'Quantity', 'Amount', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message_East', 'Validation_Message_West', 'Validation_Message_North', 'Validation_Message_South', 'Matched_Particular', 'Similarity_Score', 'Threshold_Min', 'Threshold_Max', 'Weighted_Avg_Min', 'Weighted_Avg_Max', 'Analysis']
            else:
                if client_name == 'Reliance Retail':
                    column_order = ['Sr.No', 'Service_Code', 'Original_Particular', 'Rate', 'Unit', 'Quantity', 'Amount', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message', 'Matched_Particular', 'Similarity_Score', 'Threshold_Min', 'Threshold_Max', 'Weighted_Avg_Min', 'Weighted_Avg_Max', 'Analysis']
                else:
                    column_order = ['Sr.No', 'Original_Particular', 'Rate', 'Unit', 'Quantity', 'Amount', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message', 'Matched_Particular', 'Similarity_Score', 'Threshold_Min', 'Threshold_Max', 'Weighted_Avg_Min', 'Weighted_Avg_Max', 'Analysis']
        else: # quantity
            if client_name == 'Reliance Retail':
                # ✅ ADD SERVICE CODE AS SECOND COLUMN FOR RELIANCE
                column_order = ['Sr.No', 'Service_Code', 'Original_Particular', 'Quantity', 'Unit', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message', 'Matched_Particular', 'Similarity_Score', 'Threshold_Min', 'Threshold_Max', 'Weighted_Avg_Min', 'Weighted_Avg_Max', 'Analysis']
            else:
                column_order = ['Sr.No', 'Original_Particular', 'Quantity', 'Unit', 'Sq_Feet_Range', 'Is_Valid', 'Validation_Message', 'Matched_Particular', 'Similarity_Score', 'Threshold_Min', 'Threshold_Max', 'Weighted_Avg_Min', 'Weighted_Avg_Max', 'Analysis']
        
        display_df = display_df[column_order]
        
        # >>>>>>>>>> FILTER LOGIC STARTS HERE <<<<<<<<<<
        # Filter out rows where all range columns are N/A
        check_cols = ['Weighted_Avg_Min', 'Weighted_Avg_Max', 'Threshold_Min', 'Threshold_Max']
        existing_check_cols = [c for c in check_cols if c in display_df.columns]
        
        if existing_check_cols:
            def is_row_all_na(row):
                for col in existing_check_cols:
                    val = str(row.get(col, '')).strip().upper()
                    if val not in ['NA', 'N/A', 'NAN', 'NONE', '']:
                        return False 
                return True

           # display_df = display_df[~display_df.apply(is_row_all_na, axis=1)]
        # >>>>>>>>>> FILTER LOGIC ENDS HERE <<<<<<<<<<
        
        # Generate HTML table
        results_html = display_df.to_html(classes='table table-striped table-hover', justify='left', index=False, escape=False)

        # FIX: Add CSS class to selected region column if regional mode
        if is_regional and selected_region:
            region_col_name = f'Validation_Message_{selected_region.capitalize()}'
            results_html = results_html.replace(f'<th>{region_col_name}</th>', f'<th class="selected-region-header">{region_col_name} ✓</th>')

        total = len(results_df)
        valid = results_df['Is_Valid'].sum()
        invalid = total - valid
        accuracy = (valid / total * 100) if total > 0 else 0
        
        print(f"📈 Statistics: Total={total}, Valid={valid}, Invalid={invalid}")
        print(f"✅ Rendering results with mode: {mode}")
        print("="*60 + "\n")
        
        return render_template('results.html', 
                table_html=results_html, 
                full_report_filename=full_report_filename,
                invalid_report_filename=invalid_report_filename,
                total=total, valid=valid, invalid=invalid, 
                accuracy=f"{accuracy:.2f}%",
                client=client_name,
                validation_mode=mode,
                region=selected_region)
                
    except Exception as e:
        print(f"❌ Error in switch_mode: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Error switching mode: {str(e)}')
        return redirect(url_for('upload_file'))
@app.route('/debug_model')
def debug_model():
    """Debug endpoint to inspect model structure"""
    try:
        # Get parameters
        model_type = request.args.get('model', 'icici_qty')  # Default to icici_qty
        
        # Check if model exists
        if model_type not in VALIDATORS:
            available = list(VALIDATORS.keys())
            return f"""
            <h1>❌ Model '{model_type}' not found</h1>
            <p>Available models: {', '.join(available)}</p>
            <p><a href="/debug_model?model={available[0] if available else 'none'}">Try {available[0] if available else 'none'}</a></p>
            """
        
        validator = VALIDATORS[model_type]
        
        output = []
        output.append(f"<h1>🔍 Model Debug: {model_type}</h1>")
        output.append(f"<h2>Available Square Feet Ranges:</h2>")
        output.append("<ul>")
        
        # Iterate through all ranges
        for range_name, range_data in validator.sqfeet_ranges.items():
            output.append(f"<li><b>📏 Range: {range_name}</b>")
            output.append("<ul>")
            
            # Show keys in this range
            output.append(f"<li><b>Keys in this range:</b> {list(range_data.keys())}</li>")
            
            # Check each potential data key
            for key in range_data.keys():
                if isinstance(range_data[key], dict):
                    num_items = len(range_data[key])
                    output.append(f"<li><b>{key}:</b> {num_items} items")
                    
                    # Show first 5 particulars
                    if num_items > 0:
                        output.append("<ul>")
                        for i, (particular, stats) in enumerate(list(range_data[key].items())[:5]):
                            # Show particular name and some stats
                            if isinstance(stats, dict):
                                mean = stats.get('overall_mean', 'N/A')
                                w_min = stats.get('weighted_avg_min', 'N/A')
                                w_max = stats.get('weighted_avg_max', 'N/A')
                                output.append(f"<li>{particular[:80]}...")
                                output.append(f"<ul>")
                                output.append(f"<li>Mean: {mean}</li>")
                                output.append(f"<li>Weighted Range: {w_min} - {w_max}</li>")
                                output.append(f"</ul>")
                                output.append("</li>")
                        
                        if num_items > 5:
                            output.append(f"<li>... and {num_items - 5} more items</li>")
                        output.append("</ul>")
                    else:
                        output.append(" <b style='color:red;'>⚠️ EMPTY!</b>")
                    
                    output.append("</li>")
            
            output.append("</ul>")
            output.append("</li>")
        
        output.append("</ul>")
        
        # Add links to other models
        output.append("<hr>")
        output.append("<h3>🔗 Debug Other Models:</h3>")
        output.append("<ul>")
        for model in VALIDATORS.keys():
            output.append(f"<li><a href='/debug_model?model={model}'>{model}</a></li>")
        output.append("</ul>")
        
        return "<br>".join(output)
        
    except Exception as e:
        import traceback
        return f"""
        <h1>❌ Error:</h1>
        <pre>{str(e)}</pre>
        <h2>Traceback:</h2>
        <pre>{traceback.format_exc()}</pre>
        """
        

if __name__ == '__main__':
    app.run(debug=True)


        
