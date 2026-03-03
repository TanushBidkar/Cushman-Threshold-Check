import os
import re
import pandas as pd
from openpyxl import load_workbook
import gc
import shutil

# ─── CONFIG ───────────────────────────────────────────────────────────────────
INVALID_FOLDER = r"C:\Users\Tanush.Bidkar\Downloads\Invalid_Sheets\Invalid_Sheets"
ALL_SHEETS_FOLDER = r"C:\Users\Tanush.Bidkar\Downloads\All Sheets\All Sheets"

TARGET_SHEET_KEYWORDS = [
    'boq',
    'consolidated boq',
    'bill',
    'work done as per contract',
    'measurement sheet as per cwi',
    'cwi working',
    'post audit',
    'cwi-working',
    'extracted data',
    'main boq',
]
# ──────────────────────────────────────────────────────────────────────────────


def extract_number_from_filename(filename):
    numbers = re.findall(r'\d+', filename)
    return int(numbers[0]) if numbers else None


def is_target_sheet(sheet_name):
    sheet_lower = sheet_name.lower().strip()
    for keyword in TARGET_SHEET_KEYWORDS:
        if keyword in sheet_lower:
            return True
    return False


def clean_text(text):
    if pd.isna(text) or text is None:
        return ""
    return re.sub(r'\s+', ' ', str(text).lower().strip())


def get_invalid_particulars(invalid_file_path):
    try:
        df = pd.read_excel(invalid_file_path, sheet_name='Sheet1', dtype=str)

        particular_col = None
        isvalid_col = None

        for col in df.columns:
            col_lower = str(col).lower().strip()
            if 'original_particular' in col_lower or ('original' in col_lower and 'particular' in col_lower):
                particular_col = col
            if 'is_valid' in col_lower:
                isvalid_col = col

        if particular_col is None or isvalid_col is None:
            print(f"  ⚠  Could not find required columns in {os.path.basename(invalid_file_path)}")
            print(f"     Columns found: {list(df.columns)}")
            return set()

        invalid_mask = df[isvalid_col].str.strip().str.upper() == 'FALSE'
        invalid_particulars = df.loc[invalid_mask, particular_col].dropna()

        cleaned = set(clean_text(p) for p in invalid_particulars if clean_text(p))
        print(f"  → Found {len(cleaned)} invalid particulars")
        return cleaned

    except Exception as e:
        print(f"  ✗ Error reading {os.path.basename(invalid_file_path)}: {e}")
        return set()


def particulars_match(cell_val, invalid_particulars):
    """Check if a cell value matches any invalid particular."""
    cleaned_val = clean_text(cell_val)
    if not cleaned_val or len(cleaned_val) < 5:
        return False

    for inv in invalid_particulars:
        if not inv or len(inv) < 5:
            continue
        min_len = min(len(cleaned_val), len(inv))
        if min_len >= 20:
            if inv in cleaned_val or cleaned_val in inv:
                return True
        else:
            if cleaned_val == inv:
                return True
    return False


def find_particular_col_index_in_sheet(ws):
    """
    Scan the first 10 rows of the openpyxl worksheet to find which column
    contains 'particulars' or 'description' or 'description of work'.
    Returns 1-indexed column number, or None.
    
    Also returns the header row number (1-indexed).
    """
    strong_keywords = ['particular', 'description of work', 'description', 'item', 'work', 'scope', 'activity']
    avoid_keywords = ['location', 'metro', 'city', 'zone', 'region', 'rate', 'amount',
                      'qty', 'quantity', 'unit', 'total', 'sr.no', 'sr no', 'sno',
                      'remarks', 'status', 'date', 'uom', 'vendor', 'saving', 'excess',
                      'type', 'nature', 'category', 'sub category', 'make', 'model',
                      'green', 'test', 'vender', 'cwi qty', 'wo qty', 'act. qty',
                      'current_rate', 'sub_category']

    header_row = None
    best_col = None
    best_score = -1

    # Scan first 10 rows to find header
    for row_idx in range(1, 11):
        row_cells = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        non_empty = [c for c in row_cells if c is not None and str(c).strip() != '']
        if len(non_empty) < 2:
            continue

        # Check each cell in this row as potential header
        for col_idx, cell_val in enumerate(row_cells, 1):
            if cell_val is None:
                continue
            cell_lower = str(cell_val).lower().strip()

            # Check avoid list
            is_avoid = any(av in cell_lower for av in avoid_keywords)
            if is_avoid:
                continue

            # Score based on strong keywords
            score = 0
            for kw in strong_keywords:
                if kw in cell_lower:
                    score += (10 if kw in ['particular', 'description of work'] else 5)
                    break

            if score > best_score:
                best_score = score
                best_col = col_idx
                header_row = row_idx

    return best_col, header_row


def delete_rows_openpyxl_safe(ws, rows_to_delete_set):
    """
    Delete rows by rebuilding cell data — avoids MemoryError from ws.delete_rows
    on large files with complex structures.
    
    Strategy: copy all rows EXCEPT the ones to delete into a temp structure,
    then write back. Preserves relative structure but NOTE: this cannot
    preserve full formatting/merged cells perfectly for very complex sheets.
    
    For structure-safe deletion on complex sheets, we use a different approach:
    clear the row content instead of deleting, to preserve formatting.
    """
    # Sort rows descending so we delete from bottom up
    for row_num in sorted(rows_to_delete_set, reverse=True):
        try:
            ws.delete_rows(row_num, 1)
        except MemoryError:
            # If MemoryError, just clear the cell values instead of deleting
            print(f"         ⚠  MemoryError on row {row_num} — clearing cells instead of deleting")
            for cell in ws[row_num]:
                cell.value = None
        except Exception as e:
            print(f"         ⚠  Error deleting row {row_num}: {e} — clearing instead")
            try:
                for cell in ws[row_num]:
                    cell.value = None
            except:
                pass


def process_sheet_openpyxl(wb, file_path, sheet_name, invalid_particulars):
    """
    Find and delete matching rows in a sheet using openpyxl directly.
    Preserves original structure/formatting.
    """
    ws = wb[sheet_name]

    # Find the particulars column and header row
    particular_col_idx, header_row_num = find_particular_col_index_in_sheet(ws)

    if particular_col_idx is None:
        print(f"       ⚠  Cannot find particulars column. Trying all text columns...")
        # Fallback: scan ALL columns for matching text
        particular_col_idx = None
        header_row_num = 1

    print(f"       Particulars column index: {particular_col_idx}, Header row: {header_row_num}")

    # Scan all data rows to find matches
    rows_to_delete = set()
    max_row = ws.max_row
    data_start_row = (header_row_num + 1) if header_row_num else 1

    for row_num in range(data_start_row, max_row + 1):
        if particular_col_idx:
            # Check specific column
            cell = ws.cell(row=row_num, column=particular_col_idx)
            cell_val = cell.value
            if cell_val and particulars_match(str(cell_val), invalid_particulars):
                rows_to_delete.add(row_num)
                print(f"       🗑  Row {row_num}: {str(cell_val)[:80]}...")
        else:
            # Fallback: scan entire row for any matching text
            row_matched = False
            for col in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row_num, column=col).value
                if cell_val and isinstance(cell_val, str) and len(str(cell_val).strip()) > 15:
                    if particulars_match(str(cell_val), invalid_particulars):
                        rows_to_delete.add(row_num)
                        print(f"       🗑  Row {row_num} (col {col}): {str(cell_val)[:80]}...")
                        row_matched = True
                        break
                if row_matched:
                    break

    if not rows_to_delete:
        print(f"       ℹ  No matching rows found in sheet '{sheet_name}'")
        return 0

    print(f"       Deleting {len(rows_to_delete)} rows...")
    delete_rows_openpyxl_safe(ws, rows_to_delete)
    return len(rows_to_delete)


def delete_matching_rows_from_file(all_sheet_file_path, invalid_particulars):
    """Process one All Sheets file — open once, process all target sheets, save once."""
    total_deleted = 0

    # Get sheet names first using pandas (lightweight)
    try:
        xl = pd.ExcelFile(all_sheet_file_path)
        all_sheet_names = xl.sheet_names
        xl.close()
        del xl
    except Exception as e:
        print(f"    ✗ Cannot read sheet names: {e}")
        return 0

    matched_sheets = [s for s in all_sheet_names if is_target_sheet(s)]

    if not matched_sheets:
        print(f"    ⚠  No target sheets found. Available: {all_sheet_names}")
        return 0

    # Load workbook ONCE with keep_vba=False, data_only=True for speed
    try:
        print(f"    📂 Loading workbook (preserving structure)...")
        wb = load_workbook(all_sheet_file_path, keep_vba=False, data_only=False)
    except Exception as e:
        print(f"    ✗ Cannot load workbook: {e}")
        return 0

    modified = False

    for sheet_name in matched_sheets:
        print(f"    📄 Processing sheet: '{sheet_name}'")
        try:
            deleted = process_sheet_openpyxl(wb, all_sheet_file_path, sheet_name, invalid_particulars)
            total_deleted += deleted
            if deleted > 0:
                modified = True
        except MemoryError:
            print(f"    ✗ MemoryError on sheet '{sheet_name}' — skipping")
            gc.collect()
        except Exception as e:
            print(f"    ✗ Error on sheet '{sheet_name}': {e}")
            import traceback
            traceback.print_exc()

    if modified:
        try:
            print(f"    💾 Saving {os.path.basename(all_sheet_file_path)}...")
            
            # Strategy: save to a safe temp filename in same directory, then replace original
            folder = os.path.dirname(all_sheet_file_path)
            temp_path = os.path.join(folder, f"_TEMP_SAVE_{os.getpid()}.xlsx")
            
            wb.save(temp_path)
            wb.close()
            del wb
            gc.collect()
            
            # Replace original with temp
            if os.path.exists(temp_path):
                # Remove original first
                if os.path.exists(all_sheet_file_path):
                    os.remove(all_sheet_file_path)
                os.rename(temp_path, all_sheet_file_path)
                print(f"    ✅ Saved successfully")
            else:
                print(f"    ✗ Temp file not created — save may have failed silently")
                
        except Exception as e:
            print(f"    ✗ Error saving: {e}")
            # Cleanup temp if it exists
            try:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            except:
                pass
            # Try saving to Downloads folder as fallback
            try:
                fallback_name = re.sub(r'[^\w\s\-\.]', '_', os.path.basename(all_sheet_file_path))
                fallback_path = os.path.join(
                    r"C:\Users\Tanush.Bidkar\Downloads", 
                    f"FIXED_{fallback_name}"
                )
                wb.save(fallback_path)
                print(f"    ⚠  Saved to fallback location: {fallback_path}")
                print(f"    ⚠  Please manually replace the original file with this one")
            except Exception as e2:
                print(f"    ✗ Fallback save also failed: {e2}")
            finally:
                try:
                    wb.close()
                    del wb
                    gc.collect()
                except:
                    pass
            return total_deleted
    else:
        wb.close()
        del wb
        gc.collect()

    return total_deleted


def build_number_to_file_map(folder_path):
    mapping = {}
    for fname in os.listdir(folder_path):
        if not fname.endswith(('.xlsx', '.xls')):
            continue
        num = extract_number_from_filename(fname)
        if num is not None:
            if num not in mapping:
                mapping[num] = []
            mapping[num].append(os.path.join(folder_path, fname))
    return mapping


def main():
    print("=" * 70)
    print("INVALID ROW DELETION — STRUCTURE PRESERVING VERSION")
    print("=" * 70)

    print("\n📂 Scanning Invalid Sheets folder...")
    invalid_map = build_number_to_file_map(INVALID_FOLDER)
    print(f"   Found {len(invalid_map)} unique sqft numbers in Invalid Sheets")

    print("\n📂 Scanning All Sheets folder...")
    allsheets_map = build_number_to_file_map(ALL_SHEETS_FOLDER)
    print(f"   Found {len(allsheets_map)} unique sqft numbers in All Sheets")

    grand_total_deleted = 0
    files_modified = 0
    unmatched_invalid = []

    print("\n" + "=" * 70)
    print("PROCESSING...")
    print("=" * 70)

    for sqft_num in sorted(invalid_map.keys()):
        invalid_files = invalid_map[sqft_num]

        if sqft_num not in allsheets_map:
            print(f"\n⚠  sqft={sqft_num}: No matching All Sheets file — SKIPPING")
            unmatched_invalid.extend(invalid_files)
            continue

        all_sheet_files = allsheets_map[sqft_num]

        for invalid_file in invalid_files:
            print(f"\n{'─'*70}")
            print(f"🔍 sqft={sqft_num} | Invalid: {os.path.basename(invalid_file)}")

            invalid_particulars = get_invalid_particulars(invalid_file)

            if not invalid_particulars:
                print(f"   ℹ  No invalid particulars — skipping")
                continue

            for all_file in all_sheet_files:
                print(f"\n   🗂  All Sheets: {os.path.basename(all_file)}")
                deleted = delete_matching_rows_from_file(all_file, invalid_particulars)
                grand_total_deleted += deleted
                if deleted > 0:
                    files_modified += 1

            gc.collect()

    print("\n" + "=" * 70)
    print("SUMMARY")
    print("=" * 70)
    print(f"✅ Total rows deleted : {grand_total_deleted}")
    print(f"✅ Files modified     : {files_modified}")

    if unmatched_invalid:
        print(f"\n⚠  {len(unmatched_invalid)} Invalid file(s) with no matching All Sheets file:")
        for f in unmatched_invalid:
            print(f"   - {os.path.basename(f)}")

    print("\nDone!")


if __name__ == "__main__":
    main()