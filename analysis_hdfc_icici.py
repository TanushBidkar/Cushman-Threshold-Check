import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

# Define source directories and output path
hdfc_dir = r'C:\Users\Tanush.Bidkar\Downloads\training_files 2\training_files 2\training_files\1000-1250'
icici_dir = r'C:\Users\Tanush.Bidkar\Downloads\Zone Wise Grouped ICICI\1000-1250'
output_path = r'C:\Users\Tanush.Bidkar\Downloads\Analysis ICICI VS HDFC.xlsx'

def consolidate_folder(folder_path):
    files = [f for f in os.listdir(folder_path) if f.endswith(('.xlsx', '.xls'))]
    master_df = None
    branch_names = []

    for file in files:
        file_path = os.path.join(folder_path, file)
        branch_name = os.path.splitext(file)[0]
        branch_names.append(branch_name)
        
        try:
            # Load the Summary sheet
            df = pd.read_excel(file_path, sheet_name='Summary Sheet')
            
            # Clean column names (strip spaces)
            df.columns = [str(c).strip() for c in df.columns]
            
            # Extract necessary columns
            # We assume columns are: 'Sr. No', 'Particulars', 'Amount As per CWI', 'Rate per sq.feet'
            # Rename value columns to include branch name for merging
            temp_df = df[['Sr. No', 'Particulars', 'Amount As per CWI', 'Rate per sq.feet']].copy()
            amt_col = f"{branch_name}_Amt"
            rate_col = f"{branch_name}_Rate"
            temp_df.rename(columns={
                'Amount As per CWI': amt_col,
                'Rate per sq.feet': rate_col
            }, inplace=True)

            if master_df is None:
                master_df = temp_df
            else:
                # Merge on Particulars to avoid repetition. 
                # Sr. No is kept from the first file or updated.
                master_df = pd.merge(master_df, temp_df[['Particulars', amt_col, rate_col]], 
                                     on='Particulars', how='outer')
        except Exception as e:
            print(f"Skipping {file} due to error: {e}")

    return master_df, branch_names

def write_custom_excel(writer, df, branch_names, sheet_name):
    # Create sheet
    workbook = writer.book
    if sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
    else:
        ws = workbook.create_sheet(sheet_name)
    
    # Row 1: File Names
    # Starting from Column C (Index 3), each file gets 2 columns
    ws.cell(row=1, column=1, value="Sr. No").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Particulars").font = Font(bold=True)
    
    col_idx = 3
    for name in branch_names:
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx+1)
        col_idx += 2

    # Row 3: Column Headers
    ws.cell(row=3, column=1, value="Sr. No").font = Font(bold=True)
    ws.cell(row=3, column=2, value="Particulars").font = Font(bold=True)
    
    col_idx = 3
    for _ in branch_names:
        ws.cell(row=3, column=col_idx, value="Amount As per CWI").font = Font(bold=True)
        ws.cell(row=3, column=col_idx+1, value="Rate per sq.feet").font = Font(bold=True)
        col_idx += 2

    # Data starting from Row 4
    # Reorder df to ensure Total is at the bottom if it exists
    if 'Particulars' in df.columns:
        total_mask = df['Particulars'].str.contains('Total', case=False, na=False)
        df_no_total = df[~total_mask]
        df_total = df[total_mask]
        df = pd.concat([df_no_total, df_total])

    for r_idx, row in enumerate(df.values, start=4):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

# Process Folders
hdfc_master, hdfc_branches = consolidate_folder(hdfc_dir)
icici_master, icici_branches = consolidate_folder(icici_dir)

# Save to Excel
with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    if hdfc_master is not None:
        write_custom_excel(writer, hdfc_master, hdfc_branches, 'HDFC')
    if icici_master is not None:
        write_custom_excel(writer, icici_master, icici_branches, 'ICICI')

print(f"Analysis file created at: {output_path}")