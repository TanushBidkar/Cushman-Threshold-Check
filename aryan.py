import pandas as pd
import os
import re
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --- Configuration ---
data_dir = r"C:\Users\Tanush.Bidkar\Downloads\Zone Wise Grouped ICICI"
output_file = r"C:\Users\Tanush.Bidkar\Downloads\Final_Rate_Linked_ICICI.xlsx"

# Canonical Particulars
PARTICULARS_MAP = {
    "Civil & related works": (1, "I"),
    "POP and false ceiling work": (2, "II"),
    "CARPENTRY AND INTERIOR WORKS": (3, "III"),
    "Painting works": (4, "IV"),
    "ROLLING SHUTTER AND MS WORK": (5, "V"),
    "ELECTRIFICATION AND ALLIED WORKS": (6, "VI"),
    "Additional Work": (7, "VII"),
    "Others/Uncategorized": (8, "VIII"),
    "Total": (9, "") 
}

LOWER_MAP = {k.lower(): v for k, v in PARTICULARS_MAP.items()}

# Visual Styling
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
cell_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center', indent=1)

def get_area_bin(filename):
    match = re.match(r"(\d+)", filename)
    if not match: return "Others"
    sqft = int(match.group(1))
    if sqft <= 750: return "500-750sqft"
    start = 751 + ((sqft - 751) // 250) * 250
    return f"{start}-{start+249}sqft" if sqft <= 5000 else "5000+sqft"

def process_data():
    binned_data = {}
    files = [f for f in os.listdir(data_dir) if f.endswith('.xlsx') and not f.startswith('~')]
    
    for filename in files:
        bin_name = get_area_bin(filename)
        if bin_name not in binned_data: binned_data[bin_name] = []
        try:
            df = pd.read_excel(os.path.join(data_dir, filename), sheet_name='Summary Sheet')
            df.columns = [c.strip() for c in df.columns]
            df['match_key'] = df['Particulars'].astype(str).str.strip().str.lower()
            binned_data[bin_name].append({'title': filename.replace('.xlsx', ''), 'df': df})
        except Exception as e:
            print(f"Skipping {filename}: {e}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_locations = {}

    # 1. Create Detailed Area Sheets (Same structure as before)
    for bin_name in sorted(binned_data.keys()):
        ws = wb.create_sheet(bin_name)
        ws.cell(row=2, column=1, value="Sr. No").fill = header_fill
        ws.cell(row=2, column=2, value="Particulars").fill = header_fill
        for c in [1, 2]: ws.cell(row=2, column=c).font = header_font

        for part, (idx, roman) in PARTICULARS_MAP.items():
            r = idx + 2
            ws.cell(row=r, column=1, value=roman).border = cell_border
            ws.cell(row=r, column=2, value=part).border = cell_border
            ws.cell(row=r, column=2).alignment = left_align
            if part == "Total": ws.cell(row=r, column=2).font = Font(bold=True)

        curr_col = 3
        for data in binned_data[bin_name]:
            ws.merge_cells(start_row=1, start_column=curr_col, end_row=1, end_column=curr_col+1)
            h_cell = ws.cell(row=1, column=curr_col, value=data['title'])
            h_cell.fill, h_cell.font, h_cell.alignment = header_fill, header_font, center_align

            ws.cell(row=2, column=curr_col, value="Amount As per CWI").fill = header_fill
            ws.cell(row=2, column=curr_col+1, value="Rate per sq.feet").fill = header_fill
            for c in [curr_col, curr_col+1]: ws.cell(row=2, column=c).font = header_font

            for r_idx in range(3, 11):
                for c_idx in [curr_col, curr_col+1]:
                    ws.cell(row=r_idx, column=c_idx, value=0).border = cell_border
                    ws.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'

            for _, row in data['df'].iterrows():
                key = str(row['match_key'])
                if key in LOWER_MAP:
                    r_idx = LOWER_MAP[key][0] + 2
                    ws.cell(row=r_idx, column=curr_col, value=row.get('Amount As per CWI', 0))
                    ws.cell(row=r_idx, column=curr_col+1, value=row.get('Rate per sq.feet', 0))

            # Track Locations: amt_col is the Amount, rate_col is the Rate
            sheet_locations[data['title']] = {
                'sheet': bin_name, 
                'rate_col': get_column_letter(curr_col + 1) # This is the key change
            }
            ws.column_dimensions[get_column_letter(curr_col)].width = 25
            ws.column_dimensions[get_column_letter(curr_col+1)].width = 25
            ws.column_dimensions[get_column_letter(curr_col+2)].width = 4 
            curr_col += 3

        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        for r in range(1, 12): ws.row_dimensions[r].height = 25

    # 2. Create Master Summary for RATES
    ms = wb.create_sheet("Master Summary", 0)
    ms_static_headers = ["Sr. No", "File / Branch Name", "Area Group"]
    category_list = list(PARTICULARS_MAP.keys())
    
    col_ptr = 1
    for h in ms_static_headers:
        c = ms.cell(row=1, column=col_ptr, value=h)
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center_align, cell_border
        col_ptr += 1
    
    for cat in category_list:
        # Header now clearly indicates Rate
        c = ms.cell(row=1, column=col_ptr, value=f"{cat} (RATE)")
        c.font, c.fill, c.alignment, c.border = header_font, header_fill, center_align, cell_border
        ms.column_dimensions[get_column_letter(col_ptr)].width = 30
        col_ptr += 1

    # Write Data with Formulas linked to the RATE column
    for i, (title, loc) in enumerate(sheet_locations.items(), 2):
        ms.cell(row=i, column=1, value=i-1).border = cell_border
        ms.cell(row=i, column=2, value=title).border = cell_border
        ms.cell(row=i, column=3, value=loc['sheet']).border = cell_border
        
        for j, cat in enumerate(category_list):
            row_in_sheet = PARTICULARS_MAP[cat][0] + 2
            # Formula links to loc['rate_col'] instead of amt_col
            formula = f"='{loc['sheet']}'!{loc['rate_col']}{row_in_sheet}"
            cell = ms.cell(row=i, column=4 + j, value=formula)
            cell.number_format = '#,##0.00'
            cell.border = cell_border

    ms.column_dimensions['B'].width = 50
    ms.column_dimensions['C'].width = 20
    
    wb.save(output_file)
    print(f"Rate-based Master Summary created: {output_file}")

process_data()